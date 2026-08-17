#!/usr/bin/env python3
"""
batch_analyze_multi_v3.py

Multi-solver analyzer that:
1) Produces *per-solver* Excel sheets in the "old" format:
   - CoverageStats_GTvs<LABEL>
   - DistanceStats_GTvs<LABEL>

   Columns (same as your previous batch_analyze):
     patch_key, mask_type, coverage_bin/distance_bin_m, n_pixels,
     mean_signed, median_signed, mean_abs, std_abs, median_abs,
     p90_abs, p99_abs, linf_abs,
     l1_rae_sum, l1_abs_mmph_sum, l1_abs_mmph_sum_norm_hw

   Conventions:
   - Rainy pixels: signed_rel = (GT - PRED)/GT, abs_rel = |GT-PRED|/GT
   - Non-rainy pixels: signed_diff = (PRED - GT), abs_diff = |PRED-GT|
   l1_rae_sum = sum(|GT-PRED|/GT), using denom=1 when GT=0.
   l1_abs_mmph_sum = sum(|GT-PRED|).
   l1_abs_mmph_sum_norm_hw = l1_abs_mmph_sum / (H*W) for the patch.

2) Adds per-solver link-based sheets:
   - LinkStats_GTvs<LABEL>
     patch_key, n_links_valid, attn_l1_all, J_atten_all, attn_l1_ge10km, J_atten_ge10km,
     max_abs_diff, p95_abs_diff, p99_abs_diff

   Where:
     attn_l1 = sum_{links} |A_hat - A_obs|
     J_atten = (1 / #links) * sum_{links} ((A_hat - A_obs)^2 / L_km)
   and the *_ge10km versions restrict to links with L_km >= 10.

3) Keeps the nice final plots (IQR of per-patch medians by distance bin) across all solvers.

Relative paths in the YAML are resolved relative to the YAML file location.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import json
import math
import os
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np

from cml_attenuation.analysis_metrics import (
    absolute_difference_summary as abs_diff_summary,
    attenuation_error_per_km as abs_attn_error_per_km_metrics,
    attenuation_l1_and_legacy_j1 as attn_l1_and_J1,
    distribution_stats as stats_row,
    pixel_errors as compute_pixel_errors,
)
from cml_attenuation.config_io import deep_get, load_config_file, resolve_path
from cml_attenuation.pipeline_validation import validate_analysis_config

DEFAULT_REFERENCE_W_SMOOTH = 1.0
DEFAULT_REFERENCE_W_SHRINK = 1.0
DEFAULT_REFERENCE_W_SECOND_DER = 1.0
DEFAULT_REFERENCE_EPS = 0.01


# ----------------------------
# Config utilities
# ----------------------------

def _module_basename(module_name: str) -> str:
    return str(module_name or "").strip().rsplit(".", 1)[-1]


def objective_scaling_from_module(module_name: str) -> str:
    m = _module_basename(module_name)
    normalized_modules = {
        "solve_rain_lbfgsb_normalized_ildw_multipliers",
        "solve_rain_lbfgsb_normalized_ildw_multipliers",
        "solve_rain_lbfgsb_normalized_ildw_multipliers_virtual_convex",
        "solve_rain_lbfgsb_normalized_ildw_multipliers_virtual_homotopy",
        # backward-compatible names
        "solve_rain_lbfgsb_j1norm_j2j3j4",
        "solve_rain_lbfgsb_j1norm_j2j3lin_j4",
    }
    unnormalized_modules = {
        "solve_rain_lbfgsb",
        "solve_rain_lbfgsb_j2",
        "solve_rain_lbfgsb_j3",
        "solve_rain_lbfgsb_j2_j3",
    }
    baseline_modules = {"idw_baseline", "ildw_baseline"}
    if m in normalized_modules:
        return "NORMALIZED"
    if m in unnormalized_modules:
        return "UNNORMALIZED"
    if m in baseline_modules:
        return "N/A_BASELINE"
    return "UNKNOWN"


def objective_scaling_from_meta(meta: Dict[str, float]) -> str:
    if (
        "meta_alphas_1d" in meta
        or "meta_alphas_2d" in meta
        or "meta_alphas_total" in meta
        or "meta_alpha_1d" in meta
        or "meta_alpha_2d" in meta
        or "meta_alpha_total" in meta
    ):
        return "CONSTRAINED_NORMALIZED"
    if "meta_j2_w" in meta or "meta_j3_w" in meta or "meta_j4_w" in meta:
        return "NORMALIZED"
    if "meta_lambda" in meta or "meta_mu" in meta or "meta_eta" in meta:
        return "UNNORMALIZED"
    return "N/A_BASELINE" if len(meta) == 0 else "UNKNOWN"


def solver_objective_formula_text(*, scaling: str, meta: Dict[str, float]) -> str:
    if scaling == "N/A_BASELINE":
        return "N/A (baseline interpolation; no optimization objective)."
    if scaling == "NORMALIZED":
        return (
            "J_weighted_sum = J_atten + w_1d*J_1d + w_total*J_total + w_2d*J_2d; "
            "J_atten is normalized by #valid_links; J_1d, J_total, J_2d are normalized by #pixels."
        )
    if scaling == "CONSTRAINED_NORMALIZED":
        has_alpha_atten = "meta_alpha_atten" in meta
        if has_alpha_atten:
            return (
                "J_weighted_sum = w_atten*J_atten + w_1d*J_1d + w_total*J_total + w_2d*J_2d; "
                "for ILDW-multiplier solvers these w_* are per-instance alpha_*."
            )
        return (
            "J_weighted_sum = w_1d*J_1d + w_2d*J_2d + w_total*J_total, "
            "with attenuation fit J_atten handled as a constraint."
        )
    if scaling == "UNNORMALIZED":
        return (
            "J_weighted_sum = J_atten + w_1d*J_1d + w_total*J_total + w_2d*J_2d "
            "(some solver variants may omit terms). Terms are raw sums (not normalized)."
        )
    return "Unknown objective form (insufficient metadata)."


def objective_term_presence_from_module(module_name: str, *, solver_label: str = "") -> Dict[str, bool]:
    """
    Infer which objective terms are part of the solver's native objective.
    Keys: J1, J2, J3, J4
    """
    m = _module_basename(module_name)
    s = str(solver_label or "").strip().upper()
    if s == "GT":
        return {"J1": True, "J2": True, "J3": True, "J4": True}
    if m in {"idw_baseline", "ildw_baseline"}:
        return {"J1": False, "J2": False, "J3": False, "J4": False}
    if m in {"solve_rain_lbfgsb_j2"}:
        return {"J1": True, "J2": True, "J3": False, "J4": False}
    if m in {"solve_rain_lbfgsb_j3"}:
        return {"J1": True, "J2": False, "J3": True, "J4": False}
    if m in {"solve_rain_lbfgsb_j2_j3"}:
        return {"J1": True, "J2": True, "J3": True, "J4": False}
    if m in {
        "solve_rain_lbfgsb",
        "solve_rain_lbfgsb_normalized_ildw_multipliers",
        "solve_rain_lbfgsb_normalized_ildw_multipliers",
        "solve_rain_lbfgsb_normalized_ildw_multipliers_virtual_convex",
        "solve_rain_lbfgsb_normalized_ildw_multipliers_virtual_homotopy",
        "solve_rain_lbfgsb_j1norm_j2j3j4",
        "solve_rain_lbfgsb_j1norm_j2j3lin_j4",
    }:
        return {"J1": True, "J2": True, "J3": True, "J4": True}
    # Unknown module: keep terms available to avoid dropping potentially valid data.
    return {"J1": True, "J2": True, "J3": True, "J4": True}


def summarize_solver_settings(meta: Dict[str, float]) -> str:
    if not meta:
        return "No meta_* settings found in solution npz."
    has_constrained = any(
        k in meta
        for k in (
            "meta_alphas_1d",
            "meta_alphas_2d",
            "meta_alphas_total",
            "meta_alpha_1d",
            "meta_alpha_2d",
            "meta_alpha_total",
        )
    )
    has_norm = any(k in meta for k in ("meta_j2_w", "meta_j3_w", "meta_j4_w"))
    has_unnorm = any(k in meta for k in ("meta_lambda", "meta_mu", "meta_eta"))

    parts: List[str] = []
    if has_constrained:
        parts.append(f"w_atten={float(meta.get('meta_alpha_atten', 1.0))}")
        parts.append(f"w_1d={float(meta.get('meta_alpha_1d', meta.get('meta_alphas_1d', 0.0)))}")
        parts.append(f"w_total={float(meta.get('meta_alpha_total', meta.get('meta_alphas_total', 0.0)))}")
        parts.append(f"w_2d={float(meta.get('meta_alpha_2d', meta.get('meta_alphas_2d', 0.0)))}")
    elif has_norm:
        parts.append(f"w_1d={float(meta.get('meta_j2_w', 0.0))}")
        parts.append(f"w_total={float(meta.get('meta_j3_w', 0.0))}")
        parts.append(f"w_2d={float(meta.get('meta_j4_w', 0.0))}")
    elif has_unnorm:
        parts.append(f"w_1d={float(meta.get('meta_lambda', 0.0))}")
        parts.append(f"w_total={float(meta.get('meta_mu', 0.0))}")
        parts.append(f"w_2d={float(meta.get('meta_eta', 0.0))}")

    if "meta_eps" in meta and not bool(meta.get("meta_use_linear_j3", False)):
        parts.append(f"eps={float(meta['meta_eps'])}")
    rain_init_mode = str(meta.get("meta_rain_init_mode", "")) if "meta_rain_init_mode" in meta else ""
    if rain_init_mode:
        parts.append(f"rain_init.mode={rain_init_mode}")
    if "meta_rain_init_value" in meta and rain_init_mode in {"", "fixed"}:
        parts.append(f"rain_init.value={float(meta['meta_rain_init_value'])}")
    if "meta_rain_init_multiplier" in meta and rain_init_mode == "idw_mean":
        parts.append(f"rain_init.multiplier={float(meta['meta_rain_init_multiplier'])}")
    if "meta_R0" in meta:
        parts.append(f"R0={float(meta['meta_R0'])}")
    if "meta_R0_from_IDW" in meta:
        parts.append(f"R0_from_IDW={bool(meta['meta_R0_from_IDW'])}")
    if "meta_R0_from_ILDW" in meta:
        parts.append(f"R0_from_ILDW={bool(meta['meta_R0_from_ILDW'])}")
    if "meta_ftol" in meta:
        parts.append(f"ftol={float(meta['meta_ftol'])}")
    if "meta_gtol" in meta:
        parts.append(f"gtol={float(meta['meta_gtol'])}")
    if "meta_maxiter" in meta:
        parts.append(f"maxiter={int(meta['meta_maxiter'])}")
    if "meta_maxls" in meta:
        parts.append(f"maxls={int(meta['meta_maxls'])}")

    if not parts:
        return "meta_* found, but no recognized optimization setting keys."
    return "; ".join(parts)


def normalize_solvers(obj: Any, *, path_for_err: str) -> List[dict]:
    """
    Accept either:
      solvers: [ {..}, {..} ]
    or
      solvers:
        keyA: {..}
        keyB: {..}
    Returns list of solver dicts.
    """
    if obj is None:
        raise SystemExit(f"Config missing {path_for_err}.")
    if isinstance(obj, list):
        if not obj:
            raise SystemExit(f"{path_for_err} must be a non-empty list.")
        if not all(isinstance(x, dict) for x in obj):
            raise SystemExit(f"{path_for_err} list elements must be dicts.")
        return obj
    if isinstance(obj, dict):
        vals = list(obj.values())
        if not vals:
            raise SystemExit(f"{path_for_err} must be a non-empty mapping.")
        if not all(isinstance(x, dict) for x in vals):
            raise SystemExit(f"{path_for_err} mapping values must be dicts.")
        return vals
    raise SystemExit(f"{path_for_err} must be a list or mapping.")


def progress_iter(items: Iterable[Any], *, total: Optional[int] = None, desc: str = "") -> Iterable[Any]:
    try:
        from tqdm import tqdm  # type: ignore
        return tqdm(items, total=total, desc=desc)
    except Exception:
        def gen():
            count = 0
            for count, item in enumerate(items, 1):
                if total is not None:
                    msg = f"{desc} {count}/{total}"
                else:
                    msg = f"{desc} {count}"
                if count == 1 or count % 5 == 0 or (total is not None and count == total):
                    sys.stdout.write("\r" + msg)
                    sys.stdout.flush()
                yield item
            if count > 0:
                sys.stdout.write("\r" + msg + "\n")
                sys.stdout.flush()
        return gen()


# ----------------------------
# IO utilities
# ----------------------------

def load_npz_first_key(path: Path, key_preference: Sequence[str]) -> np.ndarray:
    with np.load(path, allow_pickle=False) as z:
        for k in key_preference:
            if k in z:
                return np.asarray(z[k])
        # fallback: first ndarray-like key
        for k in z.files:
            if isinstance(z[k], np.ndarray):
                return np.asarray(z[k])
        raise KeyError(f"None of keys {list(key_preference)} found in {path.name}; keys={z.files}")


def list_npz(dirp: Path, prefix: str) -> List[Path]:
    return sorted(dirp.glob(f"{prefix}_*.npz"))


def list_json(dirp: Path, prefix: str) -> List[Path]:
    return sorted(dirp.glob(f"{prefix}_*.json"))


def patch_key_from_filename(name: str) -> str:
    # expected patterns: gt_..._patch001.npz OR est_input_..._patch001_solution.npz etc.
    stem = Path(name).stem
    # remove trailing _solution if present
    if stem.endswith("_solution"):
        stem = stem[: -len("_solution")]
    # remove leading gt_/est_input_ if present
    if stem.startswith("gt_"):
        stem = stem[len("gt_") :]
    if stem.startswith("est_input_"):
        stem = stem[len("est_input_") :]
    return stem


def load_npz_meta_scalars(path: Path) -> Dict[str, float]:
    """
    Read scalar meta_* fields from solver npz, if present.
    """
    out: Dict[str, float] = {}
    with np.load(path, allow_pickle=False) as z:
        for k in z.files:
            if not str(k).startswith("meta_"):
                continue
            try:
                v = z[k]
                if np.isscalar(v):
                    out[str(k)] = float(v)
                elif isinstance(v, np.ndarray) and v.size == 1:
                    out[str(k)] = float(v.reshape(-1)[0])
            except (TypeError, ValueError):
                continue
    return out


def load_npz_optional_link_arrays(path: Path) -> Dict[str, np.ndarray]:
    wanted = (
        "A_obs_virtual",
        "A_hat_virtual",
        "L_km_virtual",
        "valid_links_virtual",
    )
    out: Dict[str, np.ndarray] = {}
    try:
        with np.load(path, allow_pickle=False) as z:
            for name in wanted:
                if name in z.files:
                    out[name] = np.asarray(z[name])
    except Exception:
        return {}
    return out


def native_objective_params_from_meta(meta: Dict[str, float]) -> Optional[Dict[str, float | str]]:
    """
    Infer solver-native objective weights from scalar meta_* fields saved in solution NPZ.
    Returns standardized weights (w_1d, w_total, w_2d) plus scaling label.
    """
    has_constrained = any(
        k in meta
        for k in (
            "meta_alphas_1d",
            "meta_alphas_2d",
            "meta_alphas_total",
            "meta_alpha_1d",
            "meta_alpha_2d",
            "meta_alpha_total",
        )
    )
    if has_constrained:
        return dict(
            objective_scaling="CONSTRAINED_NORMALIZED",
            w_atten=float(meta.get("meta_alpha_atten", 1.0)),
            w_1d=float(meta.get("meta_alpha_1d", meta.get("meta_alphas_1d", 0.0))),
            w_total=float(meta.get("meta_alpha_total", meta.get("meta_alphas_total", 0.0))),
            w_2d=float(meta.get("meta_alpha_2d", meta.get("meta_alphas_2d", 0.0))),
        )

    has_norm = any(k in meta for k in ("meta_j2_w", "meta_j3_w", "meta_j4_w"))
    if has_norm:
        return dict(
            objective_scaling="NORMALIZED",
            w_1d=float(meta.get("meta_j2_w", 0.0)),
            w_total=float(meta.get("meta_j3_w", 0.0)),
            w_2d=float(meta.get("meta_j4_w", 0.0)),
        )

    has_unnorm = any(k in meta for k in ("meta_lambda", "meta_mu", "meta_eta"))
    if has_unnorm:
        return dict(
            objective_scaling="UNNORMALIZED",
            w_1d=float(meta.get("meta_lambda", 0.0)),
            w_total=float(meta.get("meta_mu", 0.0)),
            w_2d=float(meta.get("meta_eta", 0.0)),
        )

    return None


def eps_applicable_from_meta(meta: Dict[str, float]) -> bool:
    # Linear-neighbor objectives do not use epsilon in their objective terms.
    if bool(meta.get("meta_use_linear_j3", False)):
        return False
    if "meta_eps" not in meta:
        return False
    return True


def load_json_dict(path: Path) -> Dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8") as f:
            obj = json.load(f)
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def build_neighbor_triplets(H: int, W: int) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    a_idx: List[int] = []
    b_idx: List[int] = []
    c_idx: List[int] = []
    for i in range(H):
        base = i * W
        for j in range(W - 2):
            a_idx.append(base + j)
            b_idx.append(base + j + 1)
            c_idx.append(base + j + 2)
    for i in range(H - 2):
        base0 = i * W
        base1 = (i + 1) * W
        base2 = (i + 2) * W
        for j in range(W):
            a_idx.append(base0 + j)
            b_idx.append(base1 + j)
            c_idx.append(base2 + j)
    if not a_idx:
        z = np.zeros(0, dtype=np.int64)
        return z, z, z
    return (
        np.asarray(a_idx, dtype=np.int64),
        np.asarray(b_idx, dtype=np.int64),
        np.asarray(c_idx, dtype=np.int64),
    )


# ----------------------------
# Coverage + distance from est_input JSON
# ----------------------------

def load_est_payload(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def compute_coverage_map(est: dict) -> Tuple[np.ndarray, int, int]:
    """
    coverage(p) = #distinct links that intersect pixel p
    Uses segments_by_link with pixel indices (i,j).
    """
    header = est["header"]
    H = int(header["H"])
    W = int(header["W"])
    P = H * W

    segs: Dict[str, list] = est.get("segments_by_link", {})
    pairs: List[Tuple[int, int]] = []
    for li_str, seg_list in segs.items():
        li = int(li_str)
        for s in seg_list:
            i = int(s["i"]); j = int(s["j"])
            p = i * W + j
            pairs.append((p, li))

    if not pairs:
        return np.zeros((H, W), dtype=np.int32), H, W

    arr = np.array(pairs, dtype=np.int64)
    # unique pixel-link pairs
    arr_u = np.unique(arr, axis=0)
    cov = np.bincount(arr_u[:, 0], minlength=P).astype(np.int32)
    return cov.reshape(H, W), H, W


def point_to_segment_dist(px: np.ndarray, py: np.ndarray,
                          x0: np.ndarray, y0: np.ndarray,
                          x1: np.ndarray, y1: np.ndarray) -> np.ndarray:
    """
    Vectorized distance from points (px,py) to segments (x0,y0)-(x1,y1), per-segment.
    px,py are scalars or arrays broadcastable to segment arrays.
    Returns distances in same units as inputs.
    """
    vx = x1 - x0
    vy = y1 - y0
    wx = px - x0
    wy = py - y0
    vv = vx * vx + vy * vy
    # avoid div0
    vv = np.where(vv <= 1e-12, 1e-12, vv)
    t = (wx * vx + wy * vy) / vv
    t = np.clip(t, 0.0, 1.0)
    projx = x0 + t * vx
    projy = y0 + t * vy
    dx = px - projx
    dy = py - projy
    return np.sqrt(dx * dx + dy * dy)



def compute_link_kth_neighbor_distances(
    est: dict,
    k_values: Sequence[int],
) -> Dict[int, np.ndarray]:
    """
    For each link, compute distance (meters) to its k-th closest other link
    using exact segment-to-segment distance.
    """
    links = est.get("links", []) or []
    n_links = int(len(links))
    ks = sorted({int(k) for k in k_values if int(k) >= 1})
    if not ks:
        return {}
    if n_links == 0:
        return {k: np.zeros((0,), dtype=np.float64) for k in ks}

    x0 = np.asarray([float(L["x0_m"]) for L in links], dtype=np.float64)
    y0 = np.asarray([float(L["y0_m"]) for L in links], dtype=np.float64)
    x1 = np.asarray([float(L["x1_m"]) for L in links], dtype=np.float64)
    y1 = np.asarray([float(L["y1_m"]) for L in links], dtype=np.float64)

    out = {k: np.full((n_links,), np.inf, dtype=np.float64) for k in ks}

    for i in range(n_links):
        ax0 = float(x0[i]); ay0 = float(y0[i]); ax1 = float(x1[i]); ay1 = float(y1[i])
        ax0v = np.full_like(x0, ax0)
        ay0v = np.full_like(y0, ay0)
        ax1v = np.full_like(x1, ax1)
        ay1v = np.full_like(y1, ay1)

        d1 = point_to_segment_dist(ax0, ay0, x0, y0, x1, y1)
        d2 = point_to_segment_dist(ax1, ay1, x0, y0, x1, y1)
        d3 = point_to_segment_dist(x0, y0, ax0v, ay0v, ax1v, ay1v)
        d4 = point_to_segment_dist(x1, y1, ax0v, ay0v, ax1v, ay1v)
        d = np.minimum(np.minimum(d1, d2), np.minimum(d3, d4))
        d[i] = np.inf

        for k in ks:
            if (n_links - 1) < k:
                out[k][i] = np.inf
            else:
                out[k][i] = float(np.partition(d, k - 1)[k - 1])

    return out


def compute_dk_maps_sampled_points(
    est: dict,
    k_values: Sequence[int],
    *,
    sample_spacing_m: float = 250.0,
    k_query_samples: int = 48,
    chunk_size: int = 8000,
    max_samples_per_link: int = 200,
    warn_shortfall: bool = True,
    warn_threshold_frac: float = 0.01,
    debug_label: Optional[str] = None,
) -> Tuple[Dict[int, np.ndarray], float]:
    """
    Compute dk(p) = distance in meters from pixel center to the k-th closest link (point-to-segment),
    for each k in k_values.

    This method matches the *older* analyzer: it samples points along each segment at a fixed spacing,
    builds a KD-tree over those sampled points, then uses nearest sampled points to propose candidate
    links. Exact point-to-segment distances are computed for candidate links, and the k-th smallest is taken.

    Returns (dict k -> dk_map_m, pixel_size_m)
    """
    header = est["header"]
    links = est["links"]
    H = int(header["H"])
    W = int(header["W"])
    pix = float(header["pixel_size_m"])

    k_values = sorted({int(k) for k in k_values if int(k) >= 1})
    if not k_values:
        return {}, pix
    k_max = max(k_values)

    if not links or len(links) < k_max:
        return {k: np.full((H, W), np.inf, dtype=np.float64) for k in k_values}, pix

    try:
        from scipy.spatial import cKDTree  # type: ignore
    except Exception as e:
        raise RuntimeError("scipy is required for dk computation (scipy.spatial.cKDTree).") from e

    x0 = np.array([float(L["x0_m"]) for L in links], dtype=np.float64)
    y0 = np.array([float(L["y0_m"]) for L in links], dtype=np.float64)
    x1 = np.array([float(L["x1_m"]) for L in links], dtype=np.float64)
    y1 = np.array([float(L["y1_m"]) for L in links], dtype=np.float64)

    dx = x1 - x0
    dy = y1 - y0
    lengths = np.sqrt(dx * dx + dy * dy)
    spacing = max(1.0, float(sample_spacing_m))

    sample_xy_list = []
    sample_to_link_list = []

    for li in range(len(links)):
        L = float(lengths[li])
        n = max(2, int(math.ceil(L / spacing)) + 1)
        n = min(n, int(max_samples_per_link))
        ts = np.linspace(0.0, 1.0, n, dtype=np.float64)
        xs = x0[li] + ts * dx[li]
        ys = y0[li] + ts * dy[li]
        sample_xy_list.append(np.stack([xs, ys], axis=1))
        sample_to_link_list.append(np.full((n,), li, dtype=np.int32))

    sample_xy = np.concatenate(sample_xy_list, axis=0).astype(np.float64)
    sample_to_link = np.concatenate(sample_to_link_list, axis=0)

    tree = cKDTree(sample_xy)

    # pixel centers in meters
    xs = (np.arange(W, dtype=np.float64) + 0.5) * pix
    ys = (np.arange(H, dtype=np.float64) + 0.5) * pix
    X, Y = np.meshgrid(xs, ys)
    pts = np.stack([X.ravel(), Y.ravel()], axis=1).astype(np.float64)

    out = {k: np.full((H * W,), np.inf, dtype=np.float64) for k in k_values}

    kq = int(k_query_samples)
    kq = max(kq, 12)
    # cap kq to number of samples
    kq = min(kq, int(sample_xy.shape[0]))

    total_queries = 0
    shortfall_queries = 0

    for start in range(0, pts.shape[0], int(chunk_size)):
        end = min(pts.shape[0], start + int(chunk_size))
        q = pts[start:end]

        # query nearest sample points
        _, nn_idx = tree.query(q, k=kq, workers=-1)
        if nn_idx.ndim == 1:
            nn_idx = nn_idx[:, None]

        for bi in range(q.shape[0]):
            cand_links = np.unique(sample_to_link[nn_idx[bi]])

            # ensure enough candidates
            if cand_links.size < k_max:
                shortfall_queries += 1
                kq2 = min(int(sample_xy.shape[0]), kq * 4)
                _, nn_idx2 = tree.query(q[bi], k=kq2, workers=-1)
                cand_links = np.unique(sample_to_link[np.atleast_1d(nn_idx2)])

            if cand_links.size == 0:
                continue

            # exact point-to-segment distances for candidate links
            px = q[bi, 0]
            py = q[bi, 1]
            ds = point_to_segment_dist(px, py, x0[cand_links], y0[cand_links], x1[cand_links], y1[cand_links])

            if ds.size == 0:
                continue

            for k in k_values:
                if ds.size < k:
                    out[k][start + bi] = np.inf
                else:
                    out[k][start + bi] = float(np.partition(ds, k - 1)[k - 1])

            total_queries += 1

    if warn_shortfall and total_queries > 0:
        frac = float(shortfall_queries) / float(total_queries)
        if frac >= float(warn_threshold_frac):
            lab = f" ({debug_label})" if debug_label else ""
            print(
                f"[dist] Candidate shortfall{lab}: {shortfall_queries}/{total_queries} "
                f"pixels had <{k_max} candidate links "
                f"(k_query_samples={k_query_samples}, frac={frac:.3f})"
            )

    return {k: v.reshape(H, W) for k, v in out.items()}, pix


def compute_dk_maps(est: dict, k_values: Sequence[int], *, max_candidates: int) -> Tuple[Dict[int, np.ndarray], float]:
    """
    dk(p) = distance in meters from pixel-center to the k-th closest link (point-to-segment),
            approximated by using KDTree over endpoints+midpoints to propose candidates.

    Returns (dict k -> dk_map_m, pixel_size_m)
    """
    header = est["header"]
    links = est["links"]
    H = int(header["H"])
    W = int(header["W"])
    pix = float(header["pixel_size_m"])

    k_values = sorted({int(k) for k in k_values if int(k) >= 1})
    if not k_values:
        return {}, pix
    k_max = max(k_values)

    if not links or len(links) < k_max:
        return {k: np.full((H, W), np.inf, dtype=np.float64) for k in k_values}, pix

    x0 = np.array([float(L["x0_m"]) for L in links], dtype=np.float64)
    y0 = np.array([float(L["y0_m"]) for L in links], dtype=np.float64)
    x1 = np.array([float(L["x1_m"]) for L in links], dtype=np.float64)
    y1 = np.array([float(L["y1_m"]) for L in links], dtype=np.float64)
    mx = 0.5 * (x0 + x1)
    my = 0.5 * (y0 + y1)
    n_links = len(links)

    try:
        from scipy.spatial import cKDTree  # type: ignore
    except Exception as e:
        raise RuntimeError("scipy is required for dk computation (scipy.spatial.cKDTree).") from e

    pts = np.vstack([
        np.stack([x0, y0], axis=1),
        np.stack([x1, y1], axis=1),
        np.stack([mx, my], axis=1),
    ])
    pt_to_link = np.concatenate([np.arange(n_links), np.arange(n_links), np.arange(n_links)])
    tree = cKDTree(pts)

    # pixel centers in meters; assume pixel (i,j) spans [j*pix,(j+1)*pix] etc.
    xs = (np.arange(W, dtype=np.float64) + 0.5) * pix
    ys = (np.arange(H, dtype=np.float64) + 0.5) * pix
    X, Y = np.meshgrid(xs, ys)

    # query K nearest proxy points
    K = int(max_candidates)
    K = max(k_max, min(K, pts.shape[0]))
    _, idx_proxy = tree.query(np.stack([X.ravel(), Y.ravel()], axis=1), k=K, workers=-1)
    # ensure 2d
    if K == 1:
        idx_proxy = idx_proxy[:, None]

    out = {k: np.empty(H * W, dtype=np.float64) for k in k_values}
    # compute per pixel
    for t in range(H * W):
        cand_links = np.unique(pt_to_link[idx_proxy[t]])
        # true segment distances for these links
        dist = point_to_segment_dist(X.ravel()[t], Y.ravel()[t], x0[cand_links], y0[cand_links], x1[cand_links], y1[cand_links])
        for k in k_values:
            if dist.size < k:
                out[k][t] = np.inf
            else:
                out[k][t] = float(np.partition(dist, k - 1)[k - 1])
    return {k: v.reshape(H, W) for k, v in out.items()}, pix


def parse_bins(edges: Sequence[float]) -> List[Tuple[Optional[float], Optional[float], str]]:
    """
    edges = [125,375,750,...]
    bins: <=e0, (e0,e1], ..., >elast
    """
    e = list(map(float, edges))
    out: List[Tuple[Optional[float], Optional[float], str]] = []
    if not e:
        out.append((None, None, "all"))
        return out
    out.append((None, e[0], f"≤{int(e[0])}"))
    for a, b in zip(e, e[1:]):
        out.append((a, b, f"({int(a)},{int(b)}]"))
    out.append((e[-1], None, f">{int(e[-1])}"))
    return out


def assign_bin_labels(values: np.ndarray, bins: List[Tuple[Optional[float], Optional[float], str]]) -> np.ndarray:
    lab = np.empty(values.shape, dtype=object)
    for lo, hi, name in bins:
        if lo is None:
            m = values <= hi
        elif hi is None:
            m = values > lo
        else:
            m = (values > lo) & (values <= hi)
        lab[m] = name
    return lab


def parse_coverage_bins(cfg_bins: Sequence[Any]) -> Tuple[List[int], Optional[int]]:
    exact: List[int] = []
    ge: Optional[int] = None
    for b in cfg_bins:
        if isinstance(b, int):
            exact.append(int(b))
        elif isinstance(b, str) and b.strip().endswith("+"):
            ge = int(b.strip()[:-1])
        else:
            raise ValueError(f"Invalid coverage bin {b!r}. Use ints or 'K+'.")
    exact = sorted(set(exact))
    return exact, ge


def parse_fp_fn_thresholds(cfg: dict, *, default_threshold_mmph: float) -> List[float]:
    """
    Build sorted unique thresholds for FP/FN reporting.
    Supported config:
      rain.fp_fn_thresholds_mmph: [0.5, 1.0, 2.0]
      rain.fp_fn_threshold_sweep_mmph: {start: 0.5, stop: 5.0, step: 0.5}
    If neither is provided, falls back to a default sweep [0.2, 5.0] in 0.1 mm/h steps.
    """
    vals = deep_get(cfg, "rain.fp_fn_thresholds_mmph", None)
    if vals is not None:
        if not isinstance(vals, (list, tuple)):
            raise SystemExit("rain.fp_fn_thresholds_mmph must be a list of numbers.")
        out = sorted({float(v) for v in vals if float(v) >= 0.0})
        if not out:
            raise SystemExit("rain.fp_fn_thresholds_mmph must include at least one non-negative value.")
        return out

    sweep = deep_get(cfg, "rain.fp_fn_threshold_sweep_mmph", None)
    if isinstance(sweep, dict):
        start = float(sweep.get("start", default_threshold_mmph))
        stop = float(sweep.get("stop", start))
        step = float(sweep.get("step", 0.0))
        if start < 0.0 or stop < 0.0 or step <= 0.0:
            raise SystemExit("rain.fp_fn_threshold_sweep_mmph requires start>=0, stop>=0, step>0.")
        if stop < start:
            raise SystemExit("rain.fp_fn_threshold_sweep_mmph requires stop >= start.")
        n = int(math.floor((stop - start) / step)) + 1
        out = [round(start + i * step, 10) for i in range(max(0, n))]
        if not out:
            out = [start]
        if out[-1] < stop - 1e-10:
            out.append(round(stop, 10))
        return sorted({float(v) for v in out if float(v) >= 0.0})

    start = min(0.2, float(default_threshold_mmph))
    stop = max(5.0, float(default_threshold_mmph))
    step = 0.1
    n = int(math.floor((stop - start) / step)) + 1
    out = [round(start + i * step, 10) for i in range(max(0, n))]
    if out[-1] < stop - 1e-10:
        out.append(round(stop, 10))
    return sorted({float(v) for v in out if float(v) >= 0.0})


def coverage_bin_label(v: int, exact: List[int], ge: Optional[int]) -> Optional[str]:
    if v in exact:
        return str(v)
    if ge is not None and v >= ge:
        return f"{ge}+"
    return None


# ----------------------------
# Stats helpers
# ----------------------------

def append_average_rows(
    rows: List[Dict[str, Any]],
    *,
    group_keys: List[str],
    patch_key_field: str = "patch_key",
    average_label: str = "AVERAGE",
) -> List[Dict[str, Any]]:
    """
    Append average rows over patches, grouped by group_keys.
    """
    if not rows:
        return rows

    def _is_number(v: Any) -> bool:
        return isinstance(v, (int, float, np.integer, np.floating)) and not isinstance(v, bool)

    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = {}
    for r in rows:
        if r.get(patch_key_field) == average_label:
            continue
        g = tuple(r.get(k, None) for k in group_keys)
        grouped.setdefault(g, []).append(r)

    avg_rows: List[Dict[str, Any]] = []
    for g, grp_rows in grouped.items():
        if not grp_rows:
            continue
        out: Dict[str, Any] = {patch_key_field: average_label}
        for i, k in enumerate(group_keys):
            out[k] = g[i]

        keys = set()
        for rr in grp_rows:
            keys.update(rr.keys())
        keys.discard(patch_key_field)
        for k in group_keys:
            keys.discard(k)

        for k in sorted(keys):
            vals = [rr.get(k, None) for rr in grp_rows]
            nums = [float(v) for v in vals if _is_number(v)]
            if nums:
                out[k] = float(np.mean(nums))
            else:
                out[k] = ""
        avg_rows.append(out)

    return rows + avg_rows


def append_native_objective_definition_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Append human-readable mathematical definitions for native objective terms.
    """
    defs: List[Dict[str, Any]] = [
        dict(
            patch_key="DEFINITION",
            solver="J_weighted_sum",
            definition=(
                "Weighted sum objective value reported in this sheet. For constrained solvers, "
                "J_atten is reported as a constraint metric while J_weighted_sum uses only regularization terms."
            ),
        ),
        dict(
            patch_key="DEFINITION",
            solver="J_atten",
            definition="J_atten = sum over valid links of (A_hat - A_obs)^2/L_km, normalized by #valid_links in this sheet.",
        ),
        dict(
            patch_key="DEFINITION",
            solver="J_1d",
            definition="J_1d = first-order smoothness term over neighboring pixels, normalized by #pixels in this sheet.",
        ),
        dict(
            patch_key="DEFINITION",
            solver="J_total",
            definition="J_total = total-field energy term sum over pixels of R^2, normalized by #pixels in this sheet.",
        ),
        dict(
            patch_key="DEFINITION",
            solver="J_2d",
            definition="J_2d = second-order smoothness term over collinear triplets, normalized by #pixels in this sheet.",
        ),
    ]
    return rows + defs


def _safe_ratio(num: Any, den: Any) -> Optional[float]:
    try:
        n = float(num)
        d = float(den)
        if not np.isfinite(n) or not np.isfinite(d) or d == 0.0:
            return None
        return float(n / d)
    except Exception:
        return None


def enrich_linkstats_with_baseline_jatten(sheets: Dict[str, List[Dict[str, Any]]]) -> None:
    """
    For each LinkStats_GTvs<ALG> (ALG not in {IDW, ILDW}), add baseline columns:
      - J_atten_IDW
      - J_atten_ILDW
      - J_atten_ALG_over_IDW
      - J_atten_ALG_over_ILDW
    Matching is by patch_key.
    """
    idw_sheet = sheets.get("LinkStats_GTvsIDW", [])
    ildw_sheet = sheets.get("LinkStats_GTvsILDW", [])

    idw_by_patch: Dict[str, float] = {}
    for r in idw_sheet:
        pk = str(r.get("patch_key", ""))
        if not pk:
            continue
        try:
            idw_by_patch[pk] = float(r.get("J_atten_all"))
        except Exception:
            continue

    ildw_by_patch: Dict[str, float] = {}
    for r in ildw_sheet:
        pk = str(r.get("patch_key", ""))
        if not pk:
            continue
        try:
            ildw_by_patch[pk] = float(r.get("J_atten_all"))
        except Exception:
            continue

    for sheet_name, rows in sheets.items():
        if not sheet_name.startswith("LinkStats_GTvs"):
            continue
        alg = sheet_name[len("LinkStats_GTvs") :]
        if alg in {"IDW", "ILDW"}:
            continue
        for row in rows:
            pk = str(row.get("patch_key", ""))
            if not pk or pk == "DEFINITION":
                continue
            j_alg = row.get("J_atten_all", None)
            j_idw = idw_by_patch.get(pk, None)
            j_ildw = ildw_by_patch.get(pk, None)
            row["J_atten_IDW"] = j_idw
            row["J_atten_ILDW"] = j_ildw
            row["J_atten_ALG_over_IDW"] = _safe_ratio(j_alg, j_idw)
            row["J_atten_ALG_over_ILDW"] = _safe_ratio(j_alg, j_ildw)


def enrich_binned_stats_with_baseline_ratios(sheets: Dict[str, List[Dict[str, Any]]]) -> None:
    """
    For Coverage/Distance/Overall per-solver tabs, add ALG-vs-baseline ratios (ALG!=IDW,ILDW)
    for key metrics: mae, median_error, p90_error, p99_error.
    """
    metric_cols = ("mean_abs", "std_abs", "median_abs", "p90_abs", "p95_abs", "p99_abs")
    for sheet_name, rows in sheets.items():
        if "GTvs" not in sheet_name:
            continue
        if not (
            sheet_name.startswith("CoverageStats_")
            or sheet_name.startswith("DistanceStats")
            or sheet_name.startswith("OverallStats_")
        ):
            continue

        base, alg = sheet_name.split("GTvs", 1)
        if alg in {"IDW", "ILDW"}:
            continue

        idw_rows = sheets.get(f"{base}GTvsIDW", None)
        ildw_rows = sheets.get(f"{base}GTvsILDW", None)
        if idw_rows is None or ildw_rows is None:
            continue

        def row_key(r: Dict[str, Any]) -> Tuple[Any, ...]:
            return (
                r.get("patch_key", None),
                r.get("mask_type", None),
                r.get("coverage_bin", None),
                r.get("distance_bin_m", None),
            )

        idw_map = {row_key(r): r for r in idw_rows}
        ildw_map = {row_key(r): r for r in ildw_rows}
        for r in rows:
            if str(r.get("patch_key", "")) == "DEFINITION":
                continue
            k = row_key(r)
            r_idw = idw_map.get(k, {})
            r_ildw = ildw_map.get(k, {})
            for m in metric_cols:
                v_alg = r.get(m, None)
                v_idw = r_idw.get(m, None)
                v_ildw = r_ildw.get(m, None)
                r[f"{m}_IDW"] = v_idw
                r[f"{m}_ILDW"] = v_ildw
                r[f"{m}_over_IDW"] = _safe_ratio(v_alg, v_idw)
                r[f"{m}_over_ILDW"] = _safe_ratio(v_alg, v_ildw)


def enrich_overall_by_solver_ratios(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    metric_cols = ("mean_abs", "std_abs", "median_abs", "p90_abs", "p95_abs", "p99_abs")
    out: List[Dict[str, Any]] = [dict(r) for r in rows]
    key_fields = ("patch_key", "mask_type", "distance_bin_m")
    idw_map: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    ildw_map: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    for r in out:
        k = tuple(r.get(f, None) for f in key_fields)
        solver = str(r.get("solver", ""))
        if solver == "IDW":
            idw_map[k] = r
        elif solver == "ILDW":
            ildw_map[k] = r
    for r in out:
        solver = str(r.get("solver", ""))
        if solver in {"IDW", "ILDW"}:
            continue
        k = tuple(r.get(f, None) for f in key_fields)
        rb_idw = idw_map.get(k, {})
        rb_ildw = ildw_map.get(k, {})
        for m in metric_cols:
            v_alg = r.get(m, None)
            v_idw = rb_idw.get(m, None)
            v_ildw = rb_ildw.get(m, None)
            r[f"{m}_IDW"] = v_idw
            r[f"{m}_ILDW"] = v_ildw
            r[f"{m}_over_IDW"] = _safe_ratio(v_alg, v_idw)
            r[f"{m}_over_ILDW"] = _safe_ratio(v_alg, v_ildw)
    return out


def evaluate_objective_values(
    prob,
    R_field: np.ndarray,
    *,
    lam: float,
    mu: float,
    eps: float,
    eta: float = 0.0,
) -> Dict[str, float]:
    if R_field.shape != (prob.H, prob.W):
        raise ValueError(f"R_field shape {R_field.shape} != (H,W)=({prob.H},{prob.W})")
    R = np.asarray(R_field, dtype=np.float64).ravel()

    # --- J1 data term ---
    pix = prob.pix_idx
    li = prob.link_idx
    ds = prob.ds_km
    A_obs = prob.A_obs
    L_km = prob.L_km
    valid = prob.valid_links
    k = prob.k
    a = prob.alpha

    Rp = R[pix]
    pow_a = np.power(Rp, a[li], where=(Rp > 0), out=np.zeros_like(Rp))
    contrib = ds * k[li] * pow_a
    A_hat = np.bincount(li, weights=contrib, minlength=prob.L).astype(np.float64)

    r = np.zeros(prob.L, dtype=np.float64)
    r[valid] = (A_hat[valid] - A_obs[valid]) / L_km[valid]
    J1 = float(np.dot(r[valid], r[valid]))

    # --- J2 smoothness term ---
    du = R[prob.n_u] - R[prob.n_v]
    J2 = float(np.dot(du, du))

    # --- J3 shrinkage term ---
    J3 = float(np.dot(R, R))

    # --- J4 triplet term: sum(((f(b)-f(a))-(f(c)-f(b)))^2) on collinear triplets ---
    t_a, t_b, t_c = build_neighbor_triplets(prob.H, prob.W)
    if t_a.size > 0:
        d2 = (R[t_b] - R[t_a]) - (R[t_c] - R[t_b])
        J4 = float(np.dot(d2, d2))
    else:
        J4 = 0.0

    lamJ2 = float(lam * J2)
    muJ3 = float(mu * J3)
    etaJ4 = float(eta * J4)
    J = J1 + lamJ2 + muJ3 + etaJ4
    n_valid_links = int(np.sum(valid))
    n_pixels = int(prob.P)
    return dict(
        J=float(J),
        J1=float(J1),
        J2=float(J2),
        J3=float(J3),
        J4=float(J4),
        wJ1=float(J1),
        wJ2=lamJ2,
        wJ3=muJ3,
        wJ4=etaJ4,
        etaJ4=etaJ4,
        eta=float(eta),
        n_valid_links=n_valid_links,
        n_pixels=n_pixels,
    )


def analyze_single_patch(task: Dict[str, Any]) -> Dict[str, Any]:
    """
    Per-(solver,patch) worker. Returns all per-patch artifacts so callers can
    aggregate deterministically in key order.
    """
    key = str(task["key"])
    label = str(task["label"])
    name = str(task["name"])
    gt_path = Path(str(task["gt_path"]))
    sol_path = Path(str(task["sol_path"]))
    est_path = Path(str(task["est_path"]))

    gt_key_pref = list(task["gt_key_pref"])
    sol_key_pref = list(task["sol_key_pref"])
    thr = float(task["thr"])
    fp_fn_thresholds = [float(v) for v in task.get("fp_fn_thresholds", [thr])]
    cov_exact = list(task["cov_exact"])
    cov_ge = task["cov_ge"]
    k_values = [int(k) for k in task["k_values"]]
    dist_bins = list(task["dist_bins"])
    jatten_k_values = [int(k) for k in task.get("jatten_k_values", k_values)]
    jatten_dist_bins = list(task.get("jatten_dist_bins", dist_bins))
    dist_method = str(task["dist_method"])
    sample_spacing_m = float(task["sample_spacing_m"])
    k_query_samples = int(task["k_query_samples"])
    chunk_size = int(task["chunk_size"])
    max_samples_per_link = int(task["max_samples_per_link"])
    max_candidates = int(task["max_candidates"])
    rainy_bins_enabled = bool(task["rainy_bins_enabled"])
    rainy_intervals = list(task["rainy_intervals"])
    obj_enabled = bool(task["obj_enabled"])
    obj_pairs = [tuple(x) for x in task.get("obj_pairs", [])]
    obj_eps = float(task.get("obj_eps", DEFAULT_REFERENCE_EPS))
    native_obj_enabled = bool(task.get("native_obj_enabled", True))
    ref_eval_enabled = bool(task.get("ref_eval_enabled", False))
    ref_eval_lam = float(task.get("ref_eval_lam", DEFAULT_REFERENCE_W_SMOOTH))
    ref_eval_mu = float(task.get("ref_eval_mu", DEFAULT_REFERENCE_W_SHRINK))
    ref_eval_eta = float(task.get("ref_eval_eta", DEFAULT_REFERENCE_W_SECOND_DER))
    ref_eval_eps = float(task.get("ref_eval_eps", DEFAULT_REFERENCE_EPS))
    include_rae_hist = bool(task["include_rae_hist"])
    # Per-patch stopping diagnostics from solver optinfo JSON.
    optinfo_path = sol_path.with_name(f"{sol_path.stem}_optinfo.json")
    has_optinfo = optinfo_path.exists()
    opt = load_json_dict(optinfo_path) if has_optinfo else {}
    itertrace_path = sol_path.with_name(f"{sol_path.stem}_itertrace.json")
    has_itertrace = itertrace_path.exists()
    itertrace = load_json_dict(itertrace_path) if has_itertrace else {}
    iter_entries_raw = itertrace.get("iterations", [])
    iter_entries = iter_entries_raw if isinstance(iter_entries_raw, list) else []
    iter_summary_raw = itertrace.get("summary", {})
    iter_summary = iter_summary_raw if isinstance(iter_summary_raw, dict) else {}
    stop_reason_raw = opt.get("stop_reason", None)
    stop_reason = str(stop_reason_raw) if stop_reason_raw is not None and str(stop_reason_raw) != "" else None
    reason_bucket = stop_reason if stop_reason is not None else ("missing_optinfo" if not has_optinfo else "unknown")
    stop_row = dict(
        patch_key=key,
        solver=label,
        solver_name=name,
        has_optinfo=bool(has_optinfo),
        stop_reason=stop_reason,
        reason_bucket=reason_bucket,
        success=opt.get("success", None),
        status=opt.get("status", None),
        message=opt.get("message", None),
        nit=opt.get("total_inner_iterations", opt.get("nit", opt.get("iteration", None))),
        nit_last_stage=opt.get("final_stage_iterations", None),
        outer_iterations=opt.get("outer_iterations", iter_summary.get("outer_iterations", None)),
        total_inner_iterations=opt.get("total_inner_iterations", iter_summary.get("total_inner_iterations", None)),
        final_stage_iterations=opt.get("final_stage_iterations", iter_summary.get("final_stage_iterations", None)),
        stage_iteration_counts=json.dumps(iter_summary.get("stage_iteration_counts", opt.get("stage_iteration_counts", [])))
            if (
                iter_summary.get("stage_iteration_counts", None) is not None
                or opt.get("stage_iteration_counts", None) is not None
            )
            else None,
        init_method=opt.get("init_method", iter_summary.get("init_method", None)),
        nfev=opt.get("nfev", None),
        njev=opt.get("njev", None),
        proj_grad_inf=opt.get("proj_grad_inf", None),
        rel_decrease=opt.get("rel_decrease", None),
        ftol=opt.get("ftol", None),
        gtol=opt.get("gtol", None),
        ftol_met=opt.get("ftol_met", None),
        gtol_met=opt.get("gtol_met", None),
        line_search_failed=opt.get("line_search_failed", opt.get("line_search_issue", None)),
        maxiter_reached=opt.get("maxiter_reached", (str(reason_bucket) == "maxiter")),
        optinfo_json=str(optinfo_path) if has_optinfo else None,
    )

    gt = load_npz_first_key(gt_path, gt_key_pref)
    pred = load_npz_first_key(sol_path, sol_key_pref)
    if gt.shape != pred.shape:
        raise ValueError(f"Shape mismatch for {key}: GT {gt.shape} vs {label} {pred.shape}")
    gt = gt.astype(np.float64)
    pred = pred.astype(np.float64)

    finite_mask = np.isfinite(gt) & np.isfinite(pred)
    gt_flat = gt[finite_mask].ravel()
    pred_flat = pred[finite_mask].ravel()
    if gt_flat.size == 0:
        rmse_mmph = None
        bias_mmph = None
        pearson_corr = None
    else:
        diff_flat = pred_flat - gt_flat
        rmse_mmph = float(np.sqrt(np.mean(diff_flat * diff_flat)))
        bias_mmph = float(np.mean(diff_flat))
        gt_std = float(np.std(gt_flat, ddof=0))
        pred_std = float(np.std(pred_flat, ddof=0))
        if gt_flat.size >= 2 and gt_std > 0.0 and pred_std > 0.0:
            pearson_corr = float(np.corrcoef(gt_flat, pred_flat)[0, 1])
        else:
            pearson_corr = None

    rainy = gt >= thr
    gt_wet = gt >= thr
    pred_wet = pred >= thr
    fp = np.logical_and(pred_wet, ~gt_wet)  # predicted wet but GT dry
    fn = np.logical_and(~pred_wet, gt_wet)  # predicted dry but GT wet
    total_pixels = gt.size
    dry_fp_rate = (float(np.sum(fp)) / float(total_pixels)) if total_pixels > 0 else None
    dry_fn_rate = (float(np.sum(fn)) / float(total_pixels)) if total_pixels > 0 else None
    fp_fn_by_threshold: List[Dict[str, Any]] = []
    for thr_i in fp_fn_thresholds:
        gt_wet_i = gt >= float(thr_i)
        pred_wet_i = pred >= float(thr_i)
        tp_i = np.logical_and(pred_wet_i, gt_wet_i)
        fp_i = np.logical_and(pred_wet_i, ~gt_wet_i)
        fn_i = np.logical_and(~pred_wet_i, gt_wet_i)
        tn_i = np.logical_and(~pred_wet_i, ~gt_wet_i)
        n_all_i = int(gt.size)
        n_wet_i = int(np.sum(gt_wet_i))
        n_dry_i = int(n_all_i - n_wet_i)
        tp_count_i = int(np.sum(tp_i))
        fp_count_i = int(np.sum(fp_i))
        fn_count_i = int(np.sum(fn_i))
        tn_count_i = int(np.sum(tn_i))
        fp_fn_by_threshold.append(
            dict(
                threshold_mmph=float(thr_i),
                tp_count=tp_count_i,
                fp_count=fp_count_i,
                fn_count=fn_count_i,
                tn_count=tn_count_i,
                n_pixels=n_all_i,
                n_wet=n_wet_i,
                n_dry=n_dry_i,
                tp_rate_all=(float(tp_count_i) / float(n_all_i)) if n_all_i > 0 else None,
                fp_rate_all=(float(fp_count_i) / float(n_all_i)) if n_all_i > 0 else None,
                fn_rate_all=(float(fn_count_i) / float(n_all_i)) if n_all_i > 0 else None,
                tn_rate_all=(float(tn_count_i) / float(n_all_i)) if n_all_i > 0 else None,
                fp_rate_dry=(float(fp_count_i) / float(n_dry_i)) if n_dry_i > 0 else None,
                fn_rate_wet=(float(fn_count_i) / float(n_wet_i)) if n_wet_i > 0 else None,
            )
        )

    # coverage map + bins
    est = load_est_payload(est_path)
    cov_map, _, _ = compute_coverage_map(est)
    if cov_map.shape != gt.shape:
        raise ValueError(f"Coverage map shape {cov_map.shape} != GT shape {gt.shape} for {key}")

    # distance maps d_k
    if dist_method in ("sampled_points", "sampled", "samples"):
        dk_maps, _ = compute_dk_maps_sampled_points(
            est,
            k_values,
            sample_spacing_m=sample_spacing_m,
            k_query_samples=k_query_samples,
            chunk_size=chunk_size,
            max_samples_per_link=max_samples_per_link,
        )
    else:
        dk_maps, _ = compute_dk_maps(est, k_values, max_candidates=max_candidates)
    for k in k_values:
        d_map = dk_maps.get(k)
        if d_map is None:
            raise ValueError(f"Missing d{k} map for {key}")
        if d_map.shape != gt.shape:
            raise ValueError(f"d{k} map shape {d_map.shape} != GT shape {gt.shape} for {key}")

    # Precompute full-field error arrays aligned to pixels for bin masking
    signed_full = np.empty_like(gt, dtype=np.float64)
    abs_full = np.empty_like(gt, dtype=np.float64)
    abs_rae_full = np.empty_like(gt, dtype=np.float64)
    abs_mmph_full = np.abs(gt - pred)
    denom = np.where(gt == 0.0, 1.0, gt)
    signed_full[rainy] = (gt[rainy] - pred[rainy]) / denom[rainy]
    abs_full[rainy] = np.abs(signed_full[rainy])
    abs_rae_full[rainy] = abs_full[rainy]
    signed_full[~rainy] = pred[~rainy] - gt[~rainy]
    abs_full[~rainy] = np.abs(signed_full[~rainy])
    abs_rae_full[~rainy] = abs_mmph_full[~rainy] / denom[~rainy]

    gtbin_counts_by_lab: Dict[str, int] = {}
    gtbin_rel_means_by_lab: Dict[str, Optional[float]] = {}
    gtbin_abs_means_by_lab: Dict[str, Optional[float]] = {}
    if rainy_bins_enabled:
        for lo, hi, bin_lab in rainy_intervals:
            gmask_bin = (gt >= lo) & (gt < hi)
            n_bin = int(np.sum(gmask_bin))
            gtbin_counts_by_lab[bin_lab] = n_bin
            if n_bin == 0:
                gtbin_rel_means_by_lab[bin_lab] = None
                gtbin_abs_means_by_lab[bin_lab] = None
                continue
            abs_vals = abs_mmph_full[gmask_bin]
            if lo < 1.0:
                rel_vals = abs_vals
            else:
                den_bin = np.where(gt[gmask_bin] == 0.0, 1.0, gt[gmask_bin])
                rel_vals = abs_vals / den_bin
            gtbin_rel_means_by_lab[bin_lab] = float(np.mean(rel_vals))
            gtbin_abs_means_by_lab[bin_lab] = float(np.mean(abs_vals))

    objective_pairs: List[Tuple[float, float, float, Dict[str, float], Dict[str, float]]] = []
    native_objective: Optional[Dict[str, Any]] = None
    reference_objective_pred: Optional[Dict[str, Any]] = None
    reference_objective_gt: Optional[Dict[str, Any]] = None
    prob = None
    need_prob = bool(obj_enabled)
    native_meta = load_npz_meta_scalars(sol_path) if native_obj_enabled else {}
    native_params = native_objective_params_from_meta(native_meta) if native_obj_enabled else None
    if native_params is not None:
        need_prob = True
    if ref_eval_enabled:
        need_prob = True

    if need_prob:
        from cml_attenuation.solvers.solve_rain_lbfgsb import load_est_input_json as load_est_for_obj  # type: ignore
        prob = load_est_for_obj(est_path, warn=False)

    if obj_enabled:
        assert prob is not None
        for lam, mu, eta in obj_pairs:
            j_gt = evaluate_objective_values(prob, gt, lam=lam, mu=mu, eps=obj_eps, eta=eta)
            j_sol = evaluate_objective_values(prob, pred, lam=lam, mu=mu, eps=obj_eps, eta=eta)
            objective_pairs.append((float(lam), float(mu), float(eta), j_gt, j_sol))
    if native_params is not None and prob is not None:
        w_atten = float(native_params.get("w_atten", 1.0))  # type: ignore[arg-type]
        lam = float(native_params["w_1d"])
        mu = float(native_params["w_total"])
        eta = float(native_params["w_2d"])
        eps_native = float(native_meta.get("meta_eps", obj_eps))
        j_native = evaluate_objective_values(
            prob,
            pred,
            lam=lam,
            mu=mu,
            eps=eps_native,
            eta=eta,
        )
        native_objective = dict(
            objective_scaling=str(native_params["objective_scaling"]),
            w_atten=w_atten,
            w_1d=lam,
            w_total=mu,
            w_2d=eta,
            **j_native,
        )
        if eps_applicable_from_meta(native_meta):
            native_objective["eps"] = eps_native
    if ref_eval_enabled and prob is not None:
        j_ref_pred = evaluate_objective_values(
            prob,
            pred,
            lam=ref_eval_lam,
            mu=ref_eval_mu,
            eps=ref_eval_eps,
            eta=ref_eval_eta,
        )
        reference_objective_pred = dict(
            objective_scaling="NORMALIZED_REFERENCE",
            w_atten=1.0,
            w_1d=ref_eval_lam,
            w_total=ref_eval_mu,
            w_2d=ref_eval_eta,
            **j_ref_pred,
        )
        j_ref_gt = evaluate_objective_values(
            prob,
            gt,
            lam=ref_eval_lam,
            mu=ref_eval_mu,
            eps=ref_eval_eps,
            eta=ref_eval_eta,
        )
        reference_objective_gt = dict(
            objective_scaling="NORMALIZED_REFERENCE",
            w_atten=1.0,
            w_1d=ref_eval_lam,
            w_total=ref_eval_mu,
            w_2d=ref_eval_eta,
            **j_ref_gt,
        )

    cov_rows: List[Dict[str, Any]] = []
    for mask_name, mask in (("rainy", rainy), ("nonrainy", ~rainy)):
        cov_vals = cov_map[mask].ravel()
        for v in cov_exact + ([] if cov_ge is None else [cov_ge]):
            if cov_ge is not None and v == cov_ge:
                m = cov_vals >= cov_ge
                bin_lab = f"{cov_ge}+"
            else:
                m = cov_vals == v
                bin_lab = str(v)
            if not np.any(m):
                cov_rows.append(dict(
                    patch_key=key, mask_type=mask_name, coverage_bin=bin_lab,
                    n_pixels=0,
                    mean_signed=0.0, median_signed=0.0,
                    mean_abs=0.0, std_abs=0.0,
                    median_abs=0.0, p90_abs=0.0, p95_abs=0.0, p99_abs=0.0, linf_abs=0.0,
                    l1_rae_sum=0.0,
                    l1_abs_mmph_sum=0.0,
                    l1_abs_mmph_sum_norm_hw=0.0,
                ))
                continue
            if cov_ge is not None and bin_lab.endswith("+"):
                gmask = mask & (cov_map >= cov_ge)
            else:
                gmask = mask & (cov_map == int(bin_lab))
            e_signed = signed_full[gmask].ravel()
            e_abs = abs_full[gmask].ravel()
            l1_rae = float(np.sum(abs_rae_full[gmask]))
            l1_abs_mmph = float(np.sum(abs_mmph_full[gmask]))
            s = stats_row(e_signed, e_abs, l1_rae_sum=l1_rae, l1_abs_mmph_sum=l1_abs_mmph)
            cov_rows.append(dict(
                patch_key=key,
                mask_type=mask_name,
                coverage_bin=bin_lab,
                l1_abs_mmph_sum_norm_hw=(l1_abs_mmph / float(gt.size)),
                **s,
            ))

    dist_rows_by_k: Dict[int, List[Dict[str, Any]]] = {k: [] for k in k_values}
    overall_rows: List[Dict[str, Any]] = []
    bin_counts_entries: List[Tuple[int, str, str, int]] = []
    bin_counts_all_entries: List[Tuple[int, str, int]] = []
    medians_r_entries: List[Tuple[int, str, float]] = []
    medians_n_entries: List[Tuple[int, str, float]] = []
    jatten_medians_entries: List[Tuple[int, str, float]] = []
    jatten_link_bin_counts_entries: List[Tuple[int, str, int]] = []
    p90s_r_entries: List[Tuple[int, str, float]] = []
    p90s_n_entries: List[Tuple[int, str, float]] = []
    rae_hist_entries: List[Tuple[int, str, List[float]]] = []

    # Link-based terms used for LinkStats and J_atten-by-link-distance plots.
    try:
        A_obs, A_hat, L_km, valid, ge10 = compute_link_terms(est, pred)
        link_dists_by_k = compute_link_kth_neighbor_distances(est, jatten_k_values)
    except Exception as e:
        raise RuntimeError(f"Failed to compute link stats for {key} ({label}): {e}") from e
    virtual_link_arrays = load_npz_optional_link_arrays(sol_path)
    A_obs_virtual = np.asarray(virtual_link_arrays.get("A_obs_virtual", np.array([], dtype=np.float64)), dtype=np.float64)
    A_hat_virtual = np.asarray(virtual_link_arrays.get("A_hat_virtual", np.array([], dtype=np.float64)), dtype=np.float64)
    L_km_virtual = np.asarray(virtual_link_arrays.get("L_km_virtual", L_km), dtype=np.float64)
    valid_virtual = np.asarray(virtual_link_arrays.get("valid_links_virtual", valid), dtype=bool)
    ge10_virtual = (L_km_virtual >= 10.0) & valid_virtual if L_km_virtual.size else np.zeros(0, dtype=bool)

    n_valid_links = int(np.sum(valid)) if valid.size > 0 else 0
    den_valid_links = float(max(1, n_valid_links))
    jatten_link_contrib = np.zeros_like(L_km, dtype=np.float64)
    valid_jatten = valid & np.isfinite(L_km) & (L_km > 0.0)
    if np.any(valid_jatten):
        diff = A_hat[valid_jatten] - A_obs[valid_jatten]
        jatten_link_contrib[valid_jatten] = ((diff * diff) / L_km[valid_jatten]) / den_valid_links

    for k in jatten_k_values:
        link_d = np.asarray(link_dists_by_k.get(k, np.full((len(L_km),), np.inf, dtype=np.float64)), dtype=np.float64)
        link_labels = assign_bin_labels(link_d, jatten_dist_bins)
        for _, _, bin_lab in jatten_dist_bins:
            m = valid_jatten & (link_labels == bin_lab)
            cnt = int(np.sum(m))
            jatten_link_bin_counts_entries.append((k, bin_lab, cnt))
            if cnt > 0:
                jatten_medians_entries.append((k, bin_lab, float(np.median(jatten_link_contrib[m]))))

    for k in k_values:
        d_map = dk_maps[k]
        d_vals = d_map.ravel()
        d_labels = assign_bin_labels(d_vals, dist_bins).reshape(gt.shape)
        for _, _, bin_lab in dist_bins:
            gmask_all = (d_labels == bin_lab)
            bin_counts_all_entries.append((k, bin_lab, int(np.sum(gmask_all))))
        for mask_name, mask in (("rainy", rainy), ("nonrainy", ~rainy)):
            for _, _, bin_lab in dist_bins:
                gmask = mask & (d_labels == bin_lab)
                bin_counts_entries.append((k, mask_name, bin_lab, int(np.sum(gmask))))
                e_signed = signed_full[gmask].ravel()
                e_abs = abs_full[gmask].ravel()
                l1_rae = float(np.sum(abs_rae_full[gmask]))
                l1_abs_mmph = float(np.sum(abs_mmph_full[gmask]))
                s = stats_row(e_signed, e_abs, l1_rae_sum=l1_rae, l1_abs_mmph_sum=l1_abs_mmph)
                dist_rows_by_k[k].append(dict(
                    patch_key=key,
                    mask_type=mask_name,
                    distance_bin_m=bin_lab,
                    l1_abs_mmph_sum_norm_hw=(l1_abs_mmph / float(gt.size)),
                    **s,
                ))
                if mask_name == "rainy" and e_abs.size > 0:
                    medians_r_entries.append((k, bin_lab, float(np.median(e_abs))))
                    p90s_r_entries.append((k, bin_lab, float(np.percentile(e_abs, 90))))
                if mask_name == "nonrainy" and e_abs.size > 0:
                    medians_n_entries.append((k, bin_lab, float(np.median(e_abs))))
                    p90s_n_entries.append((k, bin_lab, float(np.percentile(e_abs, 90))))
                if include_rae_hist and mask_name == "rainy" and e_abs.size > 0:
                    rae_hist_entries.append((k, bin_lab, e_abs.tolist()))

    # Distance-like stats without binning (all distances together).
    for mask_name, mask in (("rainy", rainy), ("nonrainy", ~rainy)):
        e_signed = signed_full[mask].ravel()
        e_abs = abs_full[mask].ravel()
        l1_rae = float(np.sum(abs_rae_full[mask]))
        l1_abs_mmph = float(np.sum(abs_mmph_full[mask]))
        s = stats_row(e_signed, e_abs, l1_rae_sum=l1_rae, l1_abs_mmph_sum=l1_abs_mmph)
        overall_rows.append(dict(
            patch_key=key,
            mask_type=mask_name,
            distance_bin_m="ALL",
            l1_abs_mmph_sum_norm_hw=(l1_abs_mmph / float(gt.size)),
            **s,
        ))

    attn_all, J1_all, n_valid = attn_l1_and_J1(A_obs, A_hat, L_km, valid)
    attn_10, J1_10, n_10 = attn_l1_and_J1(A_obs, A_hat, L_km, ge10)
    mean_abs_attn_err_per_km_all, length_weighted_abs_attn_err_per_km_all = abs_attn_error_per_km_metrics(
        A_obs, A_hat, L_km, valid
    )
    max_abs_diff, p95_abs_diff, p99_abs_diff = abs_diff_summary(A_obs, A_hat, valid)
    J1_len1_all = j1_len1(A_obs, A_hat, L_km, valid)
    J1_len1_10 = j1_len1(A_obs, A_hat, L_km, ge10)
    J_atten_all = float(J1_len1_all / float(n_valid)) if n_valid > 0 else 0.0
    J_atten_10 = float(J1_len1_10 / float(n_10)) if n_10 > 0 else 0.0
    denom_L = float(np.sum(np.abs(L_km[valid])))
    if denom_L > 0:
        E_all = float(np.sum(A_obs[valid] - A_hat[valid]) / denom_L)
        E2_all = float(np.sum((A_obs[valid] - A_hat[valid]) ** 2) / denom_L)
    else:
        E_all = 0.0
        E2_all = 0.0

    link_row = dict(
        patch_key=key,
        n_links_valid=n_valid,
        attn_l1_all=attn_all,
        J_atten_all=J_atten_all,
        n_links_ge10km=n_10,
        attn_l1_ge10km=attn_10,
        J_atten_ge10km=J_atten_10,
        max_abs_diff=max_abs_diff,
        p95_abs_diff=p95_abs_diff,
        p99_abs_diff=p99_abs_diff,
        J_atten_x_num_links_all=J1_len1_all,
        J_atten_x_num_links_ge10km=J1_len1_10,
        mean_abs_attn_err_per_km_all=mean_abs_attn_err_per_km_all,
        length_weighted_abs_attn_err_per_km_all=length_weighted_abs_attn_err_per_km_all,
        E_all=E_all,
        E2_all=E2_all,
    )
    if A_obs_virtual.size and A_hat_virtual.size:
        attn_virtual_all, J1_virtual_all, n_valid_virtual = attn_l1_and_J1(
            A_obs_virtual, A_hat_virtual, L_km_virtual, valid_virtual
        )
        attn_virtual_10, J1_virtual_10, n_10_virtual = attn_l1_and_J1(
            A_obs_virtual, A_hat_virtual, L_km_virtual, ge10_virtual
        )
        mean_abs_attn_err_per_km_virtual_all, length_weighted_abs_attn_err_per_km_virtual_all = abs_attn_error_per_km_metrics(
            A_obs_virtual, A_hat_virtual, L_km_virtual, valid_virtual
        )
        max_abs_diff_virtual, p95_abs_diff_virtual, p99_abs_diff_virtual = abs_diff_summary(
            A_obs_virtual, A_hat_virtual, valid_virtual
        )
        J_atten_virtual_all = float(J1_virtual_all / float(n_valid_virtual)) if n_valid_virtual > 0 else 0.0
        J_atten_virtual_10 = float(J1_virtual_10 / float(n_10_virtual)) if n_10_virtual > 0 else 0.0
        link_row.update(
            virtual_attn_available=True,
            n_links_virtual_valid=n_valid_virtual,
            attn_l1_virtual_all=attn_virtual_all,
            J_atten_virtual_all=J_atten_virtual_all,
            n_links_virtual_ge10km=n_10_virtual,
            attn_l1_virtual_ge10km=attn_virtual_10,
            J_atten_virtual_ge10km=J_atten_virtual_10,
            max_abs_diff_virtual=max_abs_diff_virtual,
            p95_abs_diff_virtual=p95_abs_diff_virtual,
            p99_abs_diff_virtual=p99_abs_diff_virtual,
            mean_abs_attn_err_per_km_virtual_all=mean_abs_attn_err_per_km_virtual_all,
            length_weighted_abs_attn_err_per_km_virtual_all=length_weighted_abs_attn_err_per_km_virtual_all,
        )
    else:
        link_row["virtual_attn_available"] = False
    idx_valid = np.where(valid)[0]
    abs_norm_resid_valid = np.abs(A_hat[idx_valid] - A_obs[idx_valid]) / L_km[idx_valid] if idx_valid.size > 0 else np.zeros(0, dtype=np.float64)
    link_metric = dict(
        L1=attn_all,
        J1=J1_all,
        J1_len1=J1_len1_all,
        MEAN_ABS_ATTN_ERR_PER_KM=mean_abs_attn_err_per_km_all,
        LENGTH_WEIGHTED_ABS_ATTN_ERR_PER_KM=length_weighted_abs_attn_err_per_km_all,
        E=E_all,
        E2=E2_all,
        abs_norm_resid_valid=abs_norm_resid_valid.tolist(),
    )

    return dict(
        patch_key=key,
        map_metrics=dict(
            rmse_mmph=rmse_mmph,
            bias_mmph=bias_mmph,
            pearson_corr=pearson_corr,
            n_pixels=int(gt_flat.size),
        ),
        stop_row=stop_row,
        dry_fp_rate=dry_fp_rate,
        dry_fn_rate=dry_fn_rate,
        fp_fn_by_threshold=fp_fn_by_threshold,
        gtbin_counts=gtbin_counts_by_lab,
        gtbin_rel_means=gtbin_rel_means_by_lab,
        gtbin_abs_means=gtbin_abs_means_by_lab,
        objective_pairs=objective_pairs,
        native_objective=native_objective,
        reference_objective_pred=reference_objective_pred,
        reference_objective_gt=reference_objective_gt,
        cov_rows=cov_rows,
        dist_rows_by_k=dist_rows_by_k,
        overall_rows=overall_rows,
        bin_counts_entries=bin_counts_entries,
        bin_counts_all_entries=bin_counts_all_entries,
        medians_r_entries=medians_r_entries,
        medians_n_entries=medians_n_entries,
        jatten_medians_entries=jatten_medians_entries,
        jatten_link_bin_counts_entries=jatten_link_bin_counts_entries,
        p90s_r_entries=p90s_r_entries,
        p90s_n_entries=p90s_n_entries,
        rae_hist_entries=rae_hist_entries,
        link_row=link_row,
        link_metric=link_metric,
        itertrace_entries=iter_entries,
        itertrace_summary=iter_summary,
        itertrace_path=(str(itertrace_path) if has_itertrace else None),
    )


# ----------------------------
# Attenuation / J1 (link-based)
# ----------------------------

def compute_link_terms(est: dict, R_field: np.ndarray) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """
    Returns arrays (A_obs, A_hat, L_km, valid_mask, len_ge10_mask) aligned by link_index.
    """
    from cml_attenuation.itu_r_p_8383 import k_alpha  # type: ignore

    links = est["links"]
    segs: Dict[str, list] = est.get("segments_by_link", {})

    header = est["header"]
    H = int(header["H"]); W = int(header["W"])
    R = np.asarray(R_field, dtype=np.float64)
    if R.shape != (H, W):
        raise ValueError(f"R_field shape {R.shape} != (H,W)=({H},{W}) for attenuation computation.")

    L = len(links)
    A_obs = np.zeros(L, dtype=np.float64)
    A_hat = np.zeros(L, dtype=np.float64)
    L_km = np.zeros(L, dtype=np.float64)

    # gather per-link k, alpha
    k = np.zeros(L, dtype=np.float64)
    a = np.zeros(L, dtype=np.float64)

    # Polarization mapping: different ITU implementations expect different strings.
    # We try a small set of common aliases for horizontal/vertical.
    def pol_aliases(p: str) -> list:
        p = str(p).strip()
        pu = p.upper()
        if pu in ("H", "HOR", "HORIZ", "HORIZONTAL", "HH"):
            return ["H", "h", "horizontal", "HORIZONTAL", "HH"]
        if pu in ("V", "VER", "VERT", "VERTICAL", "VV"):
            return ["V", "v", "vertical", "VERTICAL", "VV"]
        return [p, pu]

    for rec in links:
        li = int(rec["link_index"])
        A_obs[li] = float(rec.get("A_db", rec.get("A", 0.0)))
        freq = float(rec["freq_ghz"])
        pol_raw = rec.get("pol", "")
        last_err = None
        for pol_try in pol_aliases(pol_raw):
            try:
                k_li, a_li = k_alpha(freq, pol_try)
                k[li] = float(k_li)
                a[li] = float(a_li)
                last_err = None
                break
            except Exception as e:
                last_err = e
                continue
        if last_err is not None:
            raise ValueError(f"Unknown polarization: {pol_raw}") from last_err

    # compute A_hat by summing over pixel segments
    for li in range(L):
        seg_list = segs.get(str(li), [])
        if not seg_list:
            continue
        for s in seg_list:
            i = int(s["i"]); j = int(s["j"])
            ds_km = float(s["ds_m"]) / 1000.0
            L_km[li] += ds_km
            r = max(float(R[i, j]), 0.0)
            A_hat[li] += ds_km * k[li] * (r ** a[li])

    valid = L_km > 0
    ge10 = (L_km >= 10.0) & valid
    return A_obs, A_hat, L_km, valid, ge10


def jatten_attributed_pixel_map(
    est: dict,
    A_obs: np.ndarray,
    A_hat: np.ndarray,
    L_km: np.ndarray,
    valid: np.ndarray,
) -> np.ndarray:
    """
    Build per-pixel attributed J_atten contribution map.

    Link contribution (normalized J_atten convention):
      c_l = ((A_hat_l - A_obs_l)^2 / L_l) / #valid_links
    Pixel attribution (for each segment s on link l):
      c_l * (ds_s / L_l)
    Summed over links crossing that pixel.
    """
    header = est["header"]
    H = int(header["H"])
    W = int(header["W"])
    out = np.zeros((H, W), dtype=np.float64)
    segs_by_link = est.get("segments_by_link", {}) or {}

    n_links = int(min(len(A_obs), len(A_hat), len(L_km), len(valid)))
    n_valid = int(np.sum(valid[:n_links])) if n_links > 0 else 0
    den_valid = float(max(1, n_valid))
    for li in range(n_links):
        if not bool(valid[li]):
            continue
        L = float(L_km[li])
        if L <= 0.0 or (not np.isfinite(L)):
            continue
        diff = float(A_hat[li] - A_obs[li])
        c_link = ((diff * diff) / L) / den_valid
        segs = segs_by_link.get(str(li), [])
        if not segs:
            continue
        for s in segs:
            i = int(s.get("i", -1))
            j = int(s.get("j", -1))
            if i < 0 or i >= H or j < 0 or j >= W:
                continue
            ds_km = float(s.get("ds_m", 0.0)) / 1000.0
            if ds_km <= 0.0 or (not np.isfinite(ds_km)):
                continue
            out[i, j] += c_link * (ds_km / L)
    return out


def j1_len1(A_obs: np.ndarray, A_hat: np.ndarray, L_km: np.ndarray, mask: np.ndarray) -> float:
    """
    J1 variant with first-power link-length denominator:
      sum_links ((A_hat - A_obs)^2 / L_km)
    """
    idx = np.where(mask)[0]
    if idx.size == 0:
        return 0.0
    diff = A_hat[idx] - A_obs[idx]
    return float(np.sum((diff ** 2) / L_km[idx]))


# ----------------------------
# Excel writing
# ----------------------------

def safe_sheet_name(name: str) -> str:
    # Excel sheet max length 31; remove invalid chars
    bad = set(r'[]:*?/\\')
    s = "".join(ch for ch in name if ch not in bad).strip()
    if not s:
        s = "Sheet"
    return s[:31]


def safe_path_token(name: str) -> str:
    out = []
    for ch in str(name):
        if ch.isalnum() or ch in ("-", "_"):
            out.append(ch)
        else:
            out.append("_")
    s = "".join(out).strip("_")
    return s or "item"


def unique_sheet_name(name: str, used: set) -> str:
    base = safe_sheet_name(name)
    if base not in used:
        used.add(base)
        return base
    idx = 2
    while True:
        suffix = f"_{idx}"
        cand = f"{base[:31 - len(suffix)]}{suffix}"
        if cand not in used:
            used.add(cand)
            return cand
        idx += 1


def write_workbook(
    path: Path,
    sheets: Dict[str, List[Dict[str, Any]]],
    header_comments: Optional[Dict[str, Dict[str, str]]] = None,
):
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.comments import Comment  # type: ignore
        from openpyxl.styles import PatternFill, Border, Side  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl required to write Excel (pip install openpyxl).") from e

    wb = Workbook()
    # remove default
    wb.remove(wb.active)
    used_titles: set = set()
    patch_row_fills = [
        PatternFill(fill_type="solid", fgColor="FFF3CD"),
        PatternFill(fill_type="solid", fgColor="D1ECF1"),
        PatternFill(fill_type="solid", fgColor="D4EDDA"),
        PatternFill(fill_type="solid", fgColor="F8D7DA"),
        PatternFill(fill_type="solid", fgColor="E2E3E5"),
        PatternFill(fill_type="solid", fgColor="FDE2E4"),
        PatternFill(fill_type="solid", fgColor="E2F0CB"),
        PatternFill(fill_type="solid", fgColor="CDE7F0"),
        PatternFill(fill_type="solid", fgColor="F9E2AE"),
        PatternFill(fill_type="solid", fgColor="E4C1F9"),
    ]
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def ordered_columns_for_sheet(sheet_name: str, cols: List[str]) -> List[str]:
        def _ordered(preferred: List[str]) -> List[str]:
            rank = {name: i for i, name in enumerate(preferred)}
            return sorted(cols, key=lambda c: (rank.get(c, 10_000), cols.index(c)))

        if sheet_name.startswith("LinkStats_"):
            return _ordered([
                "patch_key",
                "n_links_valid",
                "attn_l1_all",
                "J_atten_all",
                "n_links_ge10km",
                "attn_l1_ge10km",
                "J_atten_ge10km",
                "max_abs_diff",
                "p95_abs_diff",
                "p99_abs_diff",
                "J_atten_x_num_links_all",
                "J_atten_x_num_links_ge10km",
                "mean_abs_attn_err_per_km_all",
                "length_weighted_abs_attn_err_per_km_all",
                "E_all",
                "E2_all",
                "J_atten_IDW",
                "J_atten_ILDW",
                "J_atten_ALG_over_IDW",
                "J_atten_ALG_over_ILDW",
            ])

        if sheet_name == "AttenuationErrorPerKm_BySolver":
            return _ordered([
                "solver",
                "mean_abs_attn_err_per_km_mean",
                "mean_abs_attn_err_per_km_std",
                "length_weighted_abs_attn_err_per_km_mean",
                "length_weighted_abs_attn_err_per_km_std",
                "n_patches",
                "metric",
                "definition",
            ])

        if sheet_name == "PatchMapMetrics_ByPatch":
            return _ordered([
                "patch_key",
                "solver",
                "solver_name",
                "rmse_mmph",
                "bias_mmph",
                "pearson_corr",
                "n_pixels",
            ])

        if sheet_name == "PatchMapMetrics_BySolver":
            return _ordered([
                "solver",
                "rmse_mmph_mean",
                "rmse_mmph_std",
                "rmse_mmph_min",
                "rmse_mmph_max",
                "bias_mmph_mean",
                "bias_mmph_std",
                "bias_mmph_min",
                "bias_mmph_max",
                "pearson_corr_mean",
                "pearson_corr_std",
                "pearson_corr_min",
                "pearson_corr_max",
                "n_patches",
            ])

        if (
            sheet_name.startswith("CoverageStats_")
            or sheet_name.startswith("DistanceStats")
            or sheet_name.startswith("OverallStats_GTvs")
        ):
            return _ordered([
                "patch_key",
                "mask_type",
                "coverage_bin",
                "distance_bin_m",
                "n_pixels",
                "mean_signed",
                "median_signed",
                "mean_abs",
                "std_abs",
                "median_abs",
                "p90_abs",
                "p95_abs",
                "p99_abs",
                "linf_abs",
                "l1_rae_sum",
                "l1_abs_mmph_sum",
                "l1_abs_mmph_sum_norm_hw",
                "mean_abs_IDW",
                "mean_abs_ILDW",
                "mean_abs_over_IDW",
                "mean_abs_over_ILDW",
                "std_abs_IDW",
                "std_abs_ILDW",
                "std_abs_over_IDW",
                "std_abs_over_ILDW",
                "median_abs_IDW",
                "median_abs_ILDW",
                "median_abs_over_IDW",
                "median_abs_over_ILDW",
                "p90_abs_IDW",
                "p90_abs_ILDW",
                "p90_abs_over_IDW",
                "p90_abs_over_ILDW",
                "p95_abs_IDW",
                "p95_abs_ILDW",
                "p95_abs_over_IDW",
                "p95_abs_over_ILDW",
                "p99_abs_IDW",
                "p99_abs_ILDW",
                "p99_abs_over_IDW",
                "p99_abs_over_ILDW",
            ])

        if sheet_name == "OverallStats_BySolver":
            return _ordered([
                "patch_key",
                "solver",
                "solver_name",
                "module",
                "mask_type",
                "distance_bin_m",
                "n_pixels",
                "mean_signed",
                "median_signed",
                "mean_abs",
                "std_abs",
                "median_abs",
                "p90_abs",
                "p95_abs",
                "p99_abs",
                "linf_abs",
                "l1_rae_sum",
                "l1_abs_mmph_sum",
                "l1_abs_mmph_sum_norm_hw",
                "mean_abs_IDW",
                "mean_abs_ILDW",
                "mean_abs_over_IDW",
                "mean_abs_over_ILDW",
                "std_abs_IDW",
                "std_abs_ILDW",
                "std_abs_over_IDW",
                "std_abs_over_ILDW",
                "median_abs_IDW",
                "median_abs_ILDW",
                "median_abs_over_IDW",
                "median_abs_over_ILDW",
                "p90_abs_IDW",
                "p90_abs_ILDW",
                "p90_abs_over_IDW",
                "p90_abs_over_ILDW",
                "p95_abs_IDW",
                "p95_abs_ILDW",
                "p95_abs_over_IDW",
                "p95_abs_over_ILDW",
                "p99_abs_IDW",
                "p99_abs_ILDW",
                "p99_abs_over_IDW",
                "p99_abs_over_ILDW",
            ])

        return cols

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=unique_sheet_name(sheet_name, used_titles))
        if not rows:
            continue
        # header (union of keys in first-seen order across all rows)
        cols: List[str] = []
        seen: set = set()
        for r in rows:
            for c in r.keys():
                if c not in seen:
                    seen.add(c)
                    cols.append(c)
        cols = ordered_columns_for_sheet(sheet_name, cols)
        if sheet_name == "Objective_NativeBySolver":
            has_eps_value = any(
                (("eps" in r) and (r.get("eps") is not None) and (str(r.get("eps")) != ""))
                for r in rows
            )
            preferred = [
                "patch_key",
                "definition",
                "solver_name",
                "module",
                "objective_available",
                "objective_unavailable_reason",
                "objective_scaling",
                "objective_formula",
                "w_atten",
                "w_1d",
                "w_total",
                "w_2d",
                "n_valid_links",
                "n_pixels",
                "solver",
                "J_weighted_sum",
                "J_atten",
                "J_1d",
                "J_total",
                "J_2d",
                "weighted_J_atten",
                "weighted_J_1d",
                "weighted_J_total",
                "weighted_J_2d",
            ]
            if has_eps_value:
                preferred.insert(preferred.index("n_valid_links"), "eps")
            rank = {name: i for i, name in enumerate(preferred)}
            cols = [c for c in cols if (c != "eps" or has_eps_value)]
            cols = sorted(cols, key=lambda c: (rank.get(c, 10_000), cols.index(c)))
        if sheet_name == "StoppingInfo":
            preferred = [
                "patch_key",
                "solver",
                "solver_name",
                "has_optinfo",
                "stop_reason",
                "reason_bucket",
                "success",
                "status",
                "message",
                "nit",
                "nfev",
                "njev",
                "proj_grad_inf",
                "rel_decrease",
                "ftol",
                "gtol",
                "ftol_met",
                "gtol_met",
                "line_search_failed",
                "maxiter_reached",
                "optinfo_json",
                "definition_term",
                "definition",
            ]
            rank = {name: i for i, name in enumerate(preferred)}
            cols = sorted(cols, key=lambda c: (rank.get(c, 10_000), cols.index(c)))
        ws.append(cols)
        comments_for_sheet = (header_comments or {}).get(sheet_name, {})
        term_fill = PatternFill(fill_type="solid", fgColor="FFFDEB")
        for cidx, col_name in enumerate(cols, start=1):
            if col_name in comments_for_sheet:
                cell = ws.cell(row=1, column=cidx)
                cell.comment = Comment(comments_for_sheet[col_name], "batch_analyze_multi")
                cell.fill = term_fill
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        # Color rows by patch_key; reuse colors cyclically when palette runs out.
        if "patch_key" in cols:
            patch_col = cols.index("patch_key") + 1
            patch_to_fill: Dict[str, Any] = {}
            fill_idx = 0
            for ridx in range(2, ws.max_row + 1):
                patch_val = ws.cell(row=ridx, column=patch_col).value
                if patch_val is None:
                    continue
                patch_key = str(patch_val)
                if patch_key in {"AVERAGE", "DEFINITION"}:
                    continue
                if patch_key not in patch_to_fill:
                    patch_to_fill[patch_key] = patch_row_fills[fill_idx % len(patch_row_fills)]
                    fill_idx += 1
                row_fill = patch_to_fill[patch_key]
                for cidx in range(1, len(cols) + 1):
                    ws.cell(row=ridx, column=cidx).fill = row_fill
        # Add thin borders to all populated cells.
        for ridx in range(1, ws.max_row + 1):
            for cidx in range(1, len(cols) + 1):
                cell = ws.cell(row=ridx, column=cidx)
                if cell.value is None or cell.value == "":
                    continue
                cell.border = thin_border
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ----------------------------
# Plot: IQR of per-patch medians by distance bin
# ----------------------------

def compute_iqr_profile(per_patch_medians: Dict[str, Dict[str, List[float]]],
                        dist_labels: List[str]) -> Dict[str, Dict[str, Tuple[float,float,float]]]:
    """
    per_patch_medians[method][bin_label] = list of medians (one per patch that had pixels in that bin)
    returns summary[method][bin_label] = (p25, p50, p75)
    """
    out: Dict[str, Dict[str, Tuple[float,float,float]]] = {}
    for method, by_bin in per_patch_medians.items():
        out[method] = {}
        for lab in dist_labels:
            vals = np.array(by_bin.get(lab, []), dtype=np.float64)
            if vals.size == 0:
                out[method][lab] = (0.0, 0.0, 0.0)
            else:
                out[method][lab] = (
                    float(np.percentile(vals, 25)),
                    float(np.percentile(vals, 50)),
                    float(np.percentile(vals, 75)),
                )
    return out


def compute_p90_profile(per_patch_p90s: Dict[str, Dict[str, List[float]]],
                        dist_labels: List[str]) -> Dict[str, Dict[str, Tuple[float,float,float]]]:
    """
    per_patch_p90s[method][bin_label] = list of p90s (one per patch that had pixels in that bin)
    returns summary[method][bin_label] = (p25, p50, p75)
    """
    out: Dict[str, Dict[str, Tuple[float,float,float]]] = {}
    for method, by_bin in per_patch_p90s.items():
        out[method] = {}
        for lab in dist_labels:
            vals = np.array(by_bin.get(lab, []), dtype=np.float64)
            if vals.size == 0:
                out[method][lab] = (0.0, 0.0, 0.0)
            else:
                out[method][lab] = (
                    float(np.percentile(vals, 25)),
                    float(np.percentile(vals, 50)),
                    float(np.percentile(vals, 75)),
                )
    return out


def compute_relative_iqr_profile(
    summary: Dict[str, Dict[str, Tuple[float, float, float]]],
    *,
    idw_label: str,
    dist_labels: List[str],
) -> Dict[str, Dict[str, Tuple[float, float, float]]]:
    """
    Divide each method's (p25,p50,p75) by IDW's p50 (median-of-medians) per bin.
    """
    if idw_label not in summary:
        return {}
    out: Dict[str, Dict[str, Tuple[float, float, float]]] = {}
    for method, by_bin in summary.items():
        out[method] = {}
        for lab in dist_labels:
            p25, p50, p75 = by_bin.get(lab, (0.0, 0.0, 0.0))
            idw_med = summary[idw_label].get(lab, (0.0, 0.0, 0.0))[1]
            if idw_med == 0.0:
                out[method][lab] = (0.0, 0.0, 0.0)
            else:
                out[method][lab] = (p25 / idw_med, p50 / idw_med, p75 / idw_med)
    return out


def compute_relative_distribution_profile(
    per_patch_values: Dict[str, Dict[str, List[float]]],
    *,
    baseline_label: str,
    dist_labels: List[str],
) -> Dict[str, Dict[str, List[float]]]:
    """
    Divide each method's per-patch values by the baseline method's p50 in the same bin.
    """
    if baseline_label not in per_patch_values:
        return {}
    out: Dict[str, Dict[str, List[float]]] = {}
    baseline_bins = per_patch_values[baseline_label]
    for method, by_bin in per_patch_values.items():
        out[method] = {}
        for lab in dist_labels:
            baseline_vals = np.asarray(baseline_bins.get(lab, []), dtype=np.float64)
            if baseline_vals.size == 0:
                out[method][lab] = []
                continue
            baseline_med = float(np.percentile(baseline_vals, 50))
            if baseline_med == 0.0:
                out[method][lab] = []
                continue
            vals = np.asarray(by_bin.get(lab, []), dtype=np.float64)
            if vals.size == 0:
                out[method][lab] = []
                continue
            good = np.isfinite(vals)
            out[method][lab] = [float(v / baseline_med) for v in vals[good]]
    return out


def plot_iqr_bars(out_png: Path, title: str,
                  per_patch_values: Dict[str, Dict[str, List[float]]],
                  dist_labels: List[str],
                  method_order: List[str],
                  *, y_max: Optional[float] = None, dpi: int = 150, bin_spacing: float = 1.0,
                  tick_labels: Optional[List[str]] = None,
                  x_label: Optional[str] = None,
                  y_label: Optional[str] = None,
                  footnote: Optional[str] = None,
                  show_iqr: bool = True):
    import matplotlib.pyplot as plt  # type: ignore

    n_bins = len(dist_labels)
    n_methods = len(method_order)
    x = np.arange(n_bins) * float(bin_spacing)

    # offsets
    width = 0.8 / max(1, n_methods)
    offsets = (np.arange(n_methods) - (n_methods - 1) / 2.0) * width

    fig_w = max(6, n_bins * 1.2 * float(bin_spacing))
    fig, ax = plt.subplots(figsize=(fig_w, 4.5), dpi=dpi)
    color_cycle = plt.rcParams["axes.prop_cycle"].by_key().get("color", [])
    colors = {
        m: color_cycle[i % max(1, len(color_cycle))]
        for i, m in enumerate(method_order)
    }

    for mi, m in enumerate(method_order):
        off = offsets[mi]
        color = colors.get(m, None)
        p50: List[float] = []
        for bi, lab in enumerate(dist_labels):
            vals = np.asarray(per_patch_values.get(m, {}).get(lab, []), dtype=np.float64)
            vals = vals[np.isfinite(vals)]
            if vals.size == 0:
                p50.append(0.0)
                continue
            p50.append(float(np.percentile(vals, 50)))
            if show_iqr:
                ax.boxplot(
                    [vals],
                    positions=[x[bi] + off],
                    widths=width * 0.8,
                    patch_artist=True,
                    showfliers=False,
                    manage_ticks=False,
                    boxprops={"facecolor": color, "alpha": 0.25, "edgecolor": color, "linewidth": 1.2},
                    whiskerprops={"color": color, "linewidth": 1.2},
                    capprops={"color": color, "linewidth": 1.2},
                    medianprops={"color": color, "linewidth": 1.4},
                )

        # medians as points (no connecting line)
        ax.plot(x + off, p50, marker="o", linestyle="None", color=color, label=m)

    ax.set_xticks(x)
    if tick_labels is None:
        tick_labels = dist_labels
    has_multiline = any("\n" in str(lab) for lab in tick_labels)
    if has_multiline:
        ax.set_xticklabels(tick_labels, rotation=20, ha="right", fontsize=8)
        fig.subplots_adjust(bottom=0.28)
        ax.set_xlabel(x_label or "Distance bin (m)\nSecond line: avg pixels [avg-std, avg+std]")
    else:
        ax.set_xticklabels(tick_labels, rotation=0)
    ax.set_ylabel(y_label or "Per-patch median error")
    ax.set_title(title)
    ax.legend(loc="best")
    ax.grid(True, axis="y", linestyle=":", linewidth=0.7)

    if y_max is not None:
        ax.set_ylim(0, y_max)
    else:
        ax.set_ylim(bottom=0)

    if footnote:
        fig.text(0.5, 0.01, str(footnote), ha="center", fontsize=8)
        fig.tight_layout(rect=(0, 0.05, 1, 1))
    else:
        fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png)
    plt.close(fig)


def mask_to_nan(arr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    out = arr.astype(np.float32).copy()
    out[~mask] = np.nan
    return out


def save_png_2x2(
    out_png: Path,
    a: np.ndarray,
    b: np.ndarray,
    c: np.ndarray,
    d: np.ndarray,
    *,
    titles: Tuple[str, str, str, str],
    suptitle: str,
    cmaps: Tuple[str, str, str, str],
    vlims: Tuple[
        Tuple[float, float],
        Tuple[float, float],
        Tuple[float, float],
        Tuple[float, float],
    ],
    dpi: int = 150,
    show: bool = False,
) -> None:
    import matplotlib.pyplot as plt  # type: ignore

    fig, axes = plt.subplots(2, 2, figsize=(12, 10), dpi=dpi)
    axs = axes.ravel()
    arrays = [a, b, c, d]
    for ax, arr, t, cmap, (vmin, vmax) in zip(axs, arrays, titles, cmaps, vlims):
        im = ax.imshow(arr, cmap=cmap, vmin=vmin, vmax=vmax)
        ax.set_title(t)
        ax.set_xticks([])
        ax.set_yticks([])
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

    fig.suptitle(suptitle, fontsize=12)
    fig.tight_layout(rect=(0, 0.02, 1, 0.96))
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png)
    if show:
        plt.show()
    plt.close(fig)


def render_patch_error_maps_from_job(
    job: Dict[str, Any],
    *,
    img_dir: Path,
    render_cfg: Dict[str, Any],
) -> None:
    gt = load_npz_first_key(Path(str(job["gt_path"])), list(job["gt_key_pref"])).astype(np.float64)
    pred = load_npz_first_key(Path(str(job["sol_path"])), list(job["sol_key_pref"])).astype(np.float64)
    if gt.shape != pred.shape:
        raise ValueError(f"Shape mismatch for patch plot {job['patch_key']}: GT {gt.shape} vs SOL {pred.shape}")

    key = str(job["patch_key"])
    solver_label = str(job["solver_label"])
    thr = float(render_cfg["threshold_mmph"])
    plot_dpi = int(render_cfg["dpi"])
    cmap_gt = str(render_cfg["cmap_gt"])
    cmap_sol = str(render_cfg["cmap_sol"])
    cmap_diff = str(render_cfg["cmap_diff"])
    cmap_abs = str(render_cfg["cmap_abs_diff"])
    cmap_rel = str(render_cfg["cmap_rel"])
    cmap_abs_rel = str(render_cfg["cmap_abs_rel"])

    rainy = gt >= thr
    rel_full = np.full_like(gt, np.nan, dtype=np.float64)
    if np.any(rainy):
        rel_full[rainy] = (gt[rainy] - pred[rainy]) / np.where(gt[rainy] == 0.0, 1.0, gt[rainy])
    abs_rel_full = np.abs(rel_full)
    diff_full = pred - gt
    abs_diff_full = np.abs(diff_full)

    finite_gt_sol = np.concatenate([gt[np.isfinite(gt)], pred[np.isfinite(pred)]])
    rmax = float(np.max(finite_gt_sol)) if finite_gt_sol.size > 0 else 1.0
    rmax = max(rmax, 1e-9)
    rel_max = float(np.nanmax(np.abs(rel_full))) if np.any(np.isfinite(rel_full)) else 1.0
    rel_abs_max = float(np.nanmax(abs_rel_full)) if np.any(np.isfinite(abs_rel_full)) else 1.0
    dmax = float(np.nanmax(np.abs(diff_full[~rainy]))) if np.any(~rainy) else 1.0
    rel_max = max(rel_max, 1e-9)
    rel_abs_max = max(rel_abs_max, 1e-9)
    dmax = max(dmax, 1e-9)

    patch_plot_dir = img_dir / "patch_error_maps" / safe_path_token(solver_label)
    rainy_png = patch_plot_dir / f"{safe_path_token(key)}_rainy.png"
    save_png_2x2(
        rainy_png,
        mask_to_nan(gt, rainy),
        mask_to_nan(pred, rainy),
        mask_to_nan(rel_full, rainy),
        mask_to_nan(abs_rel_full, rainy),
        titles=("GT (rainy)", "SOL (rainy)", "(GT-SOL)/GT", "|GT-SOL|/GT"),
        suptitle=f"{key} | rainy: GT>= {thr} mm/h",
        cmaps=(cmap_gt, cmap_sol, cmap_rel, cmap_abs_rel),
        vlims=((0.0, rmax), (0.0, rmax), (-rel_max, rel_max), (0.0, rel_abs_max)),
        dpi=plot_dpi,
        show=False,
    )

    nonrainy_png = patch_plot_dir / f"{safe_path_token(key)}_nonrainy.png"
    save_png_2x2(
        nonrainy_png,
        mask_to_nan(gt, ~rainy),
        mask_to_nan(pred, ~rainy),
        mask_to_nan(diff_full, ~rainy),
        mask_to_nan(abs_diff_full, ~rainy),
        titles=("GT (non-rainy)", "SOL (non-rainy)", "SOL-GT", "|SOL-GT|"),
        suptitle=f"{key} | non-rainy: GT< {thr} mm/h",
        cmaps=(cmap_gt, cmap_sol, cmap_diff, cmap_abs),
        vlims=((0.0, rmax), (0.0, rmax), (-dmax, dmax), (0.0, rmax)),
        dpi=plot_dpi,
        show=False,
    )


def pick_biggest_est_patch(est_by_key: Dict[str, Path]) -> Tuple[Optional[str], Optional[Path], float]:
    """
    Return (patch_key, est_path, area_km2) for the largest est patch by physical area.
    Area = H * W * pixel_size_m^2.
    """
    best_key: Optional[str] = None
    best_path: Optional[Path] = None
    best_area_km2 = -1.0
    for key, path in est_by_key.items():
        try:
            est = load_est_payload(path)
            hdr = est.get("header", {})
            H = int(hdr["H"])
            W = int(hdr["W"])
            pix = float(hdr.get("pixel_size_m", 125.0))
            area_km2 = float(H) * float(W) * float(pix) * float(pix) / 1_000_000.0
        except Exception:
            continue
        if area_km2 > best_area_km2:
            best_area_km2 = area_km2
            best_key = key
            best_path = path
    if best_key is None or best_path is None:
        return None, None, 0.0
    return best_key, best_path, float(best_area_km2)


def save_biggest_patch_distance_bin_maps(
    *,
    est_path: Path,
    patch_key: str,
    out_dir: Path,
    bin_edges_m: Sequence[float],
    dist_method: str,
    sample_spacing_m: float,
    k_query_samples: int,
    chunk_size: int,
    max_samples_per_link: int,
    max_candidates: int,
    dpi: int = 150,
    show: bool = False,
) -> None:
    """
    Produce 2 images for the largest patch:
      - k=2 distance-to-kth-closest-link bins + link overlay
      - k=3 distance-to-kth-closest-link bins + link overlay
    """
    import matplotlib.pyplot as plt  # type: ignore
    from matplotlib.colors import ListedColormap  # type: ignore

    est = load_est_payload(est_path)
    hdr = est["header"]
    H = int(hdr["H"])
    W = int(hdr["W"])
    pix = float(hdr.get("pixel_size_m", 125.0))
    area_km2 = float(H) * float(W) * float(pix) * float(pix) / 1_000_000.0

    k_targets = [2, 3]
    if dist_method == "sampled_points":
        dk_maps, _ = compute_dk_maps_sampled_points(
            est,
            k_targets,
            sample_spacing_m=sample_spacing_m,
            k_query_samples=k_query_samples,
            chunk_size=chunk_size,
            max_samples_per_link=max_samples_per_link,
            debug_label=patch_key,
        )
    else:
        dk_maps, _ = compute_dk_maps(est, k_targets, max_candidates=max_candidates)

    bins = parse_bins(list(bin_edges_m))
    labels = [b[2] for b in bins]
    edges = np.asarray(list(bin_edges_m), dtype=np.float64)
    if edges.size > 0:
        edges = np.unique(edges[np.isfinite(edges)])
    n_bins = len(labels)

    # Discrete colormap (plus masked gray for inf/invalid).
    cmap_vals = plt.cm.plasma(np.linspace(0.08, 0.95, max(1, n_bins)))
    cmap = ListedColormap(cmap_vals)
    cmap.set_bad(color="#d9d9d9")

    # Link segments in local patch frame.
    links = est.get("links", []) or []
    x0 = np.asarray([float(L["x0_m"]) for L in links], dtype=np.float64) if links else np.zeros(0, dtype=np.float64)
    y0 = np.asarray([float(L["y0_m"]) for L in links], dtype=np.float64) if links else np.zeros(0, dtype=np.float64)
    x1 = np.asarray([float(L["x1_m"]) for L in links], dtype=np.float64) if links else np.zeros(0, dtype=np.float64)
    y1 = np.asarray([float(L["y1_m"]) for L in links], dtype=np.float64) if links else np.zeros(0, dtype=np.float64)

    out_dir.mkdir(parents=True, exist_ok=True)
    x_span_km = (float(W) * pix) / 1000.0
    y_span_km = (float(H) * pix) / 1000.0

    for k in k_targets:
        d_map = np.asarray(dk_maps.get(k, np.full((H, W), np.inf, dtype=np.float64)), dtype=np.float64)
        finite = np.isfinite(d_map)
        if edges.size == 0:
            bin_idx = np.zeros_like(d_map, dtype=np.int32)
        else:
            bin_idx = np.digitize(d_map, edges, right=True).astype(np.int32)
        bin_plot = np.ma.masked_where(~finite, bin_idx)

        fig, ax = plt.subplots(figsize=(9.0, 7.2), dpi=dpi)
        im = ax.imshow(
            bin_plot,
            cmap=cmap,
            vmin=-0.5,
            vmax=max(0, n_bins - 1) + 0.5,
            origin="upper",
            extent=(0.0, float(W) * pix, float(H) * pix, 0.0),
            interpolation="nearest",
            aspect="equal",
        )
        if x0.size > 0:
            for i in range(x0.size):
                ax.plot([x0[i], x1[i]], [y0[i], y1[i]], color="black", linewidth=0.55, alpha=0.55)

        cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        cbar.set_ticks(np.arange(n_bins))
        cbar.set_ticklabels(labels)
        cbar.ax.set_ylabel(f"d{k} distance bin (m)")

        ax.set_title(
            f"Largest patch distance bins (k={k}) with links\n"
            f"{patch_key} | size={x_span_km:.1f}x{y_span_km:.1f} km | area={area_km2:.1f} km^2 | links={x0.size}"
        )
        ax.set_xlabel("x_local (m)")
        ax.set_ylabel("y_local (m)")
        ax.grid(False)

        out_png = out_dir / f"largest_patch_distance_bins_k{k}.png"
        fig.tight_layout()
        fig.savefig(out_png, dpi=dpi)
        if show:
            plt.show()
        plt.close(fig)
        print(f"[largest_patch_bins] wrote {out_png}")


def plot_ratio_iqr(
    out_png: Path,
    *,
    title: str,
    entries: List[Tuple[str, str, List[float]]],
    dpi: int = 150,
) -> None:
    """
    entries: list of (solver_label, x_label, values_per_patch)
    """
    import matplotlib.pyplot as plt  # type: ignore

    if not entries:
        return

    x = np.arange(len(entries))
    fig_w = max(8, len(entries) * 0.8)
    fig, ax = plt.subplots(figsize=(fig_w, 4.5), dpi=dpi)

    # color by solver
    solver_labels = []
    for solver, _, _ in entries:
        if solver not in solver_labels:
            solver_labels.append(solver)
    color_cycle = plt.rcParams["axes.prop_cycle"].by_key().get("color", [])
    colors = {s: color_cycle[i % max(1, len(color_cycle))] for i, s in enumerate(solver_labels)}

    p50s = []
    for _, _, vals in entries:
        if vals:
            arr = np.array(vals, dtype=np.float64)
            p50s.append(float(np.percentile(arr, 50)))
        else:
            p50s.append(0.0)

    # plot
    for i, (solver, _, vals) in enumerate(entries):
        c = colors.get(solver, None)
        arr = np.asarray(vals, dtype=np.float64)
        arr = arr[np.isfinite(arr)]
        if arr.size > 0:
            ax.boxplot(
                [arr],
                positions=[x[i]],
                widths=0.5,
                patch_artist=True,
                showfliers=False,
                manage_ticks=False,
                boxprops={"facecolor": c, "alpha": 0.25, "edgecolor": c, "linewidth": 1.2},
                whiskerprops={"color": c, "linewidth": 1.2},
                capprops={"color": c, "linewidth": 1.2},
                medianprops={"color": c, "linewidth": 1.4},
            )
        ax.plot(x[i], p50s[i], marker="o", linestyle="None", color=c)

    ax.set_xticks(x)
    ax.set_xticklabels([lab for _, lab, _ in entries], rotation=30, ha="right", fontsize=8)
    ax.set_ylabel("Per-patch ratio")
    ax.set_title(title)
    ax.grid(True, axis="y", linestyle=":", linewidth=0.7)

    # legend
    handles = [plt.Line2D([0], [0], marker="o", color="w", markerfacecolor=colors[s], label=s) for s in solver_labels]
    ax.legend(handles=handles, loc="best")

    fig.text(
        0.5,
        0.01,
        "L1 = sum(|A_hat-A_obs|); J1 ratio uses J1_len1 = sum((A_hat-A_obs)^2/L_km); "
        "E = sum(A_obs-A_hat)/sum(|L_km|); "
        "E2 = sum((A_obs-A_hat)^2)/sum(|L_km|) over all links",
        ha="center",
        fontsize=8,
    )

    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    fig.savefig(out_png)
    plt.close(fig)


def plot_fp_fn_vs_threshold(
    out_png: Path,
    *,
    title: str,
    rows: List[Dict[str, Any]],
    dpi: int = 150,
) -> None:
    import matplotlib.pyplot as plt  # type: ignore
    from matplotlib.ticker import MultipleLocator, FormatStrFormatter  # type: ignore

    if not rows:
        return

    by_solver: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        by_solver.setdefault(str(r["solver"]), []).append(r)

    fig, ax = plt.subplots(figsize=(9.8, 5.2), dpi=dpi)

    for solver, vals in sorted(by_solver.items()):
        vals_sorted = sorted(vals, key=lambda x: float(x.get("threshold_mmph", 0.0)))
        xs = np.array([float(v["threshold_mmph"]) for v in vals_sorted], dtype=np.float64)
        fp = np.array([float(v.get("fp_rate_dry_mean", 0.0)) for v in vals_sorted], dtype=np.float64)
        fn = np.array([float(v.get("fn_rate_wet_mean", 0.0)) for v in vals_sorted], dtype=np.float64)
        fp_std = np.array([float(v.get("fp_rate_dry_std", 0.0)) for v in vals_sorted], dtype=np.float64)
        fn_std = np.array([float(v.get("fn_rate_wet_std", 0.0)) for v in vals_sorted], dtype=np.float64)
        fp_lo = np.clip(fp - fp_std, 0.0, 1.0)
        fp_hi = np.clip(fp + fp_std, 0.0, 1.0)
        fn_lo = np.clip(fn - fn_std, 0.0, 1.0)
        fn_hi = np.clip(fn + fn_std, 0.0, 1.0)
        marker = "o" if xs.size == 1 else None
        plot_kwargs = dict(linewidth=1.8, marker=marker, markersize=6, markeredgewidth=1.0, zorder=3, clip_on=False)
        (l_fp,) = ax.plot(xs, fp, linestyle="-", label=f"{solver} FPR", **plot_kwargs)
        (l_fn,) = ax.plot(xs, fn, linestyle="--", label=f"{solver} FNR", **plot_kwargs)
        ax.fill_between(xs, fp_lo, fp_hi, color=l_fp.get_color(), alpha=0.12, linewidth=0.0)
        ax.fill_between(xs, fn_lo, fn_hi, color=l_fn.get_color(), alpha=0.08, linewidth=0.0)

    ax.set_xlabel("Threshold (mm/h)")
    ax.set_ylabel("Rate")
    ax.grid(True, axis="y", linestyle=":", linewidth=0.7)
    ax.set_ylim(bottom=-0.02)
    ax.xaxis.set_major_locator(MultipleLocator(0.1))
    ax.xaxis.set_major_formatter(FormatStrFormatter("%.1f"))
    ax.tick_params(axis="x", labelrotation=90, labelsize=7)

    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(loc="best", fontsize=8)
    fig.suptitle(title)
    fig.text(
        0.5,
        0.01,
        "Wet is the positive class. Solid=FPR=FP/GT_dry_count; dashed=FNR=FN/GT_wet_count. "
        "Shaded band shows mean±std across patches.",
        ha="center",
        fontsize=8,
    )

    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout(rect=(0, 0.04, 1, 0.95))
    fig.savefig(out_png)
    plt.close(fig)


def plot_j_behavior(
    out_png: Path,
    *,
    title: str,
    iterations: Sequence[Dict[str, Any]],
    dpi: int = 150,
):
    if not iterations:
        return
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except Exception:
        return

    xs = [int(it.get("iter", i + 1)) for i, it in enumerate(iterations)]

    def _series(key: str) -> List[float]:
        out: List[float] = []
        for it in iterations:
            v = it.get(key, None)
            if v is None:
                out.append(float("nan"))
            else:
                try:
                    out.append(float(v))
                except Exception:
                    out.append(float("nan"))
        return out

    y_total = _series("J_weighted_sum")
    if all(np.isnan(v) for v in y_total):
        y_total = _series("J_native_total")
    y_atten = _series("J_atten")
    y_1d = _series("J_1d")
    y_total_term = _series("J_total")
    y_2d = _series("J_2d")

    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(8.5, 4.8), dpi=dpi)
    ax.plot(xs, y_total, marker="o", markersize=2.0, linewidth=1.0, label="J_weighted_sum")
    ax.plot(xs, y_atten, marker="o", markersize=2.0, linewidth=0.9, label="J_atten")
    ax.plot(xs, y_1d, marker="o", markersize=2.0, linewidth=0.9, label="J_1d")
    ax.plot(xs, y_total_term, marker="o", markersize=2.0, linewidth=0.9, label="J_total")
    ax.plot(xs, y_2d, marker="o", markersize=2.0, linewidth=0.9, label="J_2d")
    ax.set_title(title)
    ax.set_xlabel("Iteration")
    ax.set_ylabel("Value")
    ax.grid(axis="y", alpha=0.25)
    ax.legend(loc="best", fontsize=8)
    fig.tight_layout()
    fig.savefig(out_png, dpi=dpi)
    plt.close(fig)


def build_bin_tick_labels(
    dist_labels: List[str],
    counts_by_bin: Dict[str, List[int]],
    *,
    count_label: str = "px",
) -> List[str]:
    out: List[str] = []
    for lab in dist_labels:
        vals = np.array(counts_by_bin.get(lab, []), dtype=np.float64)
        if vals.size == 0:
            out.append(f"{lab}\n{count_label} avg=0 [0,0]")
            continue
        avg = float(np.mean(vals))
        std = float(np.std(vals, ddof=0))
        lo = max(0.0, avg - std)
        hi = max(0.0, avg + std)
        out.append(f"{lab}\n{count_label} avg={avg:.0f} [{lo:.0f},{hi:.0f}]")
    return out


def filter_bins_by_zero_fraction(
    dist_labels: List[str],
    counts_by_bin: Dict[str, List[int]],
    *,
    zero_frac_threshold: float,
) -> List[str]:
    out: List[str] = []
    for lab in dist_labels:
        vals = counts_by_bin.get(lab, [])
        if not vals:
            continue
        zeros = sum(1 for v in vals if v == 0)
        frac = float(zeros) / float(len(vals))
        if frac < zero_frac_threshold:
            out.append(lab)
    return out


def plot_rae_histograms(
    out_png: Path,
    *,
    title: str,
    dist_labels: List[str],
    data_by_bin: Dict[str, List[float]],
    bins: int = 50,
    dpi: int = 150,
):
    import matplotlib.pyplot as plt  # type: ignore

    n_bins = len(dist_labels)
    ncols = min(4, max(1, n_bins))
    nrows = int(math.ceil(n_bins / ncols))

    fig_w = max(8, ncols * 3.2)
    fig_h = max(3.0, nrows * 2.6)
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(fig_w, fig_h), dpi=dpi)
    if nrows == 1 and ncols == 1:
        axes = np.array([[axes]])
    elif nrows == 1:
        axes = np.array([axes])
    elif ncols == 1:
        axes = axes[:, None]

    for i, lab in enumerate(dist_labels):
        r = i // ncols
        c = i % ncols
        ax = axes[r][c]
        vals = np.array(data_by_bin.get(lab, []), dtype=np.float64)
        if vals.size > 0:
            ax.hist(vals, bins=bins, color="#4C78A8", alpha=0.85)
        ax.set_title(lab)
        ax.set_xlabel("RAE = |GT-PRED|/GT")
        ax.set_ylabel("count")
        ax.grid(True, axis="y", linestyle=":", linewidth=0.6)

    # hide unused subplots
    for i in range(n_bins, nrows * ncols):
        r = i // ncols
        c = i % ncols
        axes[r][c].axis("off")

    fig.suptitle(title)
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png)
    plt.close(fig)


def build_rain_gt_bin_edges(*, max_upper: float = 128.0) -> List[float]:
    edges: List[float] = [0.0, 1.0]
    cur = 1.0
    while cur < float(max_upper):
        cur *= 2.0
        edges.append(float(cur))
    return edges


def format_interval_label(lo: float, hi: float) -> str:
    lo_s = str(int(lo)) if float(lo).is_integer() else f"{lo:g}"
    hi_s = str(int(hi)) if float(hi).is_integer() else f"{hi:g}"
    return f"[{lo_s},{hi_s})"


def plot_gt_binned_patchavg_error(
    out_png: Path,
    *,
    title: str,
    bin_labels: List[str],
    bin_count_stats: Optional[Dict[str, Tuple[float, float]]] = None,
    solver_order: List[str],
    mean_by_solver: Dict[str, List[float]],
    std_by_solver: Dict[str, List[float]],
    y_label: str,
    footnote: Optional[str] = None,
    dpi: int = 150,
):
    import matplotlib.pyplot as plt  # type: ignore

    if not bin_labels or not solver_order:
        return
    x = np.arange(len(bin_labels), dtype=np.float64)
    nsol = max(1, len(solver_order))
    width = 0.8 / nsol
    fig_w = max(10.0, len(bin_labels) * 1.3)
    fig, ax = plt.subplots(figsize=(fig_w, 5.8), dpi=dpi)
    color_cycle = ["#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD", "#8C564B", "#E377C2", "#7F7F7F"]

    for s_idx, solver in enumerate(solver_order):
        m = np.array(mean_by_solver.get(solver, []), dtype=np.float64)
        sd = np.array(std_by_solver.get(solver, []), dtype=np.float64)
        if m.size == 0:
            continue
        offset = (s_idx - (nsol - 1) / 2.0) * width
        xpos = x + offset
        ax.bar(
            xpos,
            m,
            width=width,
            color=color_cycle[s_idx % len(color_cycle)],
            alpha=0.85,
            label=solver,
            yerr=sd,
            error_kw={"elinewidth": 1.0, "capsize": 2.0},
        )

    tick_labels = list(bin_labels)
    if bin_count_stats:
        tick_labels = []
        for lab in bin_labels:
            m_sd = bin_count_stats.get(lab, None)
            if m_sd is None:
                tick_labels.append(f"{lab}\nN_avg=NA\n[N_avg-std,N_avg+std]=NA")
                continue
            m, sd = float(m_sd[0]), float(m_sd[1])
            lo, hi = m - sd, m + sd
            tick_labels.append(f"{lab}\nN_avg={m:.1f}\n[{lo:.1f},{hi:.1f}]")

    ax.set_xticks(x)
    ax.set_xticklabels(tick_labels, rotation=35, ha="right")
    ax.set_ylabel(y_label)
    ax.set_xlabel(
        "GT rain interval (mm/h)\n"
        "Per-bin patch-pixel stats: N_avg = average #pixels per patch in bin; "
        "[N_avg-std, N_avg+std] = mean ± std interval."
    )
    ax.set_title(title)
    ax.grid(True, axis="y", linestyle=":", linewidth=0.7)
    ax.legend(loc="best", fontsize=8)
    if footnote:
        fig.text(0.01, 0.01, footnote, ha="left", va="bottom", fontsize=8)
        fig.tight_layout(rect=[0, 0.04, 1, 1])
    else:
        fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png)
    plt.close(fig)


def build_largest_patch_plot_payload(
    *,
    est_path: Path,
    patch_key: str,
    bin_edges_m: Sequence[float],
    dist_method: str,
    sample_spacing_m: float,
    k_query_samples: int,
    chunk_size: int,
    max_samples_per_link: int,
    max_candidates: int,
) -> Dict[str, Any]:
    est = load_est_payload(est_path)
    hdr = est["header"]
    H = int(hdr["H"])
    W = int(hdr["W"])
    pix = float(hdr.get("pixel_size_m", 125.0))
    area_km2 = float(H) * float(W) * float(pix) * float(pix) / 1_000_000.0

    k_targets = [2, 3]
    if dist_method == "sampled_points":
        dk_maps, _ = compute_dk_maps_sampled_points(
            est,
            k_targets,
            sample_spacing_m=sample_spacing_m,
            k_query_samples=k_query_samples,
            chunk_size=chunk_size,
            max_samples_per_link=max_samples_per_link,
            debug_label=patch_key,
        )
    else:
        dk_maps, _ = compute_dk_maps(est, k_targets, max_candidates=max_candidates)

    links = est.get("links", []) or []
    return {
        "patch_key": patch_key,
        "H": H,
        "W": W,
        "pixel_size_m": pix,
        "area_km2": area_km2,
        "bin_edges_m": [float(v) for v in bin_edges_m],
        "k_targets": k_targets,
        "dk_maps": {str(k): np.asarray(dk_maps[k], dtype=np.float64).tolist() for k in k_targets},
        "links": [
            {
                "x0_m": float(L["x0_m"]),
                "y0_m": float(L["y0_m"]),
                "x1_m": float(L["x1_m"]),
                "y1_m": float(L["y1_m"]),
            }
            for L in links
        ],
    }


def render_largest_patch_distance_bin_maps(
    payload: Dict[str, Any],
    *,
    out_dir: Path,
    dpi: int = 150,
) -> None:
    import matplotlib.pyplot as plt  # type: ignore
    from matplotlib.colors import ListedColormap  # type: ignore

    if not payload:
        return

    patch_key = str(payload["patch_key"])
    H = int(payload["H"])
    W = int(payload["W"])
    pix = float(payload["pixel_size_m"])
    area_km2 = float(payload["area_km2"])
    bins = parse_bins([float(v) for v in payload.get("bin_edges_m", [])])
    labels = [b[2] for b in bins]
    n_bins = len(labels)

    cmap_vals = plt.cm.plasma(np.linspace(0.08, 0.95, max(1, n_bins)))
    cmap = ListedColormap(cmap_vals)
    cmap.set_bad(color="#d9d9d9")

    links = payload.get("links", []) or []
    x0 = np.asarray([float(L["x0_m"]) for L in links], dtype=np.float64) if links else np.zeros(0, dtype=np.float64)
    y0 = np.asarray([float(L["y0_m"]) for L in links], dtype=np.float64) if links else np.zeros(0, dtype=np.float64)
    x1 = np.asarray([float(L["x1_m"]) for L in links], dtype=np.float64) if links else np.zeros(0, dtype=np.float64)
    y1 = np.asarray([float(L["y1_m"]) for L in links], dtype=np.float64) if links else np.zeros(0, dtype=np.float64)

    out_dir.mkdir(parents=True, exist_ok=True)
    x_span_km = (float(W) * pix) / 1000.0
    y_span_km = (float(H) * pix) / 1000.0
    edges = np.asarray([float(v) for v in payload.get("bin_edges_m", [])], dtype=np.float64)
    if edges.size > 0:
        edges = np.unique(edges[np.isfinite(edges)])

    for k in [int(v) for v in payload.get("k_targets", [2, 3])]:
        d_map = np.asarray(payload.get("dk_maps", {}).get(str(k), []), dtype=np.float64).reshape(H, W)
        finite = np.isfinite(d_map)
        if edges.size == 0:
            bin_idx = np.zeros_like(d_map, dtype=np.int32)
        else:
            bin_idx = np.digitize(d_map, edges, right=True).astype(np.int32)
        bin_plot = np.ma.masked_where(~finite, bin_idx)

        fig, ax = plt.subplots(figsize=(9.0, 7.2), dpi=dpi)
        im = ax.imshow(
            bin_plot,
            cmap=cmap,
            vmin=-0.5,
            vmax=max(0, n_bins - 1) + 0.5,
            origin="upper",
            extent=(0.0, float(W) * pix, float(H) * pix, 0.0),
            interpolation="nearest",
            aspect="equal",
        )
        if x0.size > 0:
            for i in range(x0.size):
                ax.plot([x0[i], x1[i]], [y0[i], y1[i]], color="black", linewidth=0.55, alpha=0.55)

        cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        cbar.set_ticks(np.arange(n_bins))
        cbar.set_ticklabels(labels)
        cbar.ax.set_ylabel(f"d{k} distance bin (m)")

        ax.set_title(
            f"Largest patch distance bins (k={k}) with links\n"
            f"{patch_key} | size={x_span_km:.1f}x{y_span_km:.1f} km | area={area_km2:.1f} km^2 | links={x0.size}"
        )
        ax.set_xlabel("x_local (m)")
        ax.set_ylabel("y_local (m)")
        ax.grid(False)

        out_png = out_dir / f"largest_patch_distance_bins_k{k}.png"
        fig.tight_layout()
        fig.savefig(out_png, dpi=dpi)
        plt.close(fig)
        print(f"[largest_patch_bins] wrote {out_png}")


def default_cache_path(*, out_dir: Path, excel_name: str) -> Path:
    stem = Path(excel_name).stem
    return out_dir / f"{stem}_report_cache.json"


def write_report_cache(path: Path, cache: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, sort_keys=True)


def read_report_cache(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        obj = json.load(f)
    if not isinstance(obj, dict):
        raise ValueError(f"Cache at {path} is not a JSON object.")
    return obj


def objective_header_comments_map() -> Dict[str, str]:
    return {
        "patch_key": (
            "Objective_J terms are evaluated with one common legacy objective evaluator for all solvers, "
            "for apples-to-apples comparison. Solver-native objective normalization status is listed in "
            "sheet Objective_J_Conventions."
        ),
        "solver": (
            "Solver label for this row (GT is included as a reference row). "
            "Each row is one patch-solver-parameter triple."
        ),
        "objective_scaling": (
            "Objective convention inferred per solver (NORMALIZED, UNNORMALIZED, N/A_BASELINE, or UNKNOWN)."
        ),
        "unnormalized_data_fit_term": (
            "Unnormalized data-fit term J1 = sum_valid_links ((A_hat - A_obs) / L_km)^2."
        ),
        "unnormalized_objective": (
            "Unnormalized weighted objective: J1 + w_1d*J2 + w_total*J3 + w_2d*J4."
        ),
        "unnormalized_smoothness_term": (
            "Unnormalized smoothness term J2 = sum_neighbor_pairs (log(R + eps)_u - log(R + eps)_v)^2."
        ),
        "unnormalized_shrinkage_term": (
            "Unnormalized shrinkage term J3 = sum_pixels R^2."
        ),
        "unnormalized_second_derivative_term": (
            "Unnormalized second-derivative term J4 = sum_triplets (((R_b-R_a)-(R_c-R_b))^2)."
        ),
        "normalized_data_fit_term": (
            "Normalized data-fit term: J1 / #valid_links (for NORMALIZED-objective solvers)."
        ),
        "normalized_smoothness_term": (
            "Normalized smoothness term: J2 / #pixels (for NORMALIZED-objective solvers)."
        ),
        "normalized_shrinkage_term": (
            "Normalized shrinkage term: J3 / #pixels (for NORMALIZED-objective solvers)."
        ),
        "normalized_second_der_term": (
            "Normalized second-derivative term: J4 / #pixels (for NORMALIZED-objective solvers)."
        ),
    }


def reorder_report_sheets(sheets: Dict[str, List[Dict[str, Any]]]) -> Dict[str, List[Dict[str, Any]]]:
    ordered: Dict[str, List[Dict[str, Any]]] = {}
    sheet_names = list(sheets.keys())

    for prefix in ("LinkStats_", "DistanceStats_GTvs", "DistanceStatsK2_", "CoverageStats_"):
        for name in sheet_names:
            if name.startswith(prefix) and name not in ordered:
                ordered[name] = sheets[name]

    for name in sheet_names:
        if name.startswith("OverallStats_"):
            continue
        if name not in ordered:
            ordered[name] = sheets[name]

    for name in sheet_names:
        if name.startswith("OverallStats_") and name not in ordered:
            ordered[name] = sheets[name]

    return ordered


# ----------------------------
# Main
# ----------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--analyze-only", "--analyze_only", dest="analyze_only", action="store_true")
    ap.add_argument("--output-dir", default=None)
    ap.add_argument("--cache-path", default=None)
    args = ap.parse_args()

    cfg_path = Path(args.config)
    cfg = load_config_file(cfg_path)
    validate_analysis_config(cfg)
    base_dir = cfg_path.resolve().parent

    # inputs
    gt_dir = resolve_path(deep_get(cfg, "input.gt_dir"), base_dir=base_dir)
    est_input_dir = resolve_path(deep_get(cfg, "input.est_input_dir"), base_dir=base_dir)
    if gt_dir is None or est_input_dir is None:
        raise SystemExit("Config must provide input.gt_dir and input.est_input_dir")

    gt_prefix = str(deep_get(cfg, "input.gt_prefix", "gt"))
    gt_key_pref = list(deep_get(cfg, "data.gt_key_preference", ["R_gt", "rain", "gt"]))

    # solvers
    solvers_cfg = normalize_solvers(deep_get(cfg, "input.solvers"), path_for_err="input.solvers")
    # build solver list: (name,label,dir,prefix,key_pref)
    solvers: List[Tuple[str,str,Path,str,List[str]]] = []
    solver_module_by_label: Dict[str, str] = {}
    solver_scaling_by_label: Dict[str, str] = {}
    solver_effective_scaling_by_label: Dict[str, str] = {}
    solver_cfg_by_label: Dict[str, Dict[str, Any]] = {}
    for s in solvers_cfg:
        name = str(s.get("name") or s.get("label") or "solver")
        label = str(s.get("label") or name)
        module_name = str(s.get("module", ""))
        sol_dir = resolve_path(s.get("sol_dir"), base_dir=base_dir)
        if sol_dir is None:
            raise SystemExit(f"Solver {label}: missing sol_dir")
        sol_prefix = str(s.get("sol_prefix", "est"))
        sol_key_pref = list(s.get("sol_key_preference", ["R_hat"]))
        solvers.append((name, label, sol_dir, sol_prefix, sol_key_pref))
        solver_module_by_label[label] = module_name
        solver_scaling_by_label[label] = objective_scaling_from_module(module_name)
        solver_cfg_by_label[label] = dict(s)

    # outputs
    out_dir = resolve_path(deep_get(cfg, "output.out_dir", "batch_analyze_output_multi"), base_dir=base_dir) or (base_dir / "batch_analyze_output_multi").resolve()
    if args.output_dir:
        out_dir = resolve_path(args.output_dir, base_dir=base_dir) or out_dir
    images_subdir = str(deep_get(cfg, "output.images_subdir", "images"))
    excel_name = str(deep_get(cfg, "output.excel_filename", "coverage_stats_long_multi.xlsx"))
    img_dir = (out_dir / images_subdir)
    excel_path = out_dir / excel_name
    cache_path = resolve_path(args.cache_path, base_dir=base_dir) if args.cache_path else default_cache_path(out_dir=out_dir, excel_name=excel_name)

    # params
    thr = float(deep_get(cfg, "rain.threshold_mmph", 1.0))
    fp_fn_thresholds = parse_fp_fn_thresholds(cfg, default_threshold_mmph=thr)
    cov_bins_cfg = list(deep_get(cfg, "coverage.bins", [0,1,2,3,4,"5+"]))
    cov_exact, cov_ge = parse_coverage_bins(cov_bins_cfg)

    k_values = list(deep_get(cfg, "distance.k_values", [3]))
    k_values = sorted({int(k) for k in k_values if int(k) >= 1})
    if not k_values:
        k_values = [3]

    bin_edges = list(deep_get(cfg, "distance.bin_edges_m", [125,375,750,1500,3125]))
    dist_bins = parse_bins(bin_edges)
    dist_labels = [b[2] for b in dist_bins]
    jatten_k_values = list(deep_get(cfg, "jatten_link_distance.k_values", k_values))
    jatten_k_values = sorted({int(k) for k in jatten_k_values if int(k) >= 1})
    if not jatten_k_values:
        jatten_k_values = list(k_values)
    jatten_bin_edges = list(deep_get(cfg, "jatten_link_distance.bin_edges_m", bin_edges))
    jatten_dist_bins = parse_bins(jatten_bin_edges)
    jatten_dist_labels = [b[2] for b in jatten_dist_bins]
    max_candidates = int(deep_get(cfg, "distance.max_candidates", 64))
    dist_method = str(deep_get(cfg, "distance.method", "sampled_points")).strip().lower()
    sample_spacing_m = float(deep_get(cfg, "distance.sample_spacing_m", 250.0))
    k_query_samples = int(deep_get(cfg, "distance.k_query_samples", 48))
    chunk_size = int(deep_get(cfg, "distance.chunk_size", 8000))
    max_samples_per_link = int(deep_get(cfg, "distance.max_samples_per_link", 200))
    n_jobs_raw = deep_get(cfg, "analysis.n_jobs", deep_get(cfg, "n_jobs", 1))
    try:
        n_jobs = int(n_jobs_raw)
    except Exception:
        raise SystemExit("analysis.n_jobs (or n_jobs) must be an integer")
    if n_jobs == 0:
        raise SystemExit("analysis.n_jobs must be != 0")
    if n_jobs < 0:
        n_jobs = max(1, (os.cpu_count() or 1) + 1 + n_jobs)
    n_jobs = max(1, n_jobs)

    # Legacy Objective_J is disabled; use fixed internal reference weights
    # for map-based objective evaluation in Objective_NativeBySolver.
    obj_eps = float(DEFAULT_REFERENCE_EPS)
    obj_pairs: List[Tuple[float, float, float]] = []
    ref_eval_lam = float(DEFAULT_REFERENCE_W_SMOOTH)
    ref_eval_mu = float(DEFAULT_REFERENCE_W_SHRINK)
    ref_eval_eta = float(DEFAULT_REFERENCE_W_SECOND_DER)
    ref_eval_eps = float(DEFAULT_REFERENCE_EPS)

    rae_cfg = deep_get(cfg, "rae_hist", {}) or {}
    rae_enabled = bool(rae_cfg.get("enabled", False)) and bool(deep_get(cfg, "plots.enable_rae_histograms", True))
    rae_bins = int(rae_cfg.get("bins", 50))
    rae_max_patches = rae_cfg.get("max_patches", None)
    rae_out_subdir = str(rae_cfg.get("out_dir", "images/RAE_histograms"))

    rainy_bin_cfg = deep_get(cfg, "rainy_error_bins", {}) or {}
    rainy_bins_enabled = bool(rainy_bin_cfg.get("enabled", True))
    rainy_bins_max_upper = float(rainy_bin_cfg.get("max_upper", 128.0))
    rainy_bins_zero_frac = float(rainy_bin_cfg.get("prune_zero_frac", 0.5))
    rainy_edges = build_rain_gt_bin_edges(max_upper=rainy_bins_max_upper)
    rainy_intervals: List[Tuple[float, float, str]] = []
    for i in range(len(rainy_edges) - 1):
        lo = float(rainy_edges[i])
        hi = float(rainy_edges[i + 1])
        rainy_intervals.append((lo, hi, format_interval_label(lo, hi)))

    # plot scaling
    show = bool(deep_get(cfg, "plots.show", False))
    skip_patch_plots = bool(deep_get(cfg, "plots.skip_patch_plots", False))
    auto_y = bool(deep_get(cfg, "plots.automatic_vertical_scaling", True))
    y_scale = deep_get(cfg, "plots.vertical_scale", None)
    if not auto_y:
        try:
            y_max = float(y_scale)
        except Exception:
            raise SystemExit("plots.automatic_vertical_scaling=false but plots.vertical_scale is not a valid number")
        if y_max < 0:
            raise SystemExit("plots.vertical_scale must be non-negative")
    else:
        y_max = None
    dpi = int(deep_get(cfg, "plots.dpi", 150))
    cmap_gt = str(deep_get(cfg, "plots.cmap_gt", "viridis"))
    cmap_sol = str(deep_get(cfg, "plots.cmap_sol", "viridis"))
    cmap_diff = str(deep_get(cfg, "plots.cmap_diff", "seismic"))
    cmap_abs = str(deep_get(cfg, "plots.cmap_abs_diff", "magma"))
    cmap_rel = str(deep_get(cfg, "plots.cmap_rel", "seismic"))
    cmap_abs_rel = str(deep_get(cfg, "plots.cmap_abs_rel", "magma"))
    bin_spacing = float(deep_get(cfg, "plots.bin_spacing", 1.35))
    prune_bins_enabled = bool(deep_get(cfg, "plots.prune_bins_enabled", False))
    prune_bins_zero_frac = float(deep_get(cfg, "plots.prune_bins_zero_frac", 0.5))

    # load files
    gt_files = list_npz(gt_dir, gt_prefix)
    if not gt_files:
        raise SystemExit(f"No GT files found in {gt_dir} with prefix {gt_prefix}_*.npz")

    # index GT by patch key
    gt_by_key: Dict[str, Path] = {patch_key_from_filename(p.name): p for p in gt_files}

    # also need est_input json per patch key
    est_jsons = list_json(est_input_dir, "est_input")
    est_by_key: Dict[str, Path] = {patch_key_from_filename(p.name): p for p in est_jsons}

    # Largest patch distance-bin overlay payload for the render stage.
    biggest_key, biggest_est_path, biggest_area_km2 = pick_biggest_est_patch(est_by_key)
    largest_patch_plot_payload: Optional[Dict[str, Any]] = None
    if biggest_key is not None and biggest_est_path is not None:
        largest_patch_plot_payload = build_largest_patch_plot_payload(
            est_path=biggest_est_path,
            patch_key=biggest_key,
            bin_edges_m=bin_edges,
            dist_method=dist_method,
            sample_spacing_m=sample_spacing_m,
            k_query_samples=k_query_samples,
            chunk_size=chunk_size,
            max_samples_per_link=max_samples_per_link,
            max_candidates=max_candidates,
        )
        print(
            f"[largest_patch_bins] source patch={biggest_key} "
            f"area_km2={biggest_area_km2:.2f} est={biggest_est_path}"
        )

    # sheets data
    sheets: Dict[str, List[Dict[str, Any]]] = {}
    sheet_order: List[str] = []

    # For plots: per k -> per method -> per bin -> list of per-patch medians / p90s
    medians_rainy: Dict[int, Dict[str, Dict[str, List[float]]]] = {k: {} for k in k_values}
    medians_nonrainy: Dict[int, Dict[str, Dict[str, List[float]]]] = {k: {} for k in k_values}
    p90s_rainy: Dict[int, Dict[str, Dict[str, List[float]]]] = {k: {} for k in k_values}
    p90s_nonrainy: Dict[int, Dict[str, Dict[str, List[float]]]] = {k: {} for k in k_values}
    jatten_medians: Dict[int, Dict[str, Dict[str, List[float]]]] = {k: {} for k in jatten_k_values}
    bin_counts: Dict[int, Dict[str, Dict[str, List[int]]]] = {
        k: {"rainy": {lab: [] for lab in dist_labels}, "nonrainy": {lab: [] for lab in dist_labels}}
        for k in k_values
    }
    bin_counts_all: Dict[int, Dict[str, List[int]]] = {k: {lab: [] for lab in dist_labels} for k in k_values}
    jatten_link_bin_counts: Dict[int, Dict[str, List[int]]] = {
        k: {lab: [] for lab in jatten_dist_labels} for k in jatten_k_values
    }
    bin_counts_seen: set = set()
    bin_counts_all_seen: set = set()
    jatten_bin_counts_seen: set = set()
    dry_metrics: Dict[str, Dict[str, List[float]]] = {}
    fp_fn_threshold_metrics: Dict[str, Dict[float, Dict[str, List[float]]]] = {}

    objective_gt_done: set = set()
    objective_vals: Dict[Tuple[float, float, float], Dict[str, Dict[str, float]]] = {}
    native_objective_rows: List[Dict[str, Any]] = []
    iter_feas_rows: List[Dict[str, Any]] = []
    overall_by_solver_rows: List[Dict[str, Any]] = []
    gt_rows_done: set = set()
    link_metrics: Dict[str, Dict[str, Dict[str, Any]]] = {}
    stop_rows: List[Dict[str, Any]] = []
    solver_info_rows: List[Dict[str, Any]] = []
    solver_formula_by_label: Dict[str, str] = {}
    j_behavior_plots: List[Dict[str, Any]] = []
    rae_hist_plots: List[Dict[str, Any]] = []
    patch_plot_jobs: List[Dict[str, Any]] = []
    patch_map_metric_rows: List[Dict[str, Any]] = []
    patch_map_metrics_by_solver: Dict[str, Dict[str, List[float]]] = {}
    gtbin_counts_global: Dict[str, Dict[str, int]] = {lab: {} for _, _, lab in rainy_intervals}
    gtbin_rel_patch_means: Dict[str, Dict[str, List[float]]] = {}
    gtbin_abs_patch_means: Dict[str, Dict[str, List[float]]] = {}

    # Legacy Objective_J sheet is disabled; keep native per-solver objective reporting only.
    obj_enabled = False

    # iterate per solver
    for name, label, sol_dir, sol_prefix, sol_key_pref in progress_iter(
        solvers, total=len(solvers), desc="Solvers"
    ):
        patch_map_metrics_by_solver[label] = {
            "rmse_mmph": [],
            "bias_mmph": [],
            "pearson_corr": [],
        }
        sol_files = list_npz(sol_dir, sol_prefix)
        sol_by_key = {patch_key_from_filename(p.name): p for p in sol_files}

        # Solver settings/objective summary from representative solution metadata.
        meta_sample: Dict[str, float] = {}
        sample_npz: Optional[Path] = sol_files[0] if sol_files else None
        if sample_npz is not None:
            meta_sample = load_npz_meta_scalars(sample_npz)
        scaling_cfg = solver_scaling_by_label.get(label, "UNKNOWN")
        scaling_meta = objective_scaling_from_meta(meta_sample)
        scaling = scaling_cfg if scaling_cfg != "UNKNOWN" else scaling_meta
        solver_effective_scaling_by_label[label] = scaling
        module_name = solver_module_by_label.get(label, "")
        if not module_name:
            module_name = "(not provided in analyze config)"
        solver_info_rows.append(dict(
            solver=label,
            solver_name=name,
            module=module_name,
            objective_scaling=scaling,
            objective_formula=solver_objective_formula_text(scaling=scaling, meta=meta_sample),
            settings_summary=summarize_solver_settings(meta_sample),
            sol_dir=str(sol_dir),
            sol_prefix=sol_prefix,
            sample_solution_npz=(str(sample_npz) if sample_npz is not None else None),
        ))
        solver_formula_by_label[label] = solver_objective_formula_text(scaling=scaling, meta=meta_sample)

        cov_rows: List[Dict[str, Any]] = []
        dist_rows_by_k: Dict[int, List[Dict[str, Any]]] = {k: [] for k in k_values}
        overall_rows: List[Dict[str, Any]] = []
        link_rows: List[Dict[str, Any]] = []
        link_metrics[label] = {}
        dry_metrics[label] = {"fp_rates": [], "fn_rates": []}
        fp_fn_threshold_metrics[label] = {
            float(t): {
                "tp_rate_all": [],
                "fp_rate_all": [],
                "fn_rate_all": [],
                "tn_rate_all": [],
                "fp_rate_dry": [],
                "fn_rate_wet": [],
            }
            for t in fp_fn_thresholds
        }
        gtbin_rel_patch_means[label] = {lab: [] for _, _, lab in rainy_intervals}
        gtbin_abs_patch_means[label] = {lab: [] for _, _, lab in rainy_intervals}

        for k in k_values:
            medians_rainy[k][label] = {lab: [] for lab in dist_labels}
            medians_nonrainy[k][label] = {lab: [] for lab in dist_labels}
            p90s_rainy[k][label] = {lab: [] for lab in dist_labels}
            p90s_nonrainy[k][label] = {lab: [] for lab in dist_labels}
        for k in jatten_k_values:
            jatten_medians[k][label] = {lab: [] for lab in jatten_dist_labels}

        rae_hist_data: Optional[Dict[int, Dict[str, List[float]]]] = None
        if rae_enabled:
            rae_hist_data = {k: {lab: [] for lab in dist_labels} for k in k_values}
        # match patches by keys present in both GT and solver
        keys = sorted(set(gt_by_key.keys()) & set(sol_by_key.keys()))
        if not keys:
            print(f"[{label}] No matching patches between GT and {sol_dir}")
            continue

        hist_keys = keys
        if rae_enabled and rae_max_patches is not None:
            try:
                kmax = int(rae_max_patches)
                if kmax > 0:
                    hist_keys = keys[:kmax]
            except Exception:
                pass

        tasks: List[Dict[str, Any]] = []
        hist_key_set = set(hist_keys)
        for key in keys:
            gt_path = gt_by_key[key]
            sol_path = sol_by_key[key]
            est_path = est_by_key.get(key, None)
            if est_path is None:
                raise SystemExit(f"Missing est_input JSON for patch {key} under {est_input_dir}")
            tasks.append(dict(
                key=key,
                label=label,
                name=name,
                gt_path=str(gt_path),
                sol_path=str(sol_path),
                est_path=str(est_path),
                gt_key_pref=list(gt_key_pref),
                sol_key_pref=list(sol_key_pref),
                thr=thr,
                fp_fn_thresholds=list(fp_fn_thresholds),
                cov_exact=list(cov_exact),
                cov_ge=cov_ge,
                k_values=list(k_values),
                dist_bins=list(dist_bins),
                jatten_k_values=list(jatten_k_values),
                jatten_dist_bins=list(jatten_dist_bins),
                dist_method=dist_method,
                sample_spacing_m=sample_spacing_m,
                k_query_samples=k_query_samples,
                chunk_size=chunk_size,
                max_samples_per_link=max_samples_per_link,
                max_candidates=max_candidates,
                rainy_bins_enabled=rainy_bins_enabled,
                rainy_intervals=list(rainy_intervals),
                obj_enabled=obj_enabled,
                obj_pairs=[list(x) for x in obj_pairs],
                obj_eps=obj_eps,
                native_obj_enabled=True,
                ref_eval_enabled=True,
                ref_eval_lam=float(ref_eval_lam),
                ref_eval_mu=float(ref_eval_mu),
                ref_eval_eta=float(ref_eval_eta),
                ref_eval_eps=float(ref_eval_eps),
                include_rae_hist=(rae_hist_data is not None and key in hist_key_set),
            ))
            if not skip_patch_plots:
                patch_plot_jobs.append(dict(
                    patch_key=key,
                    solver_label=label,
                    solver_name=name,
                    gt_path=str(gt_path),
                    sol_path=str(sol_path),
                    gt_key_pref=list(gt_key_pref),
                    sol_key_pref=list(sol_key_pref),
                ))

        patch_results: List[Dict[str, Any]] = []
        if n_jobs <= 1:
            for t in progress_iter(tasks, total=len(tasks), desc=f"{label} patches"):
                try:
                    patch_results.append(analyze_single_patch(t))
                except Exception as e:
                    raise SystemExit(f"Patch analysis failed for {t.get('key')} ({label}): {e}") from e
        else:
            with concurrent.futures.ProcessPoolExecutor(max_workers=n_jobs) as ex:
                for res in progress_iter(
                    ex.map(analyze_single_patch, tasks),
                    total=len(tasks),
                    desc=f"{label} patches",
                ):
                    patch_results.append(res)

        for res in patch_results:
            key = str(res["patch_key"])
            stop_rows.append(dict(res["stop_row"]))

            map_metrics = res.get("map_metrics", None)
            if isinstance(map_metrics, dict):
                map_row = dict(
                    patch_key=key,
                    solver=label,
                    solver_name=name,
                    rmse_mmph=map_metrics.get("rmse_mmph", None),
                    bias_mmph=map_metrics.get("bias_mmph", None),
                    pearson_corr=map_metrics.get("pearson_corr", None),
                    n_pixels=int(map_metrics.get("n_pixels", 0)),
                )
                patch_map_metric_rows.append(map_row)
                for metric_name in ("rmse_mmph", "bias_mmph", "pearson_corr"):
                    metric_value = map_row.get(metric_name, None)
                    if metric_value is not None:
                        metric_value_f = float(metric_value)
                        if np.isfinite(metric_value_f):
                            patch_map_metrics_by_solver[label][metric_name].append(metric_value_f)

            fp_rate = res.get("dry_fp_rate", None)
            fn_rate = res.get("dry_fn_rate", None)
            if fp_rate is not None:
                dry_metrics[label]["fp_rates"].append(float(fp_rate))
            if fn_rate is not None:
                dry_metrics[label]["fn_rates"].append(float(fn_rate))
            for row_thr in (res.get("fp_fn_by_threshold", []) or []):
                t = float(row_thr.get("threshold_mmph", thr))
                bucket = fp_fn_threshold_metrics[label].setdefault(
                    t,
                    {"tp_rate_all": [], "fp_rate_all": [], "fn_rate_all": [], "tn_rate_all": [], "fp_rate_dry": [], "fn_rate_wet": []},
                )
                for k_metric in ("tp_rate_all", "fp_rate_all", "fn_rate_all", "tn_rate_all", "fp_rate_dry", "fn_rate_wet"):
                    v_metric = row_thr.get(k_metric, None)
                    if v_metric is not None:
                        bucket[k_metric].append(float(v_metric))

            if rainy_bins_enabled:
                gt_counts = res.get("gtbin_counts", {}) or {}
                gt_rel = res.get("gtbin_rel_means", {}) or {}
                gt_abs = res.get("gtbin_abs_means", {}) or {}
                for _, _, bin_lab in rainy_intervals:
                    if key not in gtbin_counts_global[bin_lab]:
                        gtbin_counts_global[bin_lab][key] = int(gt_counts.get(bin_lab, 0))
                    rel_v = gt_rel.get(bin_lab, None)
                    abs_v = gt_abs.get(bin_lab, None)
                    if rel_v is not None:
                        gtbin_rel_patch_means[label][bin_lab].append(float(rel_v))
                    if abs_v is not None:
                        gtbin_abs_patch_means[label][bin_lab].append(float(abs_v))

            if obj_enabled:
                for lam, mu, eta, j_gt, j_sol in res.get("objective_pairs", []):
                    gt_tag = (key, float(lam), float(mu), float(eta))
                    if gt_tag not in objective_gt_done:
                        objective_vals.setdefault((float(lam), float(mu), float(eta)), {}).setdefault(key, {})["GT"] = j_gt
                        objective_gt_done.add(gt_tag)
                    objective_vals.setdefault((float(lam), float(mu), float(eta)), {}).setdefault(key, {})[label] = j_sol

            native = res.get("native_objective", None)
            ref_obj_pred = res.get("reference_objective_pred", None)
            ref_obj_gt = res.get("reference_objective_gt", None)
            module_name = solver_module_by_label.get(label, "")
            native_row: Dict[str, Any] = dict(
                patch_key=key,
                solver=label,
                definition="",
                solver_name=name,
                module=module_name,
                objective_scaling=solver_effective_scaling_by_label.get(label, "UNKNOWN"),
                objective_available=bool(native is not None),
            )
            if native is None:
                native_row["objective_unavailable_reason"] = (
                    "No scalar meta_* objective weights found in solution npz."
                )
            else:
                native_row["objective_scaling"] = str(native.get("objective_scaling", native_row["objective_scaling"]))
                native_row["w_atten"] = float(native.get("w_atten", 1.0))
                native_row["w_1d"] = float(native.get("w_1d", native.get("w_smooth", 0.0)))
                native_row["w_total"] = float(native.get("w_total", native.get("w_shrinkage", 0.0)))
                native_row["w_2d"] = float(native.get("w_2d", native.get("w_second_der", 0.0)))
                if "eps" in native:
                    native_row["eps"] = float(native["eps"])
                n_valid = native.get("n_valid_links", None)
                n_pix = native.get("n_pixels", None)
                if n_valid is not None:
                    native_row["n_valid_links"] = int(n_valid)
                if n_pix is not None:
                    native_row["n_pixels"] = int(n_pix)
                den_links = float(n_valid) if (n_valid is not None and float(n_valid) > 0.0) else None
                den_pix = float(n_pix) if (n_pix is not None and float(n_pix) > 0.0) else None
                j1 = float(native.get("J1", 0.0))
                j2 = float(native.get("J2", 0.0))
                j3 = float(native.get("J3", 0.0))
                j4 = float(native.get("J4", 0.0))
                wa = float(native_row["w_atten"])
                ws = float(native_row["w_1d"])
                wk = float(native_row["w_total"])
                wd = float(native_row["w_2d"])

                j_atten = float(j1 / den_links) if den_links is not None else j1
                j_1d = float(j2 / den_pix) if den_pix is not None else j2
                j_total = float(j3 / den_pix) if den_pix is not None else j3
                j_2d = float(j4 / den_pix) if den_pix is not None else j4

                native_row["J_atten"] = j_atten
                native_row["J_1d"] = j_1d
                native_row["J_total"] = j_total
                native_row["J_2d"] = j_2d
                native_row["weighted_J_atten"] = wa * j_atten
                native_row["weighted_J_1d"] = ws * j_1d
                native_row["weighted_J_total"] = wk * j_total
                native_row["weighted_J_2d"] = wd * j_2d

                if str(native_row.get("objective_scaling", "")) == "CONSTRAINED_NORMALIZED":
                    if "w_atten" in native_row:
                        native_row["J_weighted_sum"] = (
                            float(native_row["weighted_J_atten"])
                            + float(native_row["weighted_J_1d"])
                            + float(native_row["weighted_J_total"])
                            + float(native_row["weighted_J_2d"])
                        )
                    else:
                        native_row["J_weighted_sum"] = (
                            float(native_row["weighted_J_1d"])
                            + float(native_row["weighted_J_total"])
                            + float(native_row["weighted_J_2d"])
                        )
                else:
                    native_row["J_weighted_sum"] = (
                        float(native_row["weighted_J_atten"])
                        + float(native_row["weighted_J_1d"])
                        + float(native_row["weighted_J_total"])
                        + float(native_row["weighted_J_2d"])
                    )
            if native is None and ref_obj_pred is not None:
                n_valid_ref = float(ref_obj_pred.get("n_valid_links", 0.0))
                n_pix_ref = float(ref_obj_pred.get("n_pixels", 0.0))
                j1_ref = float(ref_obj_pred.get("J1", 0.0))
                j2_ref = float(ref_obj_pred.get("J2", 0.0))
                j3_ref = float(ref_obj_pred.get("J3", 0.0))
                j4_ref = float(ref_obj_pred.get("J4", 0.0))
                native_row["w_atten"] = 1.0
                native_row["w_1d"] = 1.0
                native_row["w_total"] = 1.0
                native_row["w_2d"] = 1.0
                if "eps" in ref_obj_pred:
                    native_row["eps"] = float(ref_obj_pred["eps"])
                native_row["n_valid_links"] = int(n_valid_ref) if n_valid_ref > 0.0 else int(ref_obj_pred.get("n_valid_links", 0))
                native_row["n_pixels"] = int(n_pix_ref) if n_pix_ref > 0.0 else int(ref_obj_pred.get("n_pixels", 0))
                if n_valid_ref > 0.0 and n_pix_ref > 0.0:
                    native_row["J_atten"] = j1_ref / n_valid_ref
                    native_row["J_1d"] = j2_ref / n_pix_ref
                    native_row["J_total"] = j3_ref / n_pix_ref
                    native_row["J_2d"] = j4_ref / n_pix_ref
                    native_row["weighted_J_atten"] = float(native_row["J_atten"])
                    native_row["weighted_J_1d"] = float(native_row["J_1d"])
                    native_row["weighted_J_total"] = float(native_row["J_total"])
                    native_row["weighted_J_2d"] = float(native_row["J_2d"])
                    native_row["J_weighted_sum"] = (
                        float(native_row["weighted_J_atten"])
                        + float(native_row["weighted_J_1d"])
                        + float(native_row["weighted_J_total"])
                        + float(native_row["weighted_J_2d"])
                    )
            native_objective_rows.append(native_row)
            if ref_obj_gt is not None and key not in gt_rows_done:
                gt_rows_done.add(key)
                n_valid_ref = float(ref_obj_gt.get("n_valid_links", 0.0))
                n_pix_ref = float(ref_obj_gt.get("n_pixels", 0.0))
                j1_ref = float(ref_obj_gt.get("J1", 0.0))
                j2_ref = float(ref_obj_gt.get("J2", 0.0))
                j3_ref = float(ref_obj_gt.get("J3", 0.0))
                j4_ref = float(ref_obj_gt.get("J4", 0.0))
                gt_row: Dict[str, Any] = dict(
                    patch_key=key,
                    solver="GT",
                    definition="",
                    solver_name="GT",
                    module="reference_only",
                    objective_scaling="N/A_BASELINE",
                    objective_available=False,
                    objective_unavailable_reason="GT is not produced by optimization; native objective is not applicable.",
                    w_atten=1.0,
                    w_1d=1.0,
                    w_total=1.0,
                    w_2d=1.0,
                    n_valid_links=int(n_valid_ref) if n_valid_ref > 0.0 else int(ref_obj_gt.get("n_valid_links", 0)),
                    n_pixels=int(n_pix_ref) if n_pix_ref > 0.0 else int(ref_obj_gt.get("n_pixels", 0)),
                )
                if "eps" in ref_obj_gt:
                    gt_row["eps"] = float(ref_obj_gt["eps"])
                if n_valid_ref > 0.0 and n_pix_ref > 0.0:
                    gt_row["J_atten"] = j1_ref / n_valid_ref
                    gt_row["J_1d"] = j2_ref / n_pix_ref
                    gt_row["J_total"] = j3_ref / n_pix_ref
                    gt_row["J_2d"] = j4_ref / n_pix_ref
                    gt_row["weighted_J_atten"] = float(gt_row["J_atten"])
                    gt_row["weighted_J_1d"] = float(gt_row["J_1d"])
                    gt_row["weighted_J_total"] = float(gt_row["J_total"])
                    gt_row["weighted_J_2d"] = float(gt_row["J_2d"])
                    gt_row["J_weighted_sum"] = (
                        float(gt_row["weighted_J_atten"])
                        + float(gt_row["weighted_J_1d"])
                        + float(gt_row["weighted_J_total"])
                        + float(gt_row["weighted_J_2d"])
                    )
                native_objective_rows.append(gt_row)

            cov_rows.extend(res.get("cov_rows", []))

            dist_rows_payload = res.get("dist_rows_by_k", {}) or {}
            for k in k_values:
                dist_rows_by_k[k].extend(dist_rows_payload.get(k, []))
            overall_rows.extend(res.get("overall_rows", []))

            for k, mask_name, bin_lab, cnt in res.get("bin_counts_entries", []):
                count_key = (int(k), str(mask_name), str(bin_lab), key)
                if count_key not in bin_counts_seen:
                    bin_counts[int(k)][str(mask_name)][str(bin_lab)].append(int(cnt))
                    bin_counts_seen.add(count_key)
            for k, bin_lab, cnt in res.get("bin_counts_all_entries", []):
                count_key_all = (int(k), str(bin_lab), key)
                if count_key_all not in bin_counts_all_seen:
                    bin_counts_all[int(k)][str(bin_lab)].append(int(cnt))
                    bin_counts_all_seen.add(count_key_all)
            for k, bin_lab, cnt in res.get("jatten_link_bin_counts_entries", []):
                jcount_key = (int(k), str(bin_lab), key)
                if jcount_key not in jatten_bin_counts_seen:
                    jatten_link_bin_counts[int(k)][str(bin_lab)].append(int(cnt))
                    jatten_bin_counts_seen.add(jcount_key)

            for k, bin_lab, v in res.get("medians_r_entries", []):
                medians_rainy[int(k)][label][str(bin_lab)].append(float(v))
            for k, bin_lab, v in res.get("medians_n_entries", []):
                medians_nonrainy[int(k)][label][str(bin_lab)].append(float(v))
            for k, bin_lab, v in res.get("jatten_medians_entries", []):
                jatten_medians[int(k)][label][str(bin_lab)].append(float(v))
            for k, bin_lab, v in res.get("p90s_r_entries", []):
                p90s_rainy[int(k)][label][str(bin_lab)].append(float(v))
            for k, bin_lab, v in res.get("p90s_n_entries", []):
                p90s_nonrainy[int(k)][label][str(bin_lab)].append(float(v))

            if rae_hist_data is not None:
                for k, bin_lab, vals in res.get("rae_hist_entries", []):
                    rae_hist_data[int(k)][str(bin_lab)].extend(list(vals))

            link_row = res.get("link_row", None)
            if link_row is not None:
                link_rows.append(link_row)
            link_metric = res.get("link_metric", None)
            if link_metric is not None:
                link_metrics[label][key] = link_metric
            iter_entries = res.get("itertrace_entries", []) or []
            iter_summary = res.get("itertrace_summary", {}) or {}
            itertrace_path = res.get("itertrace_path", None)
            if isinstance(iter_entries, list) and iter_entries:
                feasible_n = int(sum(1 for it in iter_entries if bool(it.get("feasible", False))))
                total_n = int(len(iter_entries))
                infeasible_n = int(total_n - feasible_n)
                best_iter = int(
                    iter_summary.get(
                        "best_iteration_by_weighted_sum",
                        iter_summary.get("best_iteration_by_native_total", -1),
                    )
                )
                if best_iter < 0:
                    best_idx = int(
                        np.argmin(
                            [
                                float(it.get("J_weighted_sum", it.get("J_native_total", np.inf)))
                                for it in iter_entries
                            ]
                        )
                    )
                    best_iter = int(iter_entries[best_idx].get("iter", best_idx + 1))
                iter_feas_rows.append(
                    dict(
                        patch_key=key,
                        solver=label,
                        solver_name=name,
                        total_iterations=total_n,
                        feasible_iterations=feasible_n,
                        infeasible_iterations=infeasible_n,
                        best_iteration_by_weighted_sum=best_iter,
                        itertrace_json=itertrace_path,
                    )
                )
                j_behavior_plots.append(
                    dict(
                        solver_label=label,
                        solver_name=name,
                        patch_key=key,
                        iterations=iter_entries,
                    )
                )

        # store sheets (order: LinkStats, DistanceStats, CoverageStats)
        link_rows = append_average_rows(link_rows, group_keys=[])
        cov_rows = append_average_rows(cov_rows, group_keys=["mask_type", "coverage_bin"])
        for k in k_values:
            dist_rows_by_k[k] = append_average_rows(
                dist_rows_by_k[k],
                group_keys=["mask_type", "distance_bin_m"],
            )
        overall_rows = append_average_rows(
            overall_rows,
            group_keys=["mask_type", "distance_bin_m"],
        )
        for r in overall_rows:
            rr = dict(r)
            rr["solver"] = label
            rr["solver_name"] = name
            rr["module"] = module_name
            overall_by_solver_rows.append(rr)

        sheet_name = f"LinkStats_GTvs{label}"
        sheets[sheet_name] = link_rows
        sheet_order.append(sheet_name)
        k_values_for_sheets = sorted(k_values, key=lambda kv: (0 if kv == 3 else (1 if kv == 2 else 2), kv))
        for k in k_values_for_sheets:
            if k == 3:
                sheet_name = f"DistanceStats_GTvs{label}"
            else:
                sheet_name = f"DistanceStatsK{k}_GTvs{label}"
            sheets[sheet_name] = dist_rows_by_k[k]
            sheet_order.append(sheet_name)
        sheet_name = f"CoverageStats_GTvs{label}"
        sheets[sheet_name] = cov_rows
        sheet_order.append(sheet_name)
        sheet_name = f"OverallStats_GTvs{label}"
        sheets[sheet_name] = overall_rows
        sheet_order.append(sheet_name)

        # RAE histograms (rainy only)
        if rae_hist_data is not None:
            for k in k_values:
                rae_hist_plots.append(
                    dict(
                        out_relpath=str(Path(rae_out_subdir) / f"rae_hist_k{k}_{label}.png"),
                        title=f"RAE histograms (rainy) | k={k} | {label}",
                        dist_labels=list(dist_labels),
                        data_by_bin=rae_hist_data[k],
                    )
                )

    # stopping diagnostics sheets
    if stop_rows:
        stop_defs: List[Dict[str, Any]] = [
            dict(
                patch_key="DEFINITION",
                definition_term="nfev",
                definition="Number of objective function evaluations f(x) during optimization.",
            ),
            dict(
                patch_key="DEFINITION",
                definition_term="njev",
                definition="Number of gradient/Jacobian evaluations (for scalar objectives: gradient evaluations).",
            ),
            dict(
                patch_key="DEFINITION",
                definition_term="proj_grad_inf",
                definition=(
                    "Infinity norm of the projected gradient at the final iterate "
                    "(first-order stationarity under bounds R>=0). Smaller is better."
                ),
            ),
            dict(
                patch_key="DEFINITION",
                definition_term="rel_decrease",
                definition=(
                    "Relative objective decrease between the previous and final iterate: "
                    "|f_prev-f_final| / max(1, |f_prev|, |f_final|). Smaller means progress stalled."
                ),
            ),
            dict(
                patch_key="DEFINITION",
                definition_term="gtol_vs_proj_grad_inf",
                definition=(
                    "Gradient-based stopping test uses gtol_met := (proj_grad_inf <= gtol)."
                ),
            ),
        ]
        sheets["StoppingInfo"] = list(stop_rows) + stop_defs
        sheet_order.append("StoppingInfo")

        by_solver: Dict[str, List[Dict[str, Any]]] = {}
        for r in stop_rows:
            by_solver.setdefault(str(r.get("solver", "")), []).append(r)

        reason_stats_rows: List[Dict[str, Any]] = []
        for solver in sorted(by_solver.keys()):
            rows_s = by_solver[solver]
            total_patches = len(rows_s)
            counts: Dict[str, int] = {}
            for r in rows_s:
                reason = str(r.get("reason_bucket", "unknown"))
                counts[reason] = counts.get(reason, 0) + 1
            for reason in sorted(counts.keys()):
                cnt = counts[reason]
                pct = (100.0 * float(cnt) / float(total_patches)) if total_patches > 0 else 0.0
                reason_stats_rows.append(dict(
                    solver=solver,
                    stop_reason=reason,
                    n_patches=int(cnt),
                    total_patches=int(total_patches),
                    pct_of_patches=float(pct),
                ))
        sheets["StoppingReasonStats"] = reason_stats_rows
        sheet_order.append("StoppingReasonStats")

    if solver_info_rows:
        sheets["SolverSettingsAndObjective"] = solver_info_rows
        sheet_order.append("SolverSettingsAndObjective")

    if native_objective_rows:
        native_objective_rows = append_average_rows(
            native_objective_rows,
            group_keys=["solver"],
        )
        native_objective_rows = append_native_objective_definition_rows(native_objective_rows)
        sheets["Objective_NativeBySolver"] = native_objective_rows
        sheet_order.append("Objective_NativeBySolver")

    if fp_fn_threshold_metrics:
        fpfn_rows: List[Dict[str, Any]] = []
        for label in sorted(fp_fn_threshold_metrics.keys()):
            by_thr = fp_fn_threshold_metrics[label]
            for thr_i in sorted(by_thr.keys()):
                vals = by_thr[thr_i]
                tp_all = np.array(vals.get("tp_rate_all", []), dtype=np.float64)
                fp_all = np.array(vals.get("fp_rate_all", []), dtype=np.float64)
                fn_all = np.array(vals.get("fn_rate_all", []), dtype=np.float64)
                tn_all = np.array(vals.get("tn_rate_all", []), dtype=np.float64)
                fp_dry = np.array(vals.get("fp_rate_dry", []), dtype=np.float64)
                fn_wet = np.array(vals.get("fn_rate_wet", []), dtype=np.float64)
                n_patches = int(max(tp_all.size, fp_all.size, fn_all.size, tn_all.size))

                def _sem(arr: np.ndarray) -> float:
                    if arr.size <= 1:
                        return 0.0
                    return float(np.std(arr, ddof=0) / math.sqrt(float(arr.size - 1)))

                fpfn_rows.append(
                    dict(
                        solver=label,
                        threshold_mmph=float(thr_i),
                        positive_definition="wet (gt >= threshold)",
                        fp_definition="pred wet & GT dry",
                        fn_definition="pred dry & GT wet",
                        fp_rate_all_mean=(float(np.mean(fp_all)) if fp_all.size else 0.0),
                        fp_rate_all_std=(float(np.std(fp_all, ddof=0)) if fp_all.size else 0.0),
                        fn_rate_all_mean=(float(np.mean(fn_all)) if fn_all.size else 0.0),
                        fn_rate_all_std=(float(np.std(fn_all, ddof=0)) if fn_all.size else 0.0),
                        fp_rate_dry_mean=(float(np.mean(fp_dry)) if fp_dry.size else 0.0),
                        fp_rate_dry_std=(float(np.std(fp_dry, ddof=0)) if fp_dry.size else 0.0),
                        fn_rate_wet_mean=(float(np.mean(fn_wet)) if fn_wet.size else 0.0),
                        fn_rate_wet_std=(float(np.std(fn_wet, ddof=0)) if fn_wet.size else 0.0),
                        n_patches=n_patches,
                        tp_definition="pred wet & GT wet",
                        tn_definition="pred dry & GT dry",
                        tp_rate_all_mean=(float(np.mean(tp_all)) if tp_all.size else 0.0),
                        tp_rate_all_std=(float(np.std(tp_all, ddof=0)) if tp_all.size else 0.0),
                        tp_rate_all_sem=_sem(tp_all),
                        fp_rate_all_sem=_sem(fp_all),
                        fn_rate_all_sem=_sem(fn_all),
                        tn_rate_all_mean=(float(np.mean(tn_all)) if tn_all.size else 0.0),
                        tn_rate_all_std=(float(np.std(tn_all, ddof=0)) if tn_all.size else 0.0),
                        tn_rate_all_sem=_sem(tn_all),
                    )
                )
        sheets["FPFN_ByThreshold"] = fpfn_rows
        sheet_order.append("FPFN_ByThreshold")

    if iter_feas_rows:
        iter_feas_rows = append_average_rows(iter_feas_rows, group_keys=["solver", "solver_name"])
        sheets["IterationFeasibility"] = iter_feas_rows
        sheet_order.append("IterationFeasibility")

    if overall_by_solver_rows:
        overall_by_solver_rows = append_average_rows(
            overall_by_solver_rows,
            group_keys=["solver", "solver_name", "module", "mask_type", "distance_bin_m"],
        )
        overall_by_solver_rows = enrich_overall_by_solver_ratios(overall_by_solver_rows)
        sheets["OverallStats_BySolver"] = overall_by_solver_rows
        sheet_order.append("OverallStats_BySolver")

    attenuation_error_rows: List[Dict[str, Any]] = []
    for solver_label in [label for _, label, _, _, _ in solvers if label in link_metrics]:
        per_patch = link_metrics.get(solver_label, {})
        vals_mean = [
            float(v["MEAN_ABS_ATTN_ERR_PER_KM"])
            for v in per_patch.values()
            if "MEAN_ABS_ATTN_ERR_PER_KM" in v
        ]
        vals_lenw = [
            float(v["LENGTH_WEIGHTED_ABS_ATTN_ERR_PER_KM"])
            for v in per_patch.values()
            if "LENGTH_WEIGHTED_ABS_ATTN_ERR_PER_KM" in v
        ]
        n_patches = int(max(len(vals_mean), len(vals_lenw)))
        attenuation_error_rows.append(
            dict(
                solver=solver_label,
                mean_abs_attn_err_per_km_mean=(float(np.mean(vals_mean)) if vals_mean else 0.0),
                mean_abs_attn_err_per_km_std=(float(np.std(vals_mean, ddof=0)) if vals_mean else 0.0),
                length_weighted_abs_attn_err_per_km_mean=(float(np.mean(vals_lenw)) if vals_lenw else 0.0),
                length_weighted_abs_attn_err_per_km_std=(float(np.std(vals_lenw, ddof=0)) if vals_lenw else 0.0),
                n_patches=n_patches,
            )
        )
    if attenuation_error_rows:
        attenuation_error_rows.extend([
            dict(
                solver="DEFINITION",
                metric="mean_abs_attn_err_per_km",
                definition=(
                    "Per patch: mean over valid links of |A_hat - A_obs| / L_km; "
                    "then summarized across patches by solver."
                ),
            ),
            dict(
                solver="DEFINITION",
                metric="length_weighted_abs_attn_err_per_km",
                definition=(
                    "Per patch: sum over valid links of |A_hat - A_obs| divided by "
                    "sum over valid links of L_km; then summarized across patches by solver."
                ),
            ),
        ])
        sheets["AttenuationErrorPerKm_BySolver"] = attenuation_error_rows
        sheet_order.append("AttenuationErrorPerKm_BySolver")

    patch_map_metric_summary_rows: List[Dict[str, Any]] = []
    for solver_label in [label for _, label, _, _, _ in solvers]:
        vals = patch_map_metrics_by_solver.get(solver_label, {})
        rmse_vals = np.asarray(vals.get("rmse_mmph", []), dtype=np.float64)
        bias_vals = np.asarray(vals.get("bias_mmph", []), dtype=np.float64)
        corr_vals = np.asarray(vals.get("pearson_corr", []), dtype=np.float64)

        def _mean(arr: np.ndarray) -> Optional[float]:
            return float(np.mean(arr)) if arr.size else None

        def _std(arr: np.ndarray) -> Optional[float]:
            return float(np.std(arr, ddof=0)) if arr.size else None

        def _min(arr: np.ndarray) -> Optional[float]:
            return float(np.min(arr)) if arr.size else None

        def _max(arr: np.ndarray) -> Optional[float]:
            return float(np.max(arr)) if arr.size else None

        patch_map_metric_summary_rows.append(
            dict(
                solver=solver_label,
                rmse_mmph_mean=_mean(rmse_vals),
                rmse_mmph_std=_std(rmse_vals),
                rmse_mmph_min=_min(rmse_vals),
                rmse_mmph_max=_max(rmse_vals),
                bias_mmph_mean=_mean(bias_vals),
                bias_mmph_std=_std(bias_vals),
                bias_mmph_min=_min(bias_vals),
                bias_mmph_max=_max(bias_vals),
                pearson_corr_mean=_mean(corr_vals),
                pearson_corr_std=_std(corr_vals),
                pearson_corr_min=_min(corr_vals),
                pearson_corr_max=_max(corr_vals),
                n_patches=int(max(rmse_vals.size, bias_vals.size, corr_vals.size)),
            )
        )
    if patch_map_metric_rows:
        sheets["PatchMapMetrics_ByPatch"] = patch_map_metric_rows
        sheet_order.append("PatchMapMetrics_ByPatch")
    if patch_map_metric_summary_rows:
        sheets["PatchMapMetrics_BySolver"] = patch_map_metric_summary_rows
        sheet_order.append("PatchMapMetrics_BySolver")

    # Enrich LinkStats sheets (non-baseline solvers) with J_atten comparisons vs IDW/ILDW.
    enrich_linkstats_with_baseline_jatten(sheets)
    # Enrich Coverage/Distance/Overall per-solver tabs with ALG-vs-baseline ratios.
    enrich_binned_stats_with_baseline_ratios(sheets)

    # write excel (ordered)
    ordered_from_sheet_order: Dict[str, List[Dict[str, Any]]] = {}
    for name in sheet_order:
        if name in sheets and name not in ordered_from_sheet_order:
            ordered_from_sheet_order[name] = sheets[name]
    for name in sheets:
        if name not in ordered_from_sheet_order:
            ordered_from_sheet_order[name] = sheets[name]
    ordered_sheets = reorder_report_sheets(ordered_from_sheet_order)

    link_ratio_entries: List[Dict[str, Any]] = []
    if "IDW" in link_metrics:
        metrics = ["L1", "J1", "E", "E2", "AVG_ABS_NORM_LINK_RATIO"]
        for solver_label in [label for _, label, _, _, _ in solvers if label in link_metrics]:
            per_patch = link_metrics.get(solver_label, {})
            per_patch_idw = link_metrics.get("IDW", {})
            keys = sorted(set(per_patch.keys()) & set(per_patch_idw.keys()))
            if not keys:
                continue
            ratio_vals: Dict[str, List[float]] = {m: [] for m in metrics}
            for k_patch in keys:
                v = per_patch[k_patch]
                v_idw = per_patch_idw[k_patch]
                if v_idw["L1"] != 0:
                    ratio_vals["L1"].append(v["L1"] / v_idw["L1"])
                j1_num = float(v.get("J1_len1", v["J1"]))
                j1_den = float(v_idw.get("J1_len1", v_idw["J1"]))
                if j1_den != 0:
                    ratio_vals["J1"].append(j1_num / j1_den)
                if v_idw["E"] != 0:
                    ratio_vals["E"].append(v["E"] / v_idw["E"])
                if v_idw["E2"] != 0:
                    ratio_vals["E2"].append(v["E2"] / v_idw["E2"])
                arr_alg = np.asarray(v.get("abs_norm_resid_valid", []), dtype=np.float64)
                arr_idw = np.asarray(v_idw.get("abs_norm_resid_valid", []), dtype=np.float64)
                n_common = min(arr_alg.size, arr_idw.size)
                if n_common > 0:
                    den = arr_idw[:n_common]
                    num = arr_alg[:n_common]
                    good = np.isfinite(num) & np.isfinite(den)
                    if np.any(good):
                        mean_num = float(np.mean(num[good]))
                        mean_den = float(np.mean(den[good]))
                        if mean_den > 0.0:
                            ratio_vals["AVG_ABS_NORM_LINK_RATIO"].append(mean_num / mean_den)
            link_ratio_entries.extend([
                dict(solver=solver_label, label=f"L1({solver_label})/L1(IDW)", values=ratio_vals["L1"]),
                dict(solver=solver_label, label=f"J1_len1({solver_label})/J1_len1(IDW)", values=ratio_vals["J1"]),
                dict(solver=solver_label, label=f"E({solver_label})/E(IDW)", values=ratio_vals["E"]),
                dict(solver=solver_label, label=f"E2({solver_label})/E2(IDW)", values=ratio_vals["E2"]),
                dict(
                    solver=solver_label,
                    label=f"AvgLinkRatio(|A-Ahat|/L): {solver_label}/IDW",
                    values=ratio_vals["AVG_ABS_NORM_LINK_RATIO"],
                ),
            ])

    gtbin_plot_data: Optional[Dict[str, Any]] = None
    if rainy_bins_enabled and gtbin_rel_patch_means and gtbin_abs_patch_means:
        ordered_labels = [lab for _, _, lab in rainy_intervals]
        labels_to_plot = list(ordered_labels)
        while labels_to_plot:
            tail = labels_to_plot[-1]
            counts = list(gtbin_counts_global.get(tail, {}).values())
            if not counts:
                labels_to_plot.pop()
                continue
            zfrac = float(sum(1 for c in counts if c == 0)) / float(len(counts))
            if zfrac >= rainy_bins_zero_frac:
                labels_to_plot.pop()
            else:
                break

        solver_order = [label for _, label, _, _, _ in solvers if label in gtbin_rel_patch_means]
        if labels_to_plot and solver_order:
            bin_count_stats: Dict[str, Tuple[float, float]] = {}
            for bin_lab in labels_to_plot:
                counts = np.array(list(gtbin_counts_global.get(bin_lab, {}).values()), dtype=np.float64)
                if counts.size > 0:
                    bin_count_stats[bin_lab] = (float(np.mean(counts)), float(np.std(counts, ddof=0)))

            rel_mean: Dict[str, List[float]] = {}
            rel_std: Dict[str, List[float]] = {}
            abs_mean: Dict[str, List[float]] = {}
            abs_std: Dict[str, List[float]] = {}
            for solver_label in solver_order:
                rel_mean[solver_label] = []
                rel_std[solver_label] = []
                abs_mean[solver_label] = []
                abs_std[solver_label] = []
                for bin_lab in labels_to_plot:
                    rv = np.array(gtbin_rel_patch_means.get(solver_label, {}).get(bin_lab, []), dtype=np.float64)
                    av = np.array(gtbin_abs_patch_means.get(solver_label, {}).get(bin_lab, []), dtype=np.float64)
                    rel_mean[solver_label].append(float(np.mean(rv)) if rv.size else 0.0)
                    rel_std[solver_label].append(float(np.std(rv, ddof=0)) if rv.size else 0.0)
                    abs_mean[solver_label].append(float(np.mean(av)) if av.size else 0.0)
                    abs_std[solver_label].append(float(np.std(av, ddof=0)) if av.size else 0.0)

            gtbin_plot_data = dict(
                labels_to_plot=labels_to_plot,
                bin_count_stats=bin_count_stats,
                rel_mean=rel_mean,
                rel_std=rel_std,
                abs_mean=abs_mean,
                abs_std=abs_std,
            )

    cache: Dict[str, Any] = {
        "cache_version": 1,
        "config_path": str(cfg_path.resolve()),
        "output": {
            "out_dir": str(out_dir),
            "images_subdir": images_subdir,
            "excel_filename": excel_name,
        },
        "render": {
            "dpi": dpi,
            "bin_spacing": bin_spacing,
            "y_max": y_max,
            "prune_bins_enabled": prune_bins_enabled,
            "prune_bins_zero_frac": prune_bins_zero_frac,
            "rae_bins": rae_bins,
            "threshold_mmph": thr,
            "cmap_gt": cmap_gt,
            "cmap_sol": cmap_sol,
            "cmap_diff": cmap_diff,
            "cmap_abs_diff": cmap_abs,
            "cmap_rel": cmap_rel,
            "cmap_abs_rel": cmap_abs_rel,
        },
        "solvers": {
            "order": [label for _, label, _, _, _ in solvers],
            "names": {label: name for name, label, _, _, _ in solvers},
        },
        "labels": {
            "k_values": list(k_values),
            "dist_labels": list(dist_labels),
            "jatten_k_values": list(jatten_k_values),
            "jatten_dist_labels": list(jatten_dist_labels),
            "rainy_intervals": [list(x) for x in rainy_intervals],
        },
        "ordered_sheets": ordered_sheets,
        "plot_data": {
            "medians_rainy": {str(k): medians_rainy[k] for k in k_values},
            "medians_nonrainy": {str(k): medians_nonrainy[k] for k in k_values},
            "p90s_rainy": {str(k): p90s_rainy[k] for k in k_values},
            "p90s_nonrainy": {str(k): p90s_nonrainy[k] for k in k_values},
            "jatten_medians": {str(k): jatten_medians[k] for k in jatten_k_values},
            "bin_counts": {str(k): bin_counts[k] for k in k_values},
            "jatten_link_bin_counts": {str(k): jatten_link_bin_counts[k] for k in jatten_k_values},
        },
        "j_behavior_plots": j_behavior_plots,
        "rae_hist_plots": rae_hist_plots,
        "patch_plot_jobs": patch_plot_jobs,
        "largest_patch_plot_payload": largest_patch_plot_payload,
        "link_ratio_entries": link_ratio_entries,
        "gtbin_plot_data": gtbin_plot_data,
        "patch_map_metrics_plot_data": {
            solver_label: {
                metric_name: list(values)
                for metric_name, values in metric_dict.items()
            }
            for solver_label, metric_dict in patch_map_metrics_by_solver.items()
        },
    }
    write_report_cache(cache_path, cache)
    print(f"Wrote cache: {cache_path}")
    if args.analyze_only:
        for stale_png in (img_dir / "fp_fn_vs_threshold.png",):
            if stale_png.exists():
                stale_png.unlink()
        patch_error_maps_dir = img_dir / "patch_error_maps"
        if patch_error_maps_dir.exists():
            shutil.rmtree(patch_error_maps_dir)
        return 0
    render_script = Path(__file__).resolve().with_name("render_analysis_report.py")
    subprocess.run(
        [
            sys.executable,
            str(render_script),
            "--cache",
            str(cache_path),
            "--output-dir",
            str(out_dir),
        ],
        check=True,
    )
    return 0

    objective_header_comments = {
        "patch_key": (
            "Objective_J terms are evaluated with one common legacy objective evaluator for all solvers, "
            "for apples-to-apples comparison. Solver-native objective normalization status is listed in "
            "sheet Objective_J_Conventions."
        ),
        "solver": (
            "Solver label for this row (GT is included as a reference row). "
            "Each row is one patch-solver-parameter triple."
        ),
        "objective_scaling": (
            "Objective convention inferred per solver (NORMALIZED, UNNORMALIZED, N/A_BASELINE, or UNKNOWN)."
        ),
        "unnormalized_data_fit_term": (
            "Unnormalized data-fit term J1 = sum_valid_links ((A_hat - A_obs) / L_km)^2."
        ),
        "unnormalized_objective": (
            "Unnormalized weighted objective: J1 + w_1d*J2 + w_total*J3 + w_2d*J4."
        ),
        "unnormalized_smoothness_term": (
            "Unnormalized smoothness term J2 = sum_neighbor_pairs (log(R + eps)_u - log(R + eps)_v)^2."
        ),
        "unnormalized_shrinkage_term": (
            "Unnormalized shrinkage term J3 = sum_pixels R^2."
        ),
        "unnormalized_second_derivative_term": (
            "Unnormalized second-derivative term J4 = sum_triplets (((R_b-R_a)-(R_c-R_b))^2)."
        ),
        "normalized_data_fit_term": (
            "Normalized data-fit term: J1 / #valid_links (for NORMALIZED-objective solvers)."
        ),
        "normalized_smoothness_term": (
            "Normalized smoothness term: J2 / #pixels (for NORMALIZED-objective solvers)."
        ),
        "normalized_shrinkage_term": (
            "Normalized shrinkage term: J3 / #pixels (for NORMALIZED-objective solvers)."
        ),
        "normalized_second_der_term": (
            "Normalized second-derivative term: J4 / #pixels (for NORMALIZED-objective solvers)."
        ),
    }
    objective_sheet_comments: Dict[str, str] = {}
    obj_rows = ordered_sheets.get("Objective_J", [])
    if obj_rows:
        for col_name in obj_rows[0].keys():
            for target_col, txt in objective_header_comments.items():
                if col_name == target_col:
                    objective_sheet_comments[col_name] = txt
                    break

    write_workbook(
        excel_path,
        ordered_sheets,
        header_comments={"Objective_J": objective_sheet_comments},
    )
    print(f"Wrote Excel: {excel_path}")

    # ratio plot vs IDW for link metrics
    if "IDW" in link_metrics:
        entries: List[Tuple[str, str, List[float]]] = []
        metrics = ["L1", "J1", "E", "E2", "AVG_ABS_NORM_LINK_RATIO"]
        for solver_label in [label for _, label, _, _, _ in solvers if label in link_metrics]:
            per_patch = link_metrics.get(solver_label, {})
            per_patch_idw = link_metrics.get("IDW", {})
            keys = sorted(set(per_patch.keys()) & set(per_patch_idw.keys()))
            if not keys:
                continue
            ratio_vals: Dict[str, List[float]] = {m: [] for m in metrics}
            for k in keys:
                v = per_patch[k]
                v_idw = per_patch_idw[k]
                # L1 ratio
                if v_idw["L1"] != 0:
                    ratio_vals["L1"].append(v["L1"] / v_idw["L1"])
                # J1 ratio with first-power length denominator:
                # J1_len1 = sum((A_hat - A_obs)^2 / L_km)
                j1_num = float(v.get("J1_len1", v["J1"]))
                j1_den = float(v_idw.get("J1_len1", v_idw["J1"]))
                if j1_den != 0:
                    ratio_vals["J1"].append(j1_num / j1_den)
                # E ratio
                if v_idw["E"] != 0:
                    ratio_vals["E"].append(v["E"] / v_idw["E"])
                # E2 ratio
                if v_idw["E2"] != 0:
                    ratio_vals["E2"].append(v["E2"] / v_idw["E2"])
                # Ratio of per-patch means over links:
                # mean_l(|A-Ahat_alg|/L) / mean_l(|A-Ahat_IDW|/L)
                arr_alg = np.asarray(v.get("abs_norm_resid_valid", []), dtype=np.float64)
                arr_idw = np.asarray(v_idw.get("abs_norm_resid_valid", []), dtype=np.float64)
                n_common = min(arr_alg.size, arr_idw.size)
                if n_common > 0:
                    den = arr_idw[:n_common]
                    num = arr_alg[:n_common]
                    good = np.isfinite(num) & np.isfinite(den)
                    if np.any(good):
                        mean_num = float(np.mean(num[good]))
                        mean_den = float(np.mean(den[good]))
                        if mean_den > 0.0:
                            ratio_vals["AVG_ABS_NORM_LINK_RATIO"].append(mean_num / mean_den)

            entries.append((solver_label, f"L1({solver_label})/L1(IDW)", ratio_vals["L1"]))
            entries.append((solver_label, f"J1_len1({solver_label})/J1_len1(IDW)", ratio_vals["J1"]))
            entries.append((solver_label, f"E({solver_label})/E(IDW)", ratio_vals["E"]))
            entries.append((solver_label, f"E2({solver_label})/E2(IDW)", ratio_vals["E2"]))
            entries.append((
                solver_label,
                f"AvgLinkRatio(|A-Ahat|/L): {solver_label}/IDW",
                ratio_vals["AVG_ABS_NORM_LINK_RATIO"],
            ))

        if entries:
            plot_ratio_iqr(
                img_dir / "link_ratio_summary.png",
                title="Link-metric ratios vs IDW (median and IQR across patches)",
                entries=entries,
                dpi=dpi,
            )

    # GT-binned all-pixels plots (powers-of-2 bins)
    if rainy_bins_enabled and gtbin_rel_patch_means and gtbin_abs_patch_means:
        ordered_labels = [lab for _, _, lab in rainy_intervals]
        labels_to_plot = list(ordered_labels)
        # Tail-only pruning by zero fraction across patches.
        while labels_to_plot:
            tail = labels_to_plot[-1]
            counts = list(gtbin_counts_global.get(tail, {}).values())
            if not counts:
                labels_to_plot.pop()
                continue
            zfrac = float(sum(1 for c in counts if c == 0)) / float(len(counts))
            if zfrac >= rainy_bins_zero_frac:
                labels_to_plot.pop()
            else:
                break

        solver_order = [label for _, label, _, _, _ in solvers if label in gtbin_rel_patch_means]
        if labels_to_plot and solver_order:
            bin_count_stats: Dict[str, Tuple[float, float]] = {}
            for bin_lab in labels_to_plot:
                counts = np.array(list(gtbin_counts_global.get(bin_lab, {}).values()), dtype=np.float64)
                if counts.size > 0:
                    bin_count_stats[bin_lab] = (float(np.mean(counts)), float(np.std(counts, ddof=0)))

            rel_mean: Dict[str, List[float]] = {}
            rel_std: Dict[str, List[float]] = {}
            abs_mean: Dict[str, List[float]] = {}
            abs_std: Dict[str, List[float]] = {}
            for solver_label in solver_order:
                rel_mean[solver_label] = []
                rel_std[solver_label] = []
                abs_mean[solver_label] = []
                abs_std[solver_label] = []
                for bin_lab in labels_to_plot:
                    rv = np.array(gtbin_rel_patch_means.get(solver_label, {}).get(bin_lab, []), dtype=np.float64)
                    av = np.array(gtbin_abs_patch_means.get(solver_label, {}).get(bin_lab, []), dtype=np.float64)
                    rel_mean[solver_label].append(float(np.mean(rv)) if rv.size else 0.0)
                    rel_std[solver_label].append(float(np.std(rv, ddof=0)) if rv.size else 0.0)
                    abs_mean[solver_label].append(float(np.mean(av)) if av.size else 0.0)
                    abs_std[solver_label].append(float(np.std(av, ddof=0)) if av.size else 0.0)

            plot_gt_binned_patchavg_error(
                img_dir / "gt_binned_patchavg_relative_abs_error_all_pixels.png",
                title="GT-binned all-pixels error (avg of patch-averaged error ± std)",
                bin_labels=labels_to_plot,
                bin_count_stats=bin_count_stats,
                solver_order=solver_order,
                mean_by_solver=rel_mean,
                std_by_solver=rel_std,
                y_label="Avg patch error (|GT-PRED|/GT; [0,1) uses |GT-PRED|)",
                footnote="Note: For GT in [0,1) mm/h, the metric uses absolute error |GT-PRED| (not RAE).",
                dpi=dpi,
                )
            plot_gt_binned_patchavg_error(
                img_dir / "gt_binned_patchavg_absolute_error_all_pixels.png",
                title="GT-binned all-pixels absolute error (avg of patch means ± std)",
                bin_labels=labels_to_plot,
                bin_count_stats=bin_count_stats,
                solver_order=solver_order,
                mean_by_solver=abs_mean,
                std_by_solver=abs_std,
                y_label="Avg patch absolute error |GT-PRED| (mm/h)",
                dpi=dpi,
            )

    # plots across solvers (IQR of per-patch medians)
    method_order = [label for _, label, _, _, _ in solvers if label in medians_rainy[k_values[0]]]
    if method_order:
        for k in k_values:
            summary_r = compute_iqr_profile(medians_rainy[k], dist_labels)
            summary_n = compute_iqr_profile(medians_nonrainy[k], dist_labels)
            summary_p90_r = compute_p90_profile(p90s_rainy[k], dist_labels)
            summary_p90_n = compute_p90_profile(p90s_nonrainy[k], dist_labels)
            labels_r = dist_labels
            labels_n = dist_labels
            if prune_bins_enabled:
                labels_r = filter_bins_by_zero_fraction(
                    dist_labels, bin_counts[k]["rainy"],
                    zero_frac_threshold=prune_bins_zero_frac,
                )
                labels_n = filter_bins_by_zero_fraction(
                    dist_labels, bin_counts[k]["nonrainy"],
                    zero_frac_threshold=prune_bins_zero_frac,
                )
            tick_labels_r = build_bin_tick_labels(labels_r, bin_counts[k]["rainy"])
            tick_labels_n = build_bin_tick_labels(labels_n, bin_counts[k]["nonrainy"])

            if len(k_values) == 1 and k == 3:
                rainy_name = "distance_iqr_medians_rainy_multi.png"
                nonrainy_name = "distance_iqr_medians_nonrainy_multi.png"
                rainy_title = "Rainy pixels: IQR of per-patch median |(GT-PRED)/GT| by distance bin"
                nonrainy_title = "Non-rainy pixels: IQR of per-patch median |GT-PRED| by distance bin"
                rainy_p90_name = "distance_iqr_p90s_rainy_multi.png"
                nonrainy_p90_name = "distance_iqr_p90s_nonrainy_multi.png"
                rainy_p90_title = "Rainy pixels: per-patch p90 |(GT-PRED)/GT| by distance bin (median, p25-p75)"
                nonrainy_p90_title = "Non-rainy pixels: per-patch p90 |GT-PRED| by distance bin (median, p25-p75)"
            else:
                rainy_name = f"distance_iqr_medians_rainy_multi_k{k}.png"
                nonrainy_name = f"distance_iqr_medians_nonrainy_multi_k{k}.png"
                rainy_title = f"Rainy pixels: IQR of per-patch median |(GT-PRED)/GT| by distance bin (k={k})"
                nonrainy_title = f"Non-rainy pixels: IQR of per-patch median |GT-PRED| by distance bin (k={k})"
                rainy_p90_name = f"distance_iqr_p90s_rainy_multi_k{k}.png"
                nonrainy_p90_name = f"distance_iqr_p90s_nonrainy_multi_k{k}.png"
                rainy_p90_title = f"Rainy pixels: per-patch p90 |(GT-PRED)/GT| by distance bin (median, p25-p75; k={k})"
                nonrainy_p90_title = f"Non-rainy pixels: per-patch p90 |GT-PRED| by distance bin (median, p25-p75; k={k})"

            plot_iqr_bars(
                img_dir / rainy_name,
                rainy_title,
                summary_r, labels_r, method_order,
                y_max=y_max, dpi=dpi, bin_spacing=bin_spacing,
                tick_labels=tick_labels_r,
            )
            plot_iqr_bars(
                img_dir / nonrainy_name,
                nonrainy_title,
                summary_n, labels_n, method_order,
                y_max=y_max, dpi=dpi, bin_spacing=bin_spacing,
                tick_labels=tick_labels_n,
            )

            # median-of-p90s plots (p25/p90 across per-patch p90s)
            plot_iqr_bars(
                img_dir / rainy_p90_name,
                rainy_p90_title,
                summary_p90_r, labels_r, method_order,
                y_max=y_max, dpi=dpi, bin_spacing=bin_spacing,
                tick_labels=tick_labels_r,
                y_label="Per-patch p90 error (median, p25-p75)",
            )
            plot_iqr_bars(
                img_dir / nonrainy_p90_name,
                nonrainy_p90_title,
                summary_p90_n, labels_n, method_order,
                y_max=y_max, dpi=dpi, bin_spacing=bin_spacing,
                tick_labels=tick_labels_n,
                y_label="Per-patch p90 error (median, p25-p75)",
            )

            # relative plots vs IDW (median-of-medians)
            if "IDW" in summary_r:
                summary_r_rel = compute_relative_iqr_profile(summary_r, idw_label="IDW", dist_labels=dist_labels)
                summary_n_rel = compute_relative_iqr_profile(summary_n, idw_label="IDW", dist_labels=dist_labels)

                plot_iqr_bars(
                    img_dir / rainy_name.replace(".png", "_rel.png"),
                    f"{rainy_title} (relative to IDW medians)",
                    summary_r_rel, labels_r, method_order,
                    y_max=None, dpi=dpi, bin_spacing=bin_spacing,
                    tick_labels=tick_labels_r,
                )
                plot_iqr_bars(
                    img_dir / nonrainy_name.replace(".png", "_rel.png"),
                    f"{nonrainy_title} (relative to IDW medians)",
                    summary_n_rel, labels_n, method_order,
                    y_max=None, dpi=dpi, bin_spacing=bin_spacing,
                    tick_labels=tick_labels_n,
                )

        # J_atten distance-binned profile:
        # per patch and distance bin, compute median of per-pixel attributed J_atten value;
        # plot median across patches with IQR (p25-p75), one series per solver.
        for k in jatten_k_values:
            summary_jatten = compute_iqr_profile(jatten_medians[k], jatten_dist_labels)
            labels_jatten = jatten_dist_labels
            if prune_bins_enabled:
                labels_jatten = filter_bins_by_zero_fraction(
                    jatten_dist_labels,
                    jatten_link_bin_counts[k],
                    zero_frac_threshold=prune_bins_zero_frac,
                )
            tick_labels_jatten = build_bin_tick_labels(
                labels_jatten,
                jatten_link_bin_counts[k],
                count_label="links",
            )
            jatten_img_dir = img_dir / "jatten_iqr_plots"
            if len(jatten_k_values) == 1 and k == 3:
                jatten_name = "distance_iqr_medians_jatten_multi.png"
                jatten_title = "Link-distance-binned J_atten: IQR of per-patch medians"
            else:
                jatten_name = f"distance_iqr_medians_jatten_multi_k{k}.png"
                jatten_title = f"Link-distance-binned J_atten: IQR of per-patch medians (k={k})"
            plot_iqr_bars(
                jatten_img_dir / jatten_name,
                jatten_title,
                summary_jatten,
                labels_jatten,
                method_order,
                y_max=None,
                dpi=dpi,
                bin_spacing=bin_spacing,
                tick_labels=tick_labels_jatten,
                x_label="Distance bin (m)\nSecond line: avg links [avg-std, avg+std]",
                y_label="J_atten link contribution (per-patch median, IQR across patches)",
                footnote=(
                    "Per-link contribution: ((A_hat - A_obs)^2 / L_km) / #valid_links. "
                    "Links are binned by segment-to-segment distance to the k-th closest other link."
                ),
            )
            plot_iqr_bars(
                jatten_img_dir / jatten_name.replace(".png", "_no_p25_p75.png"),
                f"{jatten_title} (medians only)",
                summary_jatten,
                labels_jatten,
                method_order,
                y_max=None,
                dpi=dpi,
                bin_spacing=bin_spacing,
                tick_labels=tick_labels_jatten,
                x_label="Distance bin (m)\nSecond line: avg links [avg-std, avg+std]",
                y_label="J_atten link contribution (per-patch median across patches)",
                footnote=(
                    "Per-link contribution: ((A_hat - A_obs)^2 / L_km) / #valid_links. "
                    "Links are binned by segment-to-segment distance to the k-th closest other link."
                ),
                show_iqr=False,
            )

            # Relative J_atten profiles using baseline median-of-medians per bin.
            for baseline_label, tag in (("IDW", "idw"), ("ILDW", "ildw")):
                if baseline_label not in summary_jatten:
                    continue
                summary_jatten_rel = compute_relative_iqr_profile(
                    summary_jatten,
                    idw_label=baseline_label,
                    dist_labels=dist_labels,
                )
                if len(jatten_k_values) == 1 and k == 3:
                    jatten_rel_name = f"distance_iqr_medians_jatten_multi_rel_{tag}.png"
                else:
                    jatten_rel_name = f"distance_iqr_medians_jatten_multi_k{k}_rel_{tag}.png"
                plot_iqr_bars(
                    jatten_img_dir / jatten_rel_name,
                    (
                        "Link-distance-binned J_atten: IQR of per-patch medians "
                        f"(relative to {baseline_label} median)"
                    ),
                    summary_jatten_rel,
                    labels_jatten,
                    method_order,
                    y_max=None,
                    dpi=dpi,
                    bin_spacing=bin_spacing,
                    tick_labels=tick_labels_jatten,
                    x_label="Distance bin (m)\nSecond line: avg links [avg-std, avg+std]",
                    y_label=(
                        "J_atten link-contribution ratio "
                        f"(per-patch median, IQR; baseline={baseline_label} p50)"
                    ),
                    footnote=(
                        "For each distance bin, p25/p50/p75 are divided by "
                        f"{baseline_label}'s p50 in that bin."
                    ),
                )
                plot_iqr_bars(
                    jatten_img_dir / jatten_rel_name.replace(".png", "_no_p25_p75.png"),
                    (
                        "Link-distance-binned J_atten: per-patch medians "
                        f"(relative to {baseline_label} median, medians only)"
                    ),
                    summary_jatten_rel,
                    labels_jatten,
                    method_order,
                    y_max=None,
                    dpi=dpi,
                    bin_spacing=bin_spacing,
                    tick_labels=tick_labels_jatten,
                    x_label="Distance bin (m)\nSecond line: avg links [avg-std, avg+std]",
                    y_label=(
                        "J_atten link-contribution ratio "
                        f"(per-patch median; baseline={baseline_label} p50)"
                    ),
                    footnote=(
                        "For each distance bin, medians are divided by "
                        f"{baseline_label}'s p50 in that bin."
                    ),
                    show_iqr=False,
                )
        print(f"Wrote plots under: {img_dir}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
