#!/usr/bin/env python3
# batch_analyze.py
#
# Batch GT vs solution rainfall analysis:
# - YAML/JSON config support (--config)
# - Coverage c(pixel) = number of links that intersect pixel (from est_input JSON)
# - Rainy pixels (GT>=thr): relative errors (signed & abs)
# - Non-rainy pixels (GT<thr): additive errors (signed & abs)
# - Separate PNG per patch for rainy and non-rainy
# - Consistent heatmap scaling across all patches
# - Excel output in long/tidy format: (patch_key, mask_type, coverage_bin) rows
#
# Dependencies: numpy, matplotlib, openpyxl, (optional) pyyaml for YAML configs

from __future__ import annotations

import argparse
import html
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Optional (but recommended) for fast nearest-neighbour queries
try:
    from scipy.spatial import cKDTree  # type: ignore
except Exception:
    cKDTree = None


# ============================================================
# Config loading
# ============================================================

def load_config_file(path: str | Path) -> dict:
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix in (".yaml", ".yml"):
        try:
            import yaml  # type: ignore
        except Exception as e:
            raise RuntimeError("YAML config requires PyYAML: pip install pyyaml") from e
        with path.open("r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        return {} if cfg is None else cfg

    if suffix == ".json":
        with path.open("r", encoding="utf-8") as f:
            cfg = json.load(f)
        return cfg

    raise ValueError("Config must be .yaml/.yml or .json")


def deep_get(d: dict, path: str, default=None):
    cur: Any = d
    for part in path.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return default
        cur = cur[part]
    return cur


# ============================================================
# Pairing
# ============================================================

PATCH_RE = re.compile(r"(patch\d+)", re.IGNORECASE)
TS_RE = re.compile(r"(\d{10,14})")  # timestamps (10-14 digits)


def make_pair_key(path: Path) -> str:
    """
    Pairing key based on filename tokens.
    Prefers patch + timestamp if both exist, else patch, else timestamp, else stem.
    """
    name = path.name
    patch = PATCH_RE.search(name)
    ts = TS_RE.search(name)

    if patch and ts:
        return f"{patch.group(1).lower()}__{ts.group(1)}"
    if patch:
        return patch.group(1).lower()
    if ts:
        return ts.group(1)

    return re.sub(r"[^a-zA-Z0-9]+", "_", path.stem).lower()


def robust_glob(pattern: str) -> List[Path]:
    pat = Path(pattern)
    if pattern.startswith("/"):
        parent = pat.parent
        return sorted(parent.glob(pat.name))
    return sorted(Path(".").glob(pattern))


def build_file_map(files: List[Path], label: str) -> Dict[str, Path]:
    """Build a key->Path map and refuse silent collisions."""
    out: Dict[str, Path] = {}
    collisions: Dict[str, List[Path]] = {}
    for p in files:
        k = make_pair_key(p)
        if k in out:
            collisions.setdefault(k, [out[k]]).append(p)
        else:
            out[k] = p
    if collisions:
        msgs = []
        for k, ps in list(collisions.items())[:10]:
            msgs.append(f"{k}: " + ", ".join(str(x) for x in ps))
        raise ValueError(
            f"{label}: pairing-key collisions detected for {len(collisions)} key(s). "

            f"This would silently overwrite files. Examples:\n" + "\n".join(msgs)

        )
    return out


def list_input_files(cfg: dict) -> Tuple[List[Path], List[Path], List[Path]]:
    # GT
    gt_glob = deep_get(cfg, "input.gt_glob", None)
    gt_dir = deep_get(cfg, "input.gt_dir", None)
    if gt_glob:
        gt_files = robust_glob(gt_glob)
    elif gt_dir:
        gt_files = sorted(Path(gt_dir).glob("*.npz"))
    else:
        gt_files = []

    # SOL
    sol_glob = deep_get(cfg, "input.sol_glob", None)
    sol_dir = deep_get(cfg, "input.sol_dir", None)
    if sol_glob:
        sol_files = robust_glob(sol_glob)
    elif sol_dir:
        sol_files = sorted(Path(sol_dir).glob("*.npz"))
    else:
        sol_files = []

    # EST_INPUT JSONs (for coverage)
    est_glob = deep_get(cfg, "input.est_input_glob", None)
    est_dir = deep_get(cfg, "input.est_input_dir", None)
    if est_glob:
        est_files = robust_glob(est_glob)
    elif est_dir:
        est_files = sorted(Path(est_dir).glob("*.json"))
    else:
        est_files = []


    # Optional prefix-based filtering (useful when GT/SOL live in the same folder)
    gt_prefix = deep_get(cfg, "input.gt_prefix", None)
    sol_prefix = deep_get(cfg, "input.sol_prefix", None)
    est_prefix = deep_get(cfg, "input.est_input_prefix", None)

    if gt_prefix:
        gp = str(gt_prefix).lower()
        gt_files = [p for p in gt_files if p.name.lower().startswith(gp)]
    if sol_prefix:
        sp = str(sol_prefix).lower()
        sol_files = [p for p in sol_files if p.name.lower().startswith(sp)]
    if est_prefix:
        ep = str(est_prefix).lower()
        est_files = [p for p in est_files if p.name.lower().startswith(ep)]

    return gt_files, sol_files, est_files


# ============================================================
# Loading arrays and computing metrics
# ============================================================

def load_first_2d_array(npz_path: Path, preferred_keys: Tuple[str, ...]) -> Tuple[np.ndarray, str]:
    data = np.load(npz_path, allow_pickle=False)

    for k in preferred_keys:
        if k in data:
            arr = data[k]
            if isinstance(arr, np.ndarray) and arr.ndim == 2:
                return arr.astype(np.float32), k

    for k in data.files:
        arr = data[k]
        if isinstance(arr, np.ndarray) and arr.ndim == 2:
            return arr.astype(np.float32), k

    raise ValueError(f"No 2D array found in {npz_path}")


def pctiles(x: np.ndarray, ps: Tuple[int, ...] = (50, 90, 99)) -> Dict[str, float]:
    if x.size == 0:
        return {f"p{p}": float("nan") for p in ps}
    vals = np.nanpercentile(x, ps)
    return {f"p{p}": float(v) for p, v in zip(ps, vals)}


@dataclass
class GroupStats:
    n_pixels: int
    mean_signed: float
    mean_abs: float
    median_signed: float
    median_abs: float
    p90_abs: float
    p99_abs: float
    linf_abs: float


def compute_group_stats(signed_err: np.ndarray, abs_err: np.ndarray) -> GroupStats:
    # signed_err, abs_err are 1D arrays (already masked to group)
    n = int(signed_err.size)
    if n == 0:
        return GroupStats(
            n_pixels=0,
            mean_signed=float("nan"),
            mean_abs=float("nan"),
            median_signed=float("nan"),
            median_abs=float("nan"),
            p90_abs=float("nan"),
            p99_abs=float("nan"),
            linf_abs=float("nan"),
        )

    med_s = float(np.nanmedian(signed_err))
    med_a = float(np.nanmedian(abs_err))
    p = pctiles(abs_err, (90, 99))
    return GroupStats(
        n_pixels=n,
        mean_signed=float(np.nanmean(signed_err)),
        mean_abs=float(np.nanmean(abs_err)),
        median_signed=med_s,
        median_abs=med_a,
        p90_abs=p["p90"],
        p99_abs=p["p99"],
        linf_abs=float(np.nanmax(abs_err)),
    )


# ============================================================
# Coverage computation from est_input JSON
# ============================================================

def load_coverage_from_est_input(est_json_path: Path) -> Tuple[np.ndarray, int, int]:
    """
    Returns:
      c_map: (H,W) int32 where c(p)=#links that intersect pixel p
    Uses unique link counting per pixel (as requested).
    Expects:
      header: {H, W, ...}
      segments_by_link: dict[str(link)] -> list of {i,j,ds_m}
    """
    with est_json_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)

    H = int(payload["header"]["H"])
    W = int(payload["header"]["W"])
    segs = payload["segments_by_link"]

    pix_list = []
    link_list = []

    # Collect (pixel, link) pairs; then unique to avoid double-counting
    for link_str, lst in segs.items():
        li = int(link_str)
        for s in lst:
            i = int(s["i"])
            j = int(s["j"])
            p = i * W + j
            pix_list.append(p)
            link_list.append(li)

    if len(pix_list) == 0:
        return np.zeros((H, W), dtype=np.int32), H, W

    pix = np.asarray(pix_list, dtype=np.int64)
    li = np.asarray(link_list, dtype=np.int64)

    # unique pairs (pix,link)
    pairs = np.stack([pix, li], axis=1)
    pairs_u = np.unique(pairs, axis=0)

    pix_u = pairs_u[:, 0]
    # count unique links per pixel
    c_flat = np.bincount(pix_u, minlength=H * W).astype(np.int32)
    c_map = c_flat.reshape(H, W)

# ============================================================
# Distance-to-links computation (for distance-bucketed stats)
# ============================================================

def parse_distance_bins_m(edges: List[float]) -> Tuple[np.ndarray, List[str]]:
    """Given internal edges (in meters), return edges array and human-readable labels.

    Example edges [125, 375, 750] -> bins:
      (-inf,125] , (125,375] , (375,750] , (750,inf)
    We clamp to >=0 distances anyway.
    """
    e = np.asarray(edges, dtype=np.float32)
    e = e[np.isfinite(e)]
    e = np.unique(e)
    e = e[e > 0]
    labels: List[str] = []
    prev = 0.0
    if e.size == 0:
        return e, [">=0"]
    labels.append(f"≤{int(e[0])}")
    for i in range(1, len(e)):
        labels.append(f"({int(e[i-1])},{int(e[i])}]")
    labels.append(f">{int(e[-1])}")
    return e, labels


def load_links_and_grid(est_json_path: Path) -> Tuple[np.ndarray, dict]:
    """Load link segments (x0,y0,x1,y1) and header (grid geometry)."""
    with est_json_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    header = payload["header"]
    links = payload.get("links", [])
    if not links:
        return np.zeros((0, 4), dtype=np.float32), header
    segs = np.zeros((len(links), 4), dtype=np.float32)
    for i, L in enumerate(links):
        segs[i, 0] = float(L["x0_m"])
        segs[i, 1] = float(L["y0_m"])
        segs[i, 2] = float(L["x1_m"])
        segs[i, 3] = float(L["y1_m"])
    return segs, header


def pixel_centers_xy(header: dict, H: int, W: int) -> Tuple[np.ndarray, np.ndarray]:
    """Return meshgrid arrays X,Y of pixel centers in the same CRS as links (meters).

    Assumes header.frame == 'local_from_NW' and origin_rd_m has x_min, y_max.
    """
    pix = float(header.get("pixel_size_m", 125.0))
    frame = header.get("frame", "local_from_NW")
    if frame != "local_from_NW":
        raise ValueError(f"Unsupported frame '{frame}'. Expected 'local_from_NW'.")
    origin = header.get("origin_rd_m", {})
    x_min = float(origin.get("x_min"))
    y_max = float(origin.get("y_max"))
    xs = x_min + (np.arange(W, dtype=np.float32) + 0.5) * pix
    ys = y_max - (np.arange(H, dtype=np.float32) + 0.5) * pix
    X, Y = np.meshgrid(xs, ys)  # shape (H,W)
    return X, Y


def point_to_segments_distance(px: float, py: float, segs: np.ndarray, idxs: np.ndarray) -> np.ndarray:
    """Exact distance from a point to multiple line segments."""
    s = segs[idxs]  # (M,4)
    x0 = s[:, 0]; y0 = s[:, 1]; x1 = s[:, 2]; y1 = s[:, 3]
    dx = x1 - x0
    dy = y1 - y0
    denom = dx * dx + dy * dy
    # handle degenerate segments
    denom = np.where(denom == 0, 1e-12, denom)
    t = ((px - x0) * dx + (py - y0) * dy) / denom
    t = np.clip(t, 0.0, 1.0)
    cx = x0 + t * dx
    cy = y0 + t * dy
    return np.sqrt((px - cx) ** 2 + (py - cy) ** 2)


def compute_d3_distance_map(
    est_json_path: Path,
    sample_spacing_m: float = 250.0,
    k_query_samples: int = 48,
    chunk_size: int = 8000,
) -> np.ndarray:
    """Compute per-pixel distance to the 3rd-nearest link segment (meters).

    Uses KD-tree over sampled points along segments to get candidate links, then computes
    exact point-to-segment distances for candidates and takes the 3rd smallest.

    Returns:
      d3_map: (H,W) float32
    """
    if cKDTree is None:
        raise RuntimeError("Distance stats require scipy. Install scipy or disable distance.enabled.")
    segs, header = load_links_and_grid(est_json_path)
    H = int(header["H"])
    W = int(header["W"])
    if segs.shape[0] < 3:
        return np.full((H, W), np.nan, dtype=np.float32)

    # Sample points along segments
    dx = segs[:, 2] - segs[:, 0]
    dy = segs[:, 3] - segs[:, 1]
    lengths = np.sqrt(dx * dx + dy * dy)
    spacing = max(1.0, float(sample_spacing_m))

    sample_pts = []
    sample_link = []
    for li in range(segs.shape[0]):
        L = float(lengths[li])
        n = max(2, int(math.ceil(L / spacing)) + 1)
        # cap samples per link to avoid blow-ups on very long links
        n = min(n, 200)
        ts = np.linspace(0.0, 1.0, n, dtype=np.float32)
        xs = segs[li, 0] + ts * dx[li]
        ys = segs[li, 1] + ts * dy[li]
        sample_pts.append(np.stack([xs, ys], axis=1))
        sample_link.append(np.full((n,), li, dtype=np.int32))

    sample_xy = np.concatenate(sample_pts, axis=0).astype(np.float32)
    sample_to_link = np.concatenate(sample_link, axis=0)

    tree = cKDTree(sample_xy)

    X, Y = pixel_centers_xy(header, H, W)
    pts = np.stack([X.ravel(), Y.ravel()], axis=1).astype(np.float32)

    d3 = np.full((H * W,), np.nan, dtype=np.float32)

    kq = int(k_query_samples)
    kq = max(kq, 12)

    for start in range(0, pts.shape[0], chunk_size):
        end = min(pts.shape[0], start + chunk_size)
        q = pts[start:end]
        # query nearest sample points
        _, nn_idx = tree.query(q, k=kq, workers=-1)
        if nn_idx.ndim == 1:
            nn_idx = nn_idx[:, None]

        for bi in range(q.shape[0]):
            cand_links = np.unique(sample_to_link[nn_idx[bi]])
            # ensure enough candidates
            if cand_links.size < 3:
                _, nn_idx2 = tree.query(q[bi], k=min(sample_xy.shape[0], kq * 4), workers=-1)
                cand_links = np.unique(sample_to_link[np.atleast_1d(nn_idx2)])
            if cand_links.size == 0:
                continue
            # exact distances
            ds = point_to_segments_distance(float(q[bi, 0]), float(q[bi, 1]), segs, cand_links)
            if ds.size < 3:
                d3[start + bi] = float(np.nanmax(ds))
            else:
                d3[start + bi] = float(np.partition(ds, 2)[2])

    return d3.reshape(H, W)

    return c_map, H, W


# ============================================================
# Plotting with consistent scaling
# ============================================================

def save_png_2x2(
    out_png: Path,
    a: np.ndarray, b: np.ndarray, c: np.ndarray, d: np.ndarray,
    titles: Tuple[str, str, str, str],
    suptitle: str,
    cmaps: Tuple[str, str, str, str],
    vlims: Tuple[Tuple[float, float], Tuple[float, float], Tuple[float, float], Tuple[float, float]],
    dpi: int,
    show: bool = False,
):
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    axs = axes.ravel()
    arrays = [a, b, c, d]

    for ax, arr, t, cmap, (vmin, vmax) in zip(axs, arrays, titles, cmaps, vlims):
        im = ax.imshow(arr, cmap=cmap, vmin=vmin, vmax=vmax)
        ax.set_title(t)
        ax.set_xticks([])
        ax.set_yticks([])
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

    fig.suptitle(suptitle, fontsize=12)
    fig.tight_layout(rect=[0, 0.02, 1, 0.96])

    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=dpi)
    if show:
        plt.show()
    plt.close(fig)


def mask_to_nan(arr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    out = arr.astype(np.float32).copy()
    out[~mask] = np.nan
    return out


# ============================================================
# Excel writing (long/tidy)
# ============================================================

def autosize_columns(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 60)


def _write_long_tidy_sheet(wb: Workbook, title: str, headers: List[str], rows: List[dict], percent_for_masks: Tuple[str, ...] = ("rainy",)) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    metric_cols = {
        "mean_signed",
        "median_signed",
        "mean_abs",
        "median_abs",
        "p90_abs",
        "p99_abs",
        "linf_abs",
    }
    col_index = {h: i + 1 for i, h in enumerate(headers)}
    mask_col = col_index.get("mask_type", None)
    metric_col_idxs = [col_index[h] for h in headers if h in metric_cols]

    for r in rows:
        ws.append([r.get(h, None) for h in headers])
        row_i = ws.max_row

        mask_type = ws.cell(row=row_i, column=mask_col).value if mask_col else None
        if mask_type in percent_for_masks:
            fmt = "0.0%"
        else:
            fmt = "0.00"

        for ci in metric_col_idxs:
            cell = ws.cell(row=row_i, column=ci)
            if cell.value is None:
                continue
            try:
                v = float(cell.value)
            except Exception:
                continue
            if not np.isfinite(v):
                continue
            cell.number_format = fmt

    autosize_columns(ws)


def write_excel_report(xlsx_path: Path, coverage_rows: List[dict], distance_rows: List[dict]) -> None:
    """Write a workbook with CoverageStats and (optionally) DistanceStats sheets."""
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    # Coverage sheet (backwards-compatible schema)
    cov_headers = [
        "patch_key",
        "mask_type",
        "coverage_bin",
        "n_pixels",
        "mean_signed",
        "median_signed",
        "mean_abs",
        "median_abs",
        "p90_abs",
        "p99_abs",
        "linf_abs",
        "gt_file",
        "sol_file",
        "est_input_json",
        "threshold_mmph",
    ]
    _write_long_tidy_sheet(wb, "CoverageStats", cov_headers, coverage_rows, percent_for_masks=("rainy",))

    if distance_rows:
        dist_headers = [
            "patch_key",
            "mask_type",
            "distance_bin_m",
            "n_pixels",
            "mean_signed",
            "median_signed",
            "mean_abs",
            "median_abs",
            "p90_abs",
            "p99_abs",
            "linf_abs",
            "gt_file",
            "sol_file",
            "est_input_json",
            "threshold_mmph",
            "distance_def",
        ]
        _write_long_tidy_sheet(wb, "DistanceStats", dist_headers, distance_rows, percent_for_masks=("rainy",))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser(description="Batch GT vs SOL analysis with rainy/non-rainy split + coverage bins.")
    ap.add_argument("--config", type=str, required=True, help="Path to YAML/JSON config file")
    args = ap.parse_args()

    cfg = load_config_file(args.config)

    out_dir = Path(deep_get(cfg, "output.out_dir", "Analysis-Report"))
    images_subdir = deep_get(cfg, "output.images_subdir", "images")
    img_dir = out_dir / images_subdir

    xlsx_name = deep_get(cfg, "output.excel_filename", "summary.xlsx")
    xlsx_path = out_dir / xlsx_name

    show = bool(deep_get(cfg, "plots.show", False))
    dpi = int(deep_get(cfg, "plots.dpi", 150))
    cmap_gt = deep_get(cfg, "plots.cmap_gt", "viridis")
    cmap_sol = deep_get(cfg, "plots.cmap_sol", "viridis")
    cmap_diff = deep_get(cfg, "plots.cmap_diff", "seismic")
    cmap_abs = deep_get(cfg, "plots.cmap_abs_diff", "magma")
    cmap_rel = deep_get(cfg, "plots.cmap_rel", "seismic")
    cmap_abs_rel = deep_get(cfg, "plots.cmap_abs_rel", "magma")

    thr = float(deep_get(cfg, "rain.threshold_mmph", 1.0))

    gt_keys = tuple(deep_get(cfg, "data.gt_key_preference", []))
    sol_keys = tuple(deep_get(cfg, "data.sol_key_preference", ["R_hat"]))

    cov_bins_cfg = deep_get(cfg, "coverage.bins", [0, 1, 2, 3, 4, "5+"])
    exact_bins, ge_bin = parse_coverage_bins(cov_bins_cfg)

    gt_files, sol_files, est_files = list_input_files(cfg)
    print(f"GT files:  {len(gt_files)}")
    print(f"SOL files: {len(sol_files)}")
    print(f"EST jsons: {len(est_files)}")

    if not gt_files or not sol_files:
        raise SystemExit("No GT or SOL files found. Check input.gt_dir/sol_dir or globs in config.")

    gt_map = build_file_map(gt_files, "GT")
    sol_map = build_file_map(sol_files, "SOL")
    est_map = build_file_map(est_files, "EST_INPUT")

    common = sorted(set(gt_map.keys()) & set(sol_map.keys()))
    print(f"Matched GT/SOL pairs: {len(common)}")
    if len(common) == 0:
        print("Example GT keys:", list(gt_map.keys())[:5])
        print("Example SOL keys:", list(sol_map.keys())[:5])
        raise SystemExit(2)

    # Require est_input for coverage stats
    need_cov = bool(deep_get(cfg, "coverage.enabled", True))
    if need_cov:
        cov_common = [k for k in common if k in est_map]
        missing = [k for k in common if k not in est_map]
        print(f"Pairs with est_input for coverage: {len(cov_common)}")
        if missing:
            print(f"[warn] Missing est_input for {len(missing)} pairs. Example keys:", missing[:5])
        # We'll compute coverage stats only for keys that have est_input.
    else:
        cov_common = common

    # --------------------------
    # PASS 1: compute global scales
    # --------------------------
    Rmax_global = 0.0
    Dmax_global = 0.0
    RelMax_global = 0.0
    RelAbsMax_global = 0.0

    for k in common:
        gt, _ = load_first_2d_array(gt_map[k], gt_keys)
        sol, _ = load_first_2d_array(sol_map[k], sol_keys)

        if gt.shape != sol.shape:
            raise ValueError(f"Shape mismatch for {k}: GT {gt.shape} vs SOL {sol.shape}")

        diff = sol - gt
        Rmax_global = max(Rmax_global, float(np.nanmax(gt)), float(np.nanmax(sol)))
        Dmax_global = max(Dmax_global, float(np.nanmax(np.abs(diff))))

        rainy = gt >= thr
        if np.any(rainy):
            # relative errors only on rainy pixels
            rel = np.zeros_like(gt, dtype=np.float32)
            abs_rel = np.zeros_like(gt, dtype=np.float32)
            rel[rainy] = (gt[rainy] - sol[rainy]) / gt[rainy]
            abs_rel[rainy] = np.abs(gt[rainy] - sol[rainy]) / gt[rainy]
            RelMax_global = max(RelMax_global, float(np.nanmax(np.abs(rel[rainy]))))
            RelAbsMax_global = max(RelAbsMax_global, float(np.nanmax(abs_rel[rainy])))

    if Rmax_global <= 0:
        Rmax_global = 1.0
    if Dmax_global <= 0:
        Dmax_global = 1.0
    if RelMax_global <= 0:
        RelMax_global = 1.0
    if RelAbsMax_global <= 0:
        RelAbsMax_global = 1.0

    print(f"Global scales: Rmax={Rmax_global:.6g}, Dmax={Dmax_global:.6g}, RelMax={RelMax_global:.6g}, RelAbsMax={RelAbsMax_global:.6g}")

    # --------------------------
    # PASS 2: render PNGs + compute excel rows (long/tidy)
    # --------------------------
    coverage_rows: List[dict] = []
    distance_rows: List[dict] = []

    # Distance-bucketed stats (optional)
    dist_enabled = bool(deep_get(cfg, "distance.enabled", True))
    dist_edges = deep_get(cfg, "distance.bins_m", [125.0, 375.0, 750.0, 1500.0, 3125.0])
    dist_edges_arr, dist_labels = parse_distance_bins_m([float(x) for x in dist_edges])
    dist_sample_spacing = float(deep_get(cfg, "distance.sample_spacing_m", 250.0))
    dist_kq = int(deep_get(cfg, "distance.k_query_samples", 48))
    dist_chunk = int(deep_get(cfg, "distance.chunk_size", 8000))

    for k in common:
        gt_path = gt_map[k]
        sol_path = sol_map[k]
        est_path = est_map.get(k, None)

        gt, gt_key = load_first_2d_array(gt_path, gt_keys)
        sol, sol_key = load_first_2d_array(sol_path, sol_keys)
        diff = sol - gt
        abs_diff = np.abs(diff)

        rainy = gt >= thr
        nonrainy = ~rainy

        # Compute relative errors on rainy pixels
        rel = np.full_like(gt, np.nan, dtype=np.float32)
        abs_rel = np.full_like(gt, np.nan, dtype=np.float32)
        if np.any(rainy):
            rel[rainy] = (gt[rainy] - sol[rainy]) / gt[rainy]
            abs_rel[rainy] = np.abs(gt[rainy] - sol[rainy]) / gt[rainy]

        # ---------- PNGs ----------
        # Rainy PNG: (GT, SOL, rel, abs_rel) masked to rainy pixels
        rainy_png = img_dir / f"{k}_rainy.png"
        save_png_2x2(
            rainy_png,
            mask_to_nan(gt, rainy),
            mask_to_nan(sol, rainy),
            mask_to_nan(rel, rainy),
            mask_to_nan(abs_rel, rainy),
            titles=("GT (rainy)", "SOL (rainy)", "(GT-SOL)/GT", "|GT-SOL|/GT"),
            suptitle=f"{k} | rainy: GT>= {thr} mm/h",
            cmaps=(cmap_gt, cmap_sol, cmap_rel, cmap_abs_rel),
            vlims=((0, Rmax_global), (0, Rmax_global), (-RelMax_global, RelMax_global), (0, RelAbsMax_global)),
            dpi=dpi,
            show=show,
        )

        # Non-rainy PNG: (GT, SOL, diff, abs_diff) masked to nonrainy pixels
        non_png = img_dir / f"{k}_nonrainy.png"
        save_png_2x2(
            non_png,
            mask_to_nan(gt, nonrainy),
            mask_to_nan(sol, nonrainy),
            mask_to_nan(diff, nonrainy),
            mask_to_nan(abs_diff, nonrainy),
            titles=("GT (non-rainy)", "SOL (non-rainy)", "SOL-GT", "|SOL-GT|"),
            suptitle=f"{k} | non-rainy: GT< {thr} mm/h",
            cmaps=(cmap_gt, cmap_sol, cmap_diff, cmap_abs),
            # IMPORTANT: ABS_DIFF uses same scale as GT/SOL (your requirement)
            vlims=((0, Rmax_global), (0, Rmax_global), (-Dmax_global, Dmax_global), (0, Rmax_global)),
            dpi=dpi,
            show=show,
        )

        # ---------- Coverage map ----------
        if need_cov and est_path is not None:
            c_map, Hc, Wc = load_coverage_from_est_input(est_path)
            if c_map.shape != gt.shape:
                raise ValueError(f"Coverage shape mismatch for {k}: coverage {c_map.shape} vs GT {gt.shape}")
        else:
            # If no coverage, treat all as one bin "ALL"
            c_map = np.zeros_like(gt, dtype=np.int32)

        # ---------- Long/tidy stats by (mask_type, coverage_bin) ----------
        # Build per-pixel coverage bins for this patch
        if need_cov and est_path is not None:
            # bins for exact and ge_bin
            all_cov_values = np.unique(c_map)
            labels_present = set()
            for cov_v in all_cov_values.tolist():
                lab = coverage_bin_label(int(cov_v), exact_bins, ge_bin)
                if lab is not None:
                    labels_present.add(lab)
            # also ensure bins exist even if empty (optional)
            labels = [str(b) for b in exact_bins]
            if ge_bin >= 0:
                labels.append(f"{ge_bin}+")
        else:
            labels = ["ALL"]

        def iter_groups(mask: np.ndarray):
            if labels == ["ALL"]:
                yield "ALL", mask
                return
            for lab in labels:
                if lab.endswith("+"):
                    base = int(lab[:-1])
                    gmask = mask & (c_map >= base)
                else:
                    base = int(lab)
                    gmask = mask & (c_map == base)
                yield lab, gmask

        # Rainy stats: signed=rel, abs=abs_rel
        for lab, gmask in iter_groups(rainy):
            signed_vals = rel[gmask]
            abs_vals = abs_rel[gmask]
            # drop nans
            signed_vals = signed_vals[np.isfinite(signed_vals)]
            abs_vals = abs_vals[np.isfinite(abs_vals)]
            st = compute_group_stats(signed_vals, abs_vals)

            coverage_rows.append({
                "patch_key": k,
                "mask_type": "rainy",
                "coverage_bin": lab,
                "n_pixels": st.n_pixels,
                "mean_signed": st.mean_signed,
                "median_signed": st.median_signed,
                "mean_abs": st.mean_abs,
                "median_abs": st.median_abs,
                "p90_abs": st.p90_abs,
                "p99_abs": st.p99_abs,
                "linf_abs": st.linf_abs,
                "gt_file": gt_path.name,
                "sol_file": sol_path.name,
                "est_input_json": est_path.name if est_path else "",
                "threshold_mmph": thr,
            })

        # Non-rainy stats: signed=diff, abs=|diff|
        for lab, gmask in iter_groups(nonrainy):
            signed_vals = diff[gmask].astype(np.float64).ravel()
            abs_vals = abs_diff[gmask].astype(np.float64).ravel()
            # finite (should be)
            signed_vals = signed_vals[np.isfinite(signed_vals)]
            abs_vals = abs_vals[np.isfinite(abs_vals)]
            st = compute_group_stats(signed_vals, abs_vals)

            coverage_rows.append({
                "patch_key": k,
                "mask_type": "nonrainy",
                "coverage_bin": lab,
                "n_pixels": st.n_pixels,
                "mean_signed": st.mean_signed,
                "median_signed": st.median_signed,
                "mean_abs": st.mean_abs,
                "median_abs": st.median_abs,
                "p90_abs": st.p90_abs,
                "p99_abs": st.p99_abs,
                "linf_abs": st.linf_abs,
                "gt_file": gt_path.name,
                "sol_file": sol_path.name,
                "est_input_json": est_path.name if est_path else "",
                "threshold_mmph": thr,
            })

        # ---------- Distance-to-3rd-nearest-link stats (Excel) ----------
        # Uses per-pixel D(x) = distance to the 3rd-nearest link segment (meters),
        # then buckets pixels by D(x).
        if dist_enabled and est_path is not None:
            d3_map = compute_d3_distance_map(
                est_path,
                sample_spacing_m=dist_sample_spacing,
                k_query_samples=dist_kq,
                chunk_size=dist_chunk,
            )
            if d3_map.shape != gt.shape:
                raise ValueError(f"d3 distance map shape mismatch for {k}: d3 {d3_map.shape} vs GT {gt.shape}")

            def dist_bin_mask(idx: int) -> np.ndarray:
                # idx in [0..len(labels)-1]
                if dist_edges_arr.size == 0:
                    return np.isfinite(d3_map)
                if idx == 0:
                    return d3_map <= dist_edges_arr[0]
                if idx == len(dist_labels) - 1:
                    return d3_map > dist_edges_arr[-1]
                lo = dist_edges_arr[idx - 1]
                hi = dist_edges_arr[idx]
                return (d3_map > lo) & (d3_map <= hi)

            # Rainy: relative errors
            for bi, lab in enumerate(dist_labels):
                gmask = rainy & dist_bin_mask(bi)
                signed_vals = rel[gmask]
                abs_vals = abs_rel[gmask]
                signed_vals = signed_vals[np.isfinite(signed_vals)]
                abs_vals = abs_vals[np.isfinite(abs_vals)]
                st = compute_group_stats(signed_vals, abs_vals)
                distance_rows.append({
                    "patch_key": k,
                    "mask_type": "rainy",
                    "distance_bin_m": lab,
                    "n_pixels": st.n_pixels,
                    "mean_signed": st.mean_signed,
                    "median_signed": st.median_signed,
                    "mean_abs": st.mean_abs,
                    "median_abs": st.median_abs,
                    "p90_abs": st.p90_abs,
                    "p99_abs": st.p99_abs,
                    "linf_abs": st.linf_abs,
                    "gt_file": gt_path.name,
                    "sol_file": sol_path.name,
                    "est_input_json": est_path.name if est_path else "",
                    "threshold_mmph": thr,
                    "distance_def": "d3(point-to-segment)",
                })

            # Non-rainy: additive errors
            for bi, lab in enumerate(dist_labels):
                gmask = nonrainy & dist_bin_mask(bi)
                signed_vals = diff[gmask].astype(np.float64).ravel()
                abs_vals = abs_diff[gmask].astype(np.float64).ravel()
                signed_vals = signed_vals[np.isfinite(signed_vals)]
                abs_vals = abs_vals[np.isfinite(abs_vals)]
                st = compute_group_stats(signed_vals, abs_vals)
                distance_rows.append({
                    "patch_key": k,
                    "mask_type": "nonrainy",
                    "distance_bin_m": lab,
                    "n_pixels": st.n_pixels,
                    "mean_signed": st.mean_signed,
                    "median_signed": st.median_signed,
                    "mean_abs": st.mean_abs,
                    "median_abs": st.median_abs,
                    "p90_abs": st.p90_abs,
                    "p99_abs": st.p99_abs,
                    "linf_abs": st.linf_abs,
                    "gt_file": gt_path.name,
                    "sol_file": sol_path.name,
                    "est_input_json": est_path.name if est_path else "",
                    "threshold_mmph": thr,
                    "distance_def": "d3(point-to-segment)",
                })


    # Write Excel
    write_excel_report(xlsx_path, coverage_rows, distance_rows)
    print(f"Wrote Excel: {xlsx_path}")
    print(f"Wrote PNGs to: {img_dir}")


if __name__ == "__main__":
    main()
